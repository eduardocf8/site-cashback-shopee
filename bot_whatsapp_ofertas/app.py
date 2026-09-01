import csv
import os
import shutil
import re
import sys
import threading
from datetime import datetime, timedelta
from html import escape
from pathlib import Path

from PySide6.QtCore import QDate, QEvent, QObject, Qt, QTimer, Signal, QtMsgType, qInstallMessageHandler
from PySide6.QtGui import QFont, QIcon, QStandardItem, QStandardItemModel, QTextCursor
from PySide6.QtWidgets import (
    QApplication,
    QAbstractItemView,
    QCheckBox,
    QComboBox,
    QDateEdit,
    QFrame,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMenu,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QStyle,
    QStyledItemDelegate,
    QStyleOptionViewItem,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)
from playwright.sync_api import sync_playwright

from afiliados import ConversorAfiliados
from browser_launch import (
    montar_kwargs_launch,
    aplicar_anti_deteccao_no_contexto,
)
from bot_runner import BotRunner
from categorias_shopee import opcoes_categoria
from database import Database
from ia_revisor import RevisorIA
from relatorio_email import MESES_PT, montar_corpo_email, enviar_email
from settings import (
    APP_DIR,
    CONFIG_BACKUP_DIR,
    RESOURCE_DIR,
    AppSettings,
    backup_user_config,
)
from whatsapp import WhatsApp


TESTE_API_SHOPEE_URL = (
    "https://shopee.com.br/Apple-Iphone-11-128GB-Local-Set-i.52377417.6309028319"
)
APP_VERSION = "1.0.0"
APP_ICON_PATH = RESOURCE_DIR / "assets" / "app_icon.ico"
LOG_RETENTION_DAYS = 30
THEME_OPTIONS = {
    "padrao": "Padrão",
    "claro": "Claro",
    "noturno": "Escuro",
}


class BotSignals(QObject):
    log = Signal(str)
    status = Signal(str)
    stats = Signal(dict)
    stopped = Signal()
    test_finished = Signal()
    indicators_loaded = Signal(dict)
    conversoes_loaded = Signal(dict)
    conversoes_button_enable = Signal()
    shopee_offers_status = Signal(str)
    relatorio_email_status = Signal(str)
    relatorio_email_button_enable = Signal()
    shopee_login_button_enable = Signal()


class ThemeSelector(QPushButton):
    currentIndexChanged = Signal(int)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._items = []
        self._current_index = -1
        self._menu = QMenu(self)
        self.setCursor(Qt.PointingHandCursor)
        self.clicked.connect(self.open_menu)

    def addItem(self, label, data):
        index = len(self._items)
        self._items.append((label, data))
        action = self._menu.addAction(label)
        action.triggered.connect(lambda checked=False, i=index: self.setCurrentIndex(i))
        if self._current_index == -1:
            self.setCurrentIndex(0, emit=False)

    def findData(self, data):
        for index, (_, item_data) in enumerate(self._items):
            if item_data == data:
                return index
        return -1

    def currentData(self):
        if 0 <= self._current_index < len(self._items):
            return self._items[self._current_index][1]
        return None

    def setCurrentIndex(self, index, emit=True):
        if not 0 <= index < len(self._items):
            return
        self._current_index = index
        self.setText(self._items[index][0])
        if emit:
            self.currentIndexChanged.emit(index)

    def open_menu(self):
        self._menu.exec(self.mapToGlobal(self.rect().bottomLeft()))


class SecondsInput(QWidget):
    def __init__(self, minimum=0, maximum=120, value=1, parent=None):
        super().__init__(parent)
        self.minimum = minimum
        self.maximum = maximum

        layout = QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        self.minus_button = QPushButton("-")
        self.minus_button.setObjectName("stepButton")
        self.minus_button.setFixedSize(38, 38)
        self.minus_button.clicked.connect(lambda: self.step(-1))

        self.input = QLineEdit()
        self.input.setObjectName("secondsInput")
        self.input.setFixedWidth(92)
        self.input.setMinimumHeight(38)
        self.input.setAlignment(Qt.AlignCenter)
        self.input.editingFinished.connect(self.normalize)

        self.plus_button = QPushButton("+")
        self.plus_button.setObjectName("stepButton")
        self.plus_button.setFixedSize(38, 38)
        self.plus_button.clicked.connect(lambda: self.step(1))

        layout.addWidget(self.minus_button)
        layout.addWidget(self.input)
        layout.addWidget(self.plus_button)
        self.setFixedWidth(184)
        self.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Fixed)

        self.setValue(value)

    def value(self):
        text = self.input.text().replace("s", "").strip()
        try:
            value = int(text)
        except ValueError:
            value = self.minimum
        return max(self.minimum, min(self.maximum, value))

    def setValue(self, value):
        value = max(self.minimum, min(self.maximum, int(value)))
        self.input.setText(f"{value} s")

    def step(self, delta):
        self.setValue(self.value() + delta)

    def normalize(self):
        self.setValue(self.value())

    def set_invalid(self, invalid):
        self.input.setProperty("invalid", invalid)
        self.input.style().unpolish(self.input)
        self.input.style().polish(self.input)


class NoFocusTableDelegate(QStyledItemDelegate):
    def initStyleOption(self, option, index):
        super().initStyleOption(option, index)
        option.state &= ~QStyle.State_HasFocus


class NumericTableItem(QTableWidgetItem):
    """
    Item de tabela cujo TEXTO exibido pode ser formatado (ex.: "R$ 11,20"),
    mas cuja ORDENAÇÃO usa o valor numérico real guardado. Sem isto, o Qt
    ordena pelo texto (alfabético), fazendo "11,20" vir antes de "1,93".
    """

    def __init__(self, valor_numerico, texto):
        super().__init__(texto)
        try:
            self._valor = float(valor_numerico)
        except (TypeError, ValueError):
            self._valor = 0.0

    def __lt__(self, other):
        if isinstance(other, NumericTableItem):
            return self._valor < other._valor
        return super().__lt__(other)


class ComboBoxMultiplaSelecao(QComboBox):
    """
    Lista suspensa com itens marcáveis (checkbox), permitindo escolher uma
    ou várias opções ao mesmo tempo sem fechar o menu a cada clique.
    """

    selecaoAlterada = Signal()

    def __init__(self, parent=None, texto_vazio="Todas as categorias"):
        super().__init__(parent)
        self._texto_vazio = texto_vazio
        self.setEditable(True)
        self.lineEdit().setReadOnly(True)
        self.lineEdit().setPlaceholderText(texto_vazio)
        self.lineEdit().installEventFilter(self)
        self.setInsertPolicy(QComboBox.NoInsert)
        self._modelo = QStandardItemModel(self)
        self.setModel(self._modelo)
        self.view().viewport().installEventFilter(self)
        self._atualizar_texto_exibido()

    def adicionar_opcao(self, texto, dado):
        item = QStandardItem(texto)
        item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled | Qt.ItemIsSelectable)
        item.setData(Qt.Unchecked, Qt.CheckStateRole)
        item.setData(dado, Qt.UserRole)
        self._modelo.appendRow(item)

    def eventFilter(self, origem, evento):
        if origem is self.lineEdit() and evento.type() == QEvent.MouseButtonPress:
            if self.view().isVisible():
                self.hidePopup()
            else:
                self.showPopup()
            return True

        if origem is self.view().viewport() and evento.type() == QEvent.MouseButtonRelease:
            indice = self.view().indexAt(evento.pos())
            if indice.isValid():
                item = self._modelo.itemFromIndex(indice)
                if item is not None:
                    novo_estado = Qt.Unchecked if item.checkState() == Qt.Checked else Qt.Checked
                    item.setCheckState(novo_estado)
                    self._atualizar_texto_exibido()
                    self.selecaoAlterada.emit()
                return True
        return super().eventFilter(origem, evento)

    def _atualizar_texto_exibido(self):
        textos = self.textos_selecionados()
        self.lineEdit().setText(", ".join(textos))
        self.lineEdit().setPlaceholderText(self._texto_vazio)

    def textos_selecionados(self):
        textos = []
        for linha in range(self._modelo.rowCount()):
            item = self._modelo.item(linha)
            if item.checkState() == Qt.Checked:
                textos.append(item.text())
        return textos

    def dados_selecionados(self):
        dados = []
        for linha in range(self._modelo.rowCount()):
            item = self._modelo.item(linha)
            if item.checkState() == Qt.Checked:
                dados.append(item.data(Qt.UserRole))
        return dados

    def definir_selecionados(self, valores):
        valores = set(valores or [])
        for linha in range(self._modelo.rowCount()):
            item = self._modelo.item(linha)
            estado = Qt.Checked if item.data(Qt.UserRole) in valores else Qt.Unchecked
            item.setCheckState(estado)
        self._atualizar_texto_exibido()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"Bot.ee v{APP_VERSION}")
        if APP_ICON_PATH.exists():
            self.setWindowIcon(QIcon(str(APP_ICON_PATH)))
        self.resize(1180, 760)

        self.settings = AppSettings.load()
        self.logs_dir = APP_DIR / "logs"
        self.log_file_path = self.logs_dir / f"bot_{datetime.now().strftime('%Y-%m-%d')}.log"
        self.cleanup_old_logs()
        self.runner = None
        self.log_lines = []
        self.filtered_log_lines = []
        self.history_rows = []
        self.filtered_history_rows = []
        self.indicators_rows = []
        self.signals = BotSignals()
        self.signals.log.connect(self.add_log)
        self.signals.status.connect(self.set_status)
        self.signals.stats.connect(self.update_stats)
        self.signals.test_finished.connect(self.finish_test)
        self.signals.indicators_loaded.connect(self.render_indicators_report)
        self.signals.conversoes_loaded.connect(self.render_conversoes_report)
        self.signals.conversoes_button_enable.connect(
            lambda: self.refresh_conversoes_button.setEnabled(True)
        )
        self.signals.shopee_offers_status.connect(self.update_shopee_offers_status)
        self.history_timer = QTimer(self)
        self.history_timer.setInterval(30000)
        self.history_timer.timeout.connect(self.refresh_history)
        self.relatorio_email_timer = QTimer(self)
        self.relatorio_email_timer.setInterval(60000)
        self.relatorio_email_timer.timeout.connect(self.verificar_envio_relatorio_agendado)

        self.build_ui()
        self.signals.relatorio_email_status.connect(self.relatorio_email_status_label.setText)
        self.signals.relatorio_email_button_enable.connect(
            lambda: self.relatorio_email_testar_button.setEnabled(True)
        )
        self.signals.shopee_login_button_enable.connect(
            lambda: self.shopee_login_button.setEnabled(True)
        )
        self.load_fields()
        self.apply_styles()
        self.refresh_history()
        self.relatorio_email_timer.start()

    def build_ui(self):
        root = QWidget()
        outer = QHBoxLayout(root)
        outer.setContentsMargins(22, 22, 22, 22)
        outer.setSpacing(18)

        left = QFrame()
        left.setObjectName("sidebar")
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(22, 22, 22, 22)
        left_layout.setSpacing(14)

        title = QLabel("Bot.ee")
        title.setObjectName("appTitle")
        subtitle = QLabel("Ofertas com link afiliado via API")
        subtitle.setObjectName("appSubtitle")
        version_label = QLabel(f"Versão {APP_VERSION}")
        version_label.setObjectName("appVersion")

        self.status_label = QLabel("Pronto para iniciar")
        self.status_label.setObjectName("statusPill")
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setWordWrap(True)

        self.processadas_label = QLabel("0")
        self.enviadas_label = QLabel("0")
        self.erros_label = QLabel("0")

        stats_grid = QGridLayout()
        stats_grid.setHorizontalSpacing(10)
        stats_grid.setVerticalSpacing(10)
        stats_grid.addWidget(self.metric_card("Processadas", self.processadas_label), 0, 0)
        stats_grid.addWidget(self.metric_card("Enviadas", self.enviadas_label), 0, 1)
        stats_grid.addWidget(self.metric_card("Erros", self.erros_label), 1, 0, 1, 2)

        self.start_button = QPushButton("Iniciar bot")
        self.start_button.setObjectName("primaryButton")
        self.start_button.clicked.connect(self.start_bot)

        self.stop_button = QPushButton("Parar bot")
        self.stop_button.setObjectName("secondaryButton")
        self.stop_button.setEnabled(False)
        self.stop_button.clicked.connect(self.stop_bot)

        self.save_button = QPushButton("Salvar configurações")
        self.save_button.setObjectName("ghostButton")
        self.save_button.clicked.connect(lambda: self.save_fields())
        self.about_button = QPushButton("Sobre")
        self.about_button.setObjectName("ghostButton")
        self.about_button.clicked.connect(self.show_about)

        self.tema_visual_label = QLabel("Layout")
        self.tema_visual_label.setObjectName("sidebarFieldLabel")
        self.tema_visual_label.setAlignment(Qt.AlignCenter)
        self.tema_visual_input = ThemeSelector()
        self.tema_visual_input.setObjectName("themeSelect")
        for theme_key, theme_label in THEME_OPTIONS.items():
            self.tema_visual_input.addItem(theme_label, theme_key)
        self.tema_visual_input.currentIndexChanged.connect(self.change_visual_theme)

        self.theme_row = QFrame()
        self.theme_row.setObjectName("themeRow")
        theme_row_layout = QHBoxLayout(self.theme_row)
        theme_row_layout.setContentsMargins(0, 0, 0, 0)
        theme_row_layout.setSpacing(10)
        theme_row_layout.setAlignment(Qt.AlignVCenter)
        self.tema_visual_label.setFixedWidth(54)
        self.tema_visual_label.setMinimumHeight(44)
        self.tema_visual_input.setMinimumHeight(44)
        theme_row_layout.addWidget(self.tema_visual_label)
        theme_row_layout.addWidget(self.tema_visual_input, 1)

        left_layout.addWidget(title)
        left_layout.addWidget(subtitle)
        left_layout.addWidget(version_label)
        left_layout.addSpacing(10)
        left_layout.addWidget(self.status_label)
        left_layout.addSpacing(12)
        left_layout.addLayout(stats_grid)
        left_layout.addStretch(1)
        left_layout.addWidget(self.theme_row)
        left_layout.addStretch(1)
        left_layout.addWidget(self.start_button)
        left_layout.addWidget(self.stop_button)
        left_layout.addWidget(self.save_button)
        left_layout.addWidget(self.about_button)

        right = QFrame()
        right.setObjectName("content")
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(22, 22, 22, 22)
        right_layout.setSpacing(16)

        config_title = QLabel("Configurações")
        config_title.setObjectName("sectionTitle")

        self.grupo_origem_input = QLineEdit()
        self.grupo_origem_input.setMinimumHeight(44)
        self.grupo_origem_input.setPlaceholderText("Nome exato do grupo de origem")
        self.destinos_input = QLineEdit()
        self.destinos_input.setMinimumHeight(44)
        self.destinos_input.setPlaceholderText("Exemplo: Ofertas @vegrassi; Ofertas shopee")
        self.canais_destino_input = QLineEdit()
        self.canais_destino_input.setMinimumHeight(44)
        self.canais_destino_input.setPlaceholderText("Opcional: nome exato do(s) canal(is) do WhatsApp")
        self.app_id_input = QLineEdit()
        self.app_id_input.setMinimumHeight(44)
        self.app_id_input.setPlaceholderText("AppID da API Shopee")
        self.secret_input = QLineEdit()
        self.secret_input.setMinimumHeight(44)
        self.secret_input.setEchoMode(QLineEdit.Password)
        self.secret_input.setPlaceholderText("Secret da API Shopee")
        self.sub_ids_input = QLineEdit()
        self.sub_ids_input.setMinimumHeight(44)
        self.sub_ids_input.setPlaceholderText("Opcional: whatsapp, olapromos, bot01")
        self.intervalo_input = SecondsInput(1, 60, 1)
        self.timeout_previa_input = SecondsInput(1, 120, 30)
        self.espera_minima_input = SecondsInput(0, 60, 6)
        self.processar_historico_input = QCheckBox("Processar histórico visível ao iniciar")
        self.ignorar_links_enviados_input = QCheckBox("Ignorar links já enviados anteriormente")
        self.aguardar_previa_input = QCheckBox("Aguardar imagem da prévia antes de enviar")
        self.modo_simulacao_input = QCheckBox("Modo simulação: não enviar ofertas")
        self.headless_input = QCheckBox("Rodar navegador minimizado")
        self.agendamento_ativo_input = QCheckBox("Usar agendamento de funcionamento")
        self.horario_inicio_input = QLineEdit()
        self.horario_inicio_input.setMinimumHeight(32)
        self.horario_inicio_input.setMaximumWidth(100)
        self.horario_inicio_input.setAlignment(Qt.AlignCenter)
        self.horario_inicio_input.setPlaceholderText("07:00")
        self.horario_fim_input = QLineEdit()
        self.horario_fim_input.setMinimumHeight(32)
        self.horario_fim_input.setMaximumWidth(100)
        self.horario_fim_input.setAlignment(Qt.AlignCenter)
        self.horario_fim_input.setPlaceholderText("23:00")
        self.recuperacao_automatica_input = QCheckBox("Recuperar automaticamente em falhas")
        self.recuperacao_max_tentativas_input = QLineEdit()
        self.recuperacao_max_tentativas_input.setMinimumHeight(32)
        self.recuperacao_max_tentativas_input.setMaximumWidth(90)
        self.recuperacao_max_tentativas_input.setAlignment(Qt.AlignCenter)
        self.recuperacao_max_tentativas_input.setPlaceholderText("5")
        self.recuperacao_intervalo_input = SecondsInput(5, 300, 30)
        self.ia_revisao_input = QCheckBox("Usar revisão por IA nas ofertas")
        self.ia_pausar_limite_input = QCheckBox("Pausar IA automaticamente ao atingir limite")
        self.ia_retomar_dia_input = QCheckBox("Retomar IA no dia seguinte")
        self.ia_api_key_input = QLineEdit()
        self.ia_api_key_input.setMinimumHeight(44)
        self.ia_api_key_input.setEchoMode(QLineEdit.Password)
        self.ia_api_key_input.setPlaceholderText("Chave da API Gemini")
        self.ia_modelo_input = QComboBox()
        self.ia_modelo_input.setMinimumHeight(44)
        self.ia_modelo_input.addItem("Gemini Flash Lite", "gemini-3.1-flash-lite")
        self.ia_modelo_input.addItem("Gemini Flash", "gemini-3.5-flash")
        self.ia_tipo_revisao_input = QComboBox()
        self.ia_tipo_revisao_input.setMinimumHeight(44)
        self.ia_tipo_revisao_input.addItem("Revisão leve", "leve")
        self.ia_tipo_revisao_input.addItem("Revisão contextual", "contextual")
        self.ia_tipo_revisao_input.addItem("Revisão criativa", "criativa")
        self.ia_limite_diario_input = QLineEdit()
        self.ia_limite_diario_input.setMinimumHeight(44)
        self.ia_limite_diario_input.setAlignment(Qt.AlignCenter)
        self.ia_limite_diario_input.setPlaceholderText("180")
        self.ia_timeout_input = SecondsInput(3, 60, 10)
        self.ia_status_box = QTextEdit()
        self.ia_status_box.setObjectName("aiStatusBox")
        self.ia_status_box.setReadOnly(True)
        self.ia_status_box.setMinimumHeight(150)
        self.ia_status_box.setPlaceholderText("As mensagens da revisão por IA aparecerão aqui.")
        self.modo_origem_ofertas_input = QComboBox()
        self.modo_origem_ofertas_input.setMinimumHeight(44)
        self.modo_origem_ofertas_input.addItem("Grupo do WhatsApp", "whatsapp")
        self.modo_origem_ofertas_input.addItem("Buscar ofertas na Shopee", "shopee")
        self.shopee_ofertas_tipo_input = QComboBox()
        self.shopee_ofertas_tipo_input.setMinimumHeight(44)
        self.shopee_ofertas_tipo_input.addItem("Ofertas gerais", "geral")
        self.shopee_ofertas_tipo_input.addItem("Por palavra-chave/nicho", "palavra_chave")
        self.shopee_ofertas_tipo_input.addItem("Por loja ou marca", "loja")
        self.shopee_ofertas_tipo_input.addItem("Por categoria", "categoria")
        self.shopee_ofertas_tipo_input.currentIndexChanged.connect(self.atualizar_campos_tipo_busca_shopee)
        self.shopee_ofertas_colecao_input = QLineEdit()
        self.shopee_ofertas_colecao_input.setMinimumHeight(44)
        self.shopee_ofertas_colecao_input.setPlaceholderText("Opcional: link da coleção Shopee")
        self.shopee_ofertas_cupom_input = QLineEdit()
        self.shopee_ofertas_cupom_input.setMinimumHeight(44)
        self.shopee_ofertas_cupom_input.setPlaceholderText("Opcional: cupom de desconto")
        self.shopee_ofertas_palavra_input = QLineEdit()
        self.shopee_ofertas_palavra_input.setMinimumHeight(44)
        self.shopee_ofertas_palavra_input.setPlaceholderText("Exemplo: casa, beleza, eletrônico")
        self.shopee_ofertas_shop_id_input = QLineEdit()
        self.shopee_ofertas_shop_id_input.setMinimumHeight(44)
        self.shopee_ofertas_shop_id_input.setPlaceholderText("Opcional: ID da loja/marca")
        self.shopee_ofertas_categoria_input = ComboBoxMultiplaSelecao(
            texto_vazio="Todas as categorias"
        )
        self.shopee_ofertas_categoria_input.setMinimumHeight(44)
        for chave, label in opcoes_categoria():
            self.shopee_ofertas_categoria_input.adicionar_opcao(label, chave)
        self.shopee_ofertas_ordenacao_input = QComboBox()
        self.shopee_ofertas_ordenacao_input.setMinimumHeight(44)
        self.shopee_ofertas_ordenacao_input.addItem("Mais vendidos", "mais_vendidos")
        self.shopee_ofertas_ordenacao_input.addItem("Maior comissão", "maior_comissao")
        self.shopee_ofertas_ordenacao_input.addItem("Maior desconto", "maior_desconto")
        self.shopee_ofertas_ordenacao_input.addItem("Mais recentes", "recentes")
        self.shopee_ofertas_comissao_input = QLineEdit()
        self.shopee_ofertas_comissao_input.setMinimumHeight(44)
        self.shopee_ofertas_comissao_input.setAlignment(Qt.AlignCenter)
        self.shopee_ofertas_comissao_input.setPlaceholderText("0")
        self.shopee_ofertas_desconto_input = QLineEdit()
        self.shopee_ofertas_desconto_input.setMinimumHeight(44)
        self.shopee_ofertas_desconto_input.setAlignment(Qt.AlignCenter)
        self.shopee_ofertas_desconto_input.setPlaceholderText("0")
        self.shopee_ofertas_preco_min_input = QLineEdit()
        self.shopee_ofertas_preco_min_input.setMinimumHeight(44)
        self.shopee_ofertas_preco_min_input.setAlignment(Qt.AlignCenter)
        self.shopee_ofertas_preco_min_input.setPlaceholderText("Opcional")
        self.shopee_ofertas_preco_max_input = QLineEdit()
        self.shopee_ofertas_preco_max_input.setMinimumHeight(44)
        self.shopee_ofertas_preco_max_input.setAlignment(Qt.AlignCenter)
        self.shopee_ofertas_preco_max_input.setPlaceholderText("Opcional")
        self.shopee_ofertas_vendas_input = QLineEdit()
        self.shopee_ofertas_vendas_input.setMinimumHeight(44)
        self.shopee_ofertas_vendas_input.setAlignment(Qt.AlignCenter)
        self.shopee_ofertas_vendas_input.setPlaceholderText("0")
        self.shopee_ofertas_avaliacao_input = QLineEdit()
        self.shopee_ofertas_avaliacao_input.setMinimumHeight(44)
        self.shopee_ofertas_avaliacao_input.setAlignment(Qt.AlignCenter)
        self.shopee_ofertas_avaliacao_input.setPlaceholderText("Exemplo: 4.5")
        self.shopee_ofertas_intervalo_input = SecondsInput(60, 3600, 300)
        self.shopee_ofertas_nao_repetir_input = QCheckBox("Não repetir produtos já enviados")
        self.shopee_ofertas_status_box = QTextEdit()
        self.shopee_ofertas_status_box.setObjectName("aiStatusBox")
        self.shopee_ofertas_status_box.setReadOnly(True)
        self.shopee_ofertas_status_box.setMinimumHeight(96)
        self.shopee_ofertas_status_box.setPlainText(
            "Etapa 1 concluída: estas opções serão usadas na busca automática de ofertas da Shopee nas próximas etapas."
        )
        self.test_ai_button = QPushButton("Testar IA Gemini")
        self.test_ai_button.setObjectName("toolButton")
        self.test_ai_button.setMinimumHeight(42)
        self.test_ai_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.test_ai_button.clicked.connect(self.test_ai_gemini)
        self.test_api_button = QPushButton("Testar API Shopee")
        self.test_api_button.setObjectName("toolButton")
        self.test_api_button.setMinimumHeight(44)
        self.test_api_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.test_api_button.clicked.connect(self.test_api_shopee)
        self.test_groups_button = QPushButton("Testar grupos")
        self.test_groups_button.setObjectName("toolButton")
        self.test_groups_button.setMinimumHeight(44)
        self.test_groups_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.test_groups_button.clicked.connect(self.test_whatsapp_groups)
        self.test_shopee_offers_button = QPushButton("Testar busca Shopee")
        self.test_shopee_offers_button.setObjectName("toolButton")
        self.test_shopee_offers_button.setMinimumHeight(40)
        self.test_shopee_offers_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.test_shopee_offers_button.clicked.connect(self.test_shopee_offers)
        self.export_config_button = QPushButton("Exportar configurações")
        self.export_config_button.setObjectName("toolButton")
        self.export_config_button.setMinimumHeight(40)
        self.export_config_button.clicked.connect(self.export_settings)
        self.import_config_button = QPushButton("Importar configurações")
        self.import_config_button.setObjectName("toolButton")
        self.import_config_button.setMinimumHeight(40)
        self.import_config_button.clicked.connect(self.import_settings)
        self.restore_backup_button = QPushButton("Restaurar backup")
        self.restore_backup_button.setObjectName("toolButton")
        self.restore_backup_button.setMinimumHeight(40)
        self.restore_backup_button.clicked.connect(self.restore_settings_backup)
        self.first_setup_button = QPushButton("Primeira configuração")
        self.first_setup_button.setObjectName("toolButton")
        self.first_setup_button.setMinimumHeight(40)
        self.first_setup_button.clicked.connect(self.show_first_setup_guide)
        self.diagnostic_button = QPushButton("Verificar ambiente")
        self.diagnostic_button.setObjectName("toolButton")
        self.diagnostic_button.setMinimumHeight(40)
        self.diagnostic_button.clicked.connect(self.run_diagnostics)
        self.api_status_label = QLabel()
        self.api_status_label.setObjectName("validationBadge")
        self.api_status_label.setMinimumHeight(34)
        self.api_status_label.setAlignment(Qt.AlignCenter)
        self.api_status_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.groups_status_label = QLabel()
        self.groups_status_label.setObjectName("validationBadge")
        self.groups_status_label.setMinimumHeight(34)
        self.groups_status_label.setAlignment(Qt.AlignCenter)
        self.groups_status_label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.validation_spacer = QFrame()
        self.validation_spacer.setObjectName("settingsPanel")
        self.validation_spacer.setFixedHeight(4)

        self.config_tabs = QTabWidget()
        self.config_tabs.setObjectName("configTabs")
        self.config_tabs.setMinimumHeight(390)

        execution_config = self.config_card("Execução")
        execution_form = QHBoxLayout(execution_config)
        execution_form.setContentsMargins(22, 16, 22, 16)
        execution_form.setSpacing(20)

        timing_panel = QFrame()
        timing_panel.setObjectName("settingsPanel")
        timing_panel.setMinimumWidth(500)
        timing_layout = QVBoxLayout(timing_panel)
        timing_layout.setContentsMargins(0, 0, 0, 0)
        timing_layout.setSpacing(8)

        timing_grid = QGridLayout()
        timing_grid.setContentsMargins(0, 0, 0, 0)
        timing_grid.setHorizontalSpacing(12)
        timing_grid.setVerticalSpacing(6)
        timing_grid.addWidget(self.label_with_help(
            "Modo de operação",
            "Escolha se o bot vai copiar ofertas de um grupo do WhatsApp ou buscar ofertas direto pela API da Shopee.",
        ), 0, 0)
        timing_grid.addWidget(self.modo_origem_ofertas_input, 0, 1)
        timing_grid.addWidget(self.label_with_help(
            "Grupos de destino",
            "Grupos/conversas normais do WhatsApp (aba padrão \"Conversas\") que "
            "receberão as ofertas. Separe mais de um usando ponto e vírgula.",
        ), 1, 0)
        timing_grid.addWidget(self.destinos_input, 1, 1)
        timing_grid.addWidget(self.label_with_help(
            "Canais de destino",
            "Opcional. Canais do WhatsApp (WhatsApp Channels), que ficam em outra aba "
            "do app (\"Canais\") e só aceitam postagem de quem administra o canal. "
            "Pode ser usado junto com os grupos acima: a mesma oferta é enviada para "
            "os dois. Separe mais de um canal usando ponto e vírgula.",
        ), 2, 0)
        timing_grid.addWidget(self.canais_destino_input, 2, 1)
        timing_grid.setRowMinimumHeight(0, 44)
        timing_grid.setRowMinimumHeight(1, 44)
        timing_grid.setRowMinimumHeight(2, 44)
        timing_grid.setColumnMinimumWidth(0, 205)
        timing_grid.setColumnStretch(1, 1)
        timing_layout.addLayout(timing_grid)
        timing_layout.addSpacing(4)

        options_panel = QFrame()
        options_panel.setObjectName("settingsPanel")
        options_panel.setMinimumWidth(430)
        options_layout = QVBoxLayout(options_panel)
        options_layout.setContentsMargins(0, 0, 0, 0)
        options_layout.setSpacing(7)
        self.aguardar_previa_input.setToolTip("Espera a prévia do link aparecer no WhatsApp antes de enviar a mensagem.")
        self.modo_simulacao_input.setToolTip("Converte e registra logs, mas não envia ofertas no grupo/canal de destino.")
        self.headless_input.setToolTip("Abre o navegador minimizado. O WhatsApp Web continua visível se você restaurar a janela.")
        options_layout.addWidget(self.aguardar_previa_input)
        options_layout.addWidget(self.modo_simulacao_input)
        options_layout.addWidget(self.headless_input)
        options_layout.addSpacing(2)
        schedule_row = QHBoxLayout()
        schedule_row.setContentsMargins(0, 0, 0, 0)
        schedule_row.setSpacing(6)
        schedule_row.addWidget(self.label_with_help(
            "Horário ativo",
            "Período em que o bot pode monitorar e enviar ofertas. Use o formato HH:MM.",
        ))
        schedule_row.addWidget(self.horario_inicio_input)
        schedule_row.addWidget(QLabel("até"))
        schedule_row.addWidget(self.horario_fim_input)
        timing_layout.addWidget(self.agendamento_ativo_input)
        timing_layout.addLayout(schedule_row)
        timing_layout.addSpacing(4)
        timing_layout.addWidget(self.recuperacao_automatica_input)
        timing_layout.addStretch(1)
        options_layout.addWidget(self.diagnostic_button)
        options_layout.addSpacing(14)

        config_file_buttons = QGridLayout()
        config_file_buttons.setContentsMargins(0, 0, 0, 0)
        config_file_buttons.setHorizontalSpacing(12)
        config_file_buttons.setVerticalSpacing(8)
        config_file_buttons.addWidget(self.first_setup_button, 0, 0)
        config_file_buttons.addWidget(self.export_config_button, 0, 1)
        config_file_buttons.addWidget(self.import_config_button, 1, 0)
        config_file_buttons.addWidget(self.restore_backup_button, 1, 1)
        config_file_buttons.setColumnStretch(0, 1)
        config_file_buttons.setColumnStretch(1, 1)
        options_layout.addLayout(config_file_buttons)
        options_layout.addStretch(1)

        execution_form.addWidget(timing_panel, 1)
        execution_form.addWidget(options_panel, 1)

        group_config = self.config_card("Modo Grupo")
        group_form = QHBoxLayout(group_config)
        group_form.setContentsMargins(22, 18, 22, 18)
        group_form.setSpacing(22)

        group_fields_panel = QFrame()
        group_fields_panel.setObjectName("settingsPanel")
        group_fields_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        group_fields_layout = QGridLayout(group_fields_panel)
        group_fields_layout.setContentsMargins(0, 0, 0, 0)
        group_fields_layout.setHorizontalSpacing(14)
        group_fields_layout.setVerticalSpacing(12)
        group_fields_layout.addWidget(self.label_with_help(
            "Grupo de origem",
            "Nome exato do grupo do WhatsApp de onde o bot vai ler as ofertas.",
        ), 0, 0)
        group_fields_layout.addWidget(self.grupo_origem_input, 0, 1)
        group_fields_layout.setRowMinimumHeight(0, 50)
        group_fields_layout.setColumnMinimumWidth(0, 210)
        group_fields_layout.setColumnStretch(1, 1)

        group_actions_panel = QFrame()
        group_actions_panel.setObjectName("settingsPanel")
        group_actions_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        group_actions_layout = QVBoxLayout(group_actions_panel)
        group_actions_layout.setContentsMargins(0, 0, 0, 0)
        group_actions_layout.setSpacing(10)
        self.processar_historico_input.setToolTip("Ao iniciar, o bot também analisa mensagens já visíveis na tela do grupo.")
        self.ignorar_links_enviados_input.setToolTip("Evita reenviar ofertas cujo link original já aparece no histórico.")
        group_actions_layout.addWidget(self.processar_historico_input)
        group_actions_layout.addWidget(self.ignorar_links_enviados_input)
        group_actions_layout.addSpacing(12)
        group_actions_layout.addWidget(self.test_groups_button)
        group_actions_layout.addWidget(self.groups_status_label)
        group_actions_layout.addStretch(1)

        group_form.addWidget(group_fields_panel, 1)
        group_form.addWidget(group_actions_panel, 1)
        group_form.setStretch(0, 1)
        group_form.setStretch(1, 1)

        ai_config = self.config_card("IA")
        ai_form = QHBoxLayout(ai_config)
        ai_form.setContentsMargins(22, 18, 22, 18)
        ai_form.setSpacing(22)

        ai_fields_panel = QFrame()
        ai_fields_panel.setObjectName("settingsPanel")
        ai_fields_layout = QGridLayout(ai_fields_panel)
        ai_fields_layout.setContentsMargins(0, 0, 0, 0)
        ai_fields_layout.setHorizontalSpacing(14)
        ai_fields_layout.setVerticalSpacing(10)
        ai_fields_layout.addWidget(self.label_with_help(
            "Chave Gemini",
            "Chave da API do Google Gemini usada para revisar o texto das ofertas.",
        ), 0, 0)
        ai_fields_layout.addWidget(self.ia_api_key_input, 0, 1)
        ai_fields_layout.addWidget(self.label_with_help(
            "Modelo",
            "Modelo usado para revisar textos. Flash Lite tende a ser mais econômico.",
        ), 1, 0)
        ai_fields_layout.addWidget(self.ia_modelo_input, 1, 1)
        ai_fields_layout.addWidget(self.label_with_help(
            "Tipo de revisão",
            "Leve corrige português. Contextual corrige frases estranhas ao produto. Criativa melhora a chamada com mais liberdade.",
        ), 2, 0)
        ai_fields_layout.addWidget(self.ia_tipo_revisao_input, 2, 1)
        ai_fields_layout.addWidget(self.label_with_help(
            "Limite diário",
            "Quantidade máxima de ofertas que o bot tentará revisar com IA por dia.",
        ), 3, 0)
        ai_fields_layout.addWidget(self.ia_limite_diario_input, 3, 1)
        ai_fields_layout.addWidget(self.label_with_help(
            "Tempo máximo",
            "Tempo máximo que o bot espera a resposta da IA antes de seguir sem revisão.",
        ), 4, 0)
        ai_fields_layout.addWidget(self.ia_timeout_input, 4, 1)
        for row in range(5):
            ai_fields_layout.setRowMinimumHeight(row, 46)
        ai_fields_layout.setColumnMinimumWidth(0, 190)
        ai_fields_layout.setColumnStretch(1, 1)

        ai_status_panel = QFrame()
        ai_status_panel.setObjectName("settingsPanel")
        ai_status_layout = QVBoxLayout(ai_status_panel)
        ai_status_layout.setContentsMargins(0, 0, 0, 0)
        ai_status_layout.setSpacing(10)
        self.ia_revisao_input.setToolTip("Quando ativo, o bot tentará revisar o texto final antes de enviar.")
        self.ia_pausar_limite_input.setToolTip("Se a API retornar limite/cota, a IA será pausada sem parar o bot.")
        self.ia_retomar_dia_input.setToolTip("Permite voltar a tentar usar IA automaticamente no próximo dia.")
        ai_status_layout.addWidget(self.ia_revisao_input)
        ai_status_layout.addWidget(self.ia_pausar_limite_input)
        ai_status_layout.addWidget(self.ia_retomar_dia_input)
        ai_status_layout.addWidget(self.test_ai_button)
        ai_status_layout.addWidget(self.ia_status_box, 1)

        ai_form.addWidget(ai_fields_panel, 1)
        ai_form.addWidget(ai_status_panel, 1)
        ai_form.setStretch(0, 1)
        ai_form.setStretch(1, 1)

        api_config = self.config_card("API")
        api_form = QHBoxLayout(api_config)
        api_form.setContentsMargins(22, 18, 22, 18)
        api_form.setSpacing(22)

        api_fields_panel = QFrame()
        api_fields_panel.setObjectName("settingsPanel")
        api_fields_layout = QGridLayout(api_fields_panel)
        api_fields_layout.setContentsMargins(0, 0, 0, 0)
        api_fields_layout.setHorizontalSpacing(14)
        api_fields_layout.setVerticalSpacing(12)
        api_fields_layout.addWidget(self.label_with_help(
            "Shopee AppID",
            "Código da API aberta da Shopee, disponível na área de afiliados.",
        ), 0, 0)
        api_fields_layout.addWidget(self.app_id_input, 0, 1)
        api_fields_layout.addWidget(self.label_with_help(
            "Shopee Secret",
            "Chave secreta da API da Shopee. Ela é usada para assinar as conversões de link.",
        ), 1, 0)
        api_fields_layout.addWidget(self.secret_input, 1, 1)
        api_fields_layout.addWidget(self.label_with_help(
            "Sub IDs",
            "Marcadores opcionais para identificar origem/campanha nos relatórios da Shopee. Use apenas letras e números.",
        ), 2, 0)
        api_fields_layout.addWidget(self.sub_ids_input, 2, 1)
        for row in range(3):
            api_fields_layout.setRowMinimumHeight(row, 50)
        api_fields_layout.setColumnMinimumWidth(0, 210)
        api_fields_layout.setColumnStretch(1, 1)

        api_actions_panel = QFrame()
        api_actions_panel.setObjectName("settingsPanel")
        api_actions_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        api_actions_layout = QVBoxLayout(api_actions_panel)
        api_actions_layout.setContentsMargins(0, 0, 0, 0)
        api_actions_layout.setSpacing(10)
        api_actions_layout.addWidget(self.test_api_button)
        api_actions_layout.addWidget(self.api_status_label)
        api_actions_layout.addStretch(1)

        api_form.addWidget(api_fields_panel, 1)
        api_form.addWidget(api_actions_panel, 1)
        api_form.setStretch(0, 1)
        api_form.setStretch(1, 1)

        offers_config = self.config_card("Modo Shopee")
        offers_form = QHBoxLayout(offers_config)
        offers_form.setContentsMargins(22, 18, 22, 18)
        offers_form.setSpacing(22)

        offers_search_panel = QFrame()
        offers_search_panel.setObjectName("settingsPanel")
        offers_search_layout = QGridLayout(offers_search_panel)
        offers_search_layout.setContentsMargins(0, 0, 0, 0)
        offers_search_layout.setHorizontalSpacing(14)
        offers_search_layout.setVerticalSpacing(6)
        offers_search_layout.addWidget(self.label_with_help(
            "Link de coleção",
            "Opcional. Link de uma coleção/campanha da Shopee. Os produtos são capturados "
            "direto da página da coleção (a API da Shopee não filtra mais por coleção). "
            "Se a coleção pedir login, use a aba \"Navegador\" para logar antes de iniciar o bot.",
        ), 0, 0)
        offers_search_layout.addWidget(self.shopee_ofertas_colecao_input, 0, 1)
        offers_search_layout.addWidget(self.label_with_help(
            "Tipo de busca",
            "Escolha qual filtro abaixo vai valer: só o campo correspondente ao tipo "
            "escolhido fica habilitado e é realmente usado na busca.",
        ), 1, 0)
        offers_search_layout.addWidget(self.shopee_ofertas_tipo_input, 1, 1)
        offers_search_layout.addWidget(self.label_with_help(
            "Palavra-chave/nicho",
            "Opcional. Termo usado para buscas por nicho, como casa, beleza, moda ou eletrônicos.",
        ), 2, 0)
        offers_search_layout.addWidget(self.shopee_ofertas_palavra_input, 2, 1)
        offers_search_layout.addWidget(self.label_with_help(
            "Loja/marca",
            "Opcional. ID para buscar ofertas de uma loja ou marca específica, quando a API permitir.",
        ), 3, 0)
        offers_search_layout.addWidget(self.shopee_ofertas_shop_id_input, 3, 1)
        offers_search_layout.addWidget(self.label_with_help(
            "Categoria/nicho",
            "Opcional. Escolha uma ou mais categorias (clique para marcar/desmarcar). "
            "O bot só deixa passar ofertas cujo nome do produto realmente combine com "
            "a(s) categoria(s) escolhida(s).",
        ), 4, 0)
        offers_search_layout.addWidget(self.shopee_ofertas_categoria_input, 4, 1)
        offers_search_layout.addWidget(self.label_with_help(
            "Ordenação",
            "Opcional. Prioridade usada para escolher quais ofertas aparecem primeiro.",
        ), 5, 0)
        offers_search_layout.addWidget(self.shopee_ofertas_ordenacao_input, 5, 1)
        offers_search_layout.addWidget(self.test_shopee_offers_button, 6, 0, 1, 2)
        for row in range(6):
            offers_search_layout.setRowMinimumHeight(row, 40)
        offers_search_layout.setRowMinimumHeight(6, 40)
        offers_search_layout.setColumnMinimumWidth(0, 190)
        offers_search_layout.setColumnStretch(1, 1)


        browser_config = self.config_card("Navegador (Shopee)")
        browser_layout = QVBoxLayout(browser_config)
        browser_layout.setContentsMargins(22, 18, 22, 18)
        browser_layout.setSpacing(12)

        browser_intro = QLabel(
            "O bot usa o Google Chrome instalado no PC para tarefas auxiliares da Shopee "
            "(como resolver links curtos e baixar a imagem do produto no Modo Grupo). A "
            "captura de produtos de uma coleção agora é feita pela API oficial da Shopee, "
            "por aproximação, sem depender do navegador. É separado do navegador do "
            "WhatsApp Web, que usa seu próprio perfil dedicado."
        )
        browser_intro.setWordWrap(True)
        browser_intro.setObjectName("historySummary")
        browser_layout.addWidget(browser_intro)

        browser_actions = QHBoxLayout()
        browser_actions.setContentsMargins(0, 0, 0, 0)
        browser_actions.setSpacing(10)
        self.shopee_login_button = QPushButton("Fazer login na Shopee no navegador do bot")
        self.shopee_login_button.setObjectName("compactToolButton")
        self.shopee_login_button.clicked.connect(self.abrir_shopee_para_login)
        browser_actions.addWidget(self.shopee_login_button)
        browser_layout.addLayout(browser_actions)

        browser_layout.addStretch(1)



        offers_filters_panel = QFrame()
        offers_filters_panel.setObjectName("settingsPanel")
        offers_filters_layout = QGridLayout(offers_filters_panel)
        offers_filters_layout.setContentsMargins(0, 0, 0, 0)
        offers_filters_layout.setHorizontalSpacing(14)
        offers_filters_layout.setVerticalSpacing(6)
        offers_filters_layout.addWidget(self.label_with_help(
            "Cupom de desconto",
            "Opcional. Cupom que será incluído no texto das ofertas enviadas pelo Modo Shopee.",
        ), 0, 0)
        offers_filters_layout.addWidget(self.shopee_ofertas_cupom_input, 0, 1, 1, 3)
        offers_filters_layout.addWidget(self.label_with_help(
            "Comissão mínima (%)",
            "Opcional. Ignora ofertas com comissão abaixo do percentual informado.",
        ), 1, 0)
        offers_filters_layout.addWidget(self.shopee_ofertas_comissao_input, 1, 1)
        offers_filters_layout.addWidget(self.label_with_help(
            "Desconto mínimo (%)",
            "Opcional. Prioriza produtos com desconto maior.",
        ), 1, 2)
        offers_filters_layout.addWidget(self.shopee_ofertas_desconto_input, 1, 3)
        offers_filters_layout.addWidget(self.label_with_help(
            "Preço mínimo",
            "Opcional. Preço mínimo desejado. Pode ficar em branco.",
        ), 2, 0)
        offers_filters_layout.addWidget(self.shopee_ofertas_preco_min_input, 2, 1)
        offers_filters_layout.addWidget(self.label_with_help(
            "Preço máximo",
            "Opcional. Preço máximo desejado. Pode ficar em branco.",
        ), 2, 2)
        offers_filters_layout.addWidget(self.shopee_ofertas_preco_max_input, 2, 3)
        offers_filters_layout.addWidget(self.label_with_help(
            "Vendas mínimas",
            "Opcional. Evita produtos com pouca saída.",
        ), 3, 0)
        offers_filters_layout.addWidget(self.shopee_ofertas_vendas_input, 3, 1)
        offers_filters_layout.addWidget(self.label_with_help(
            "Avaliação mínima",
            "Opcional. Nota mínima do produto, por exemplo 4.5. Pode ficar em branco.",
        ), 3, 2)
        offers_filters_layout.addWidget(self.shopee_ofertas_avaliacao_input, 3, 3)
        offers_filters_layout.addWidget(self.label_with_help(
            "Intervalo entre envios",
            "Opcional. Tempo mínimo entre uma oferta automática e outra.",
        ), 4, 0)
        offers_filters_layout.addWidget(self.shopee_ofertas_intervalo_input, 4, 1)
        offers_filters_layout.addWidget(self.shopee_ofertas_nao_repetir_input, 5, 0, 1, 4)
        offers_filters_layout.addWidget(self.shopee_ofertas_status_box, 6, 0, 1, 4)
        for row in range(6):
            offers_filters_layout.setRowMinimumHeight(row, 42)
        offers_filters_layout.setColumnMinimumWidth(0, 210)
        offers_filters_layout.setColumnMinimumWidth(1, 125)
        offers_filters_layout.setColumnMinimumWidth(2, 205)
        offers_filters_layout.setColumnMinimumWidth(3, 125)
        offers_filters_layout.setColumnStretch(1, 1)
        offers_filters_layout.setColumnStretch(3, 1)

        offers_form.addWidget(offers_search_panel, 3)
        offers_form.addWidget(offers_filters_panel, 6)
        offers_form.setStretch(0, 3)
        offers_form.setStretch(1, 6)

        self.config_tabs.addTab(execution_config, "Execução")
        self.config_tabs.addTab(api_config, "API")
        self.config_tabs.addTab(group_config, "Modo Grupo")
        self.config_tabs.addTab(offers_config, "Modo Shopee")
        self.config_tabs.addTab(browser_config, "Navegador")
        self.config_tabs.addTab(ai_config, "IA")

        self.data_tabs = QTabWidget()
        self.data_tabs.setObjectName("dataTabs")

        logs_tab = QWidget()
        logs_layout = QVBoxLayout(logs_tab)
        logs_layout.setContentsMargins(0, 8, 0, 0)
        logs_layout.setSpacing(10)
        logs_filters = QHBoxLayout()
        logs_filters.setContentsMargins(0, 0, 0, 0)
        logs_filters.setSpacing(10)
        logs_buttons = QHBoxLayout()
        logs_buttons.setContentsMargins(0, 0, 0, 0)
        logs_buttons.setSpacing(10)
        self.logs_search_input = QLineEdit()
        self.logs_search_input.setObjectName("logsSearch")
        self.logs_search_input.setPlaceholderText("Buscar nos logs")
        self.logs_search_input.setMinimumHeight(36)
        self.logs_search_input.textChanged.connect(self.apply_log_filters)
        self.logs_type_label = QLabel("Filtro logs")
        self.logs_type_label.setObjectName("historyFilterLabel")
        self.logs_type_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.logs_type_input = QComboBox()
        self.logs_type_input.setObjectName("logsType")
        self.logs_type_input.setMinimumHeight(36)
        self.logs_type_input.addItem("Todos", "all")
        self.logs_type_input.addItem("Erros", "error")
        self.logs_type_input.addItem("Avisos", "warning")
        self.logs_type_input.addItem("Sucessos", "success")
        self.logs_type_input.currentIndexChanged.connect(self.apply_log_filters)
        self.logs_summary_label = QLabel("0 linhas")
        self.logs_summary_label.setObjectName("historySummary")
        self.logs_summary_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.clear_logs_button = QPushButton("Limpar logs da tela")
        self.clear_logs_button.setObjectName("compactToolButton")
        self.clear_logs_button.clicked.connect(self.clear_logs)
        self.open_logs_button = QPushButton("Abrir pasta de logs")
        self.open_logs_button.setObjectName("compactToolButton")
        self.open_logs_button.clicked.connect(self.open_logs_folder)
        self.copy_logs_button = QPushButton("Copiar seleção")
        self.copy_logs_button.setObjectName("compactToolButton")
        self.copy_logs_button.clicked.connect(self.copy_selected_logs)
        self.export_logs_button = QPushButton("Exportar logs")
        self.export_logs_button.setObjectName("compactToolButton")
        self.export_logs_button.clicked.connect(self.export_visible_logs)
        logs_filters.addWidget(self.logs_search_input, 2)
        logs_filters.addWidget(self.logs_type_label)
        logs_filters.addWidget(self.logs_type_input)
        logs_filters.addWidget(self.logs_summary_label, 1)
        logs_buttons.addWidget(self.clear_logs_button)
        logs_buttons.addWidget(self.open_logs_button)
        logs_buttons.addWidget(self.copy_logs_button)
        logs_buttons.addWidget(self.export_logs_button)
        logs_buttons.addStretch(1)
        self.logs = QTextEdit()
        self.logs.setReadOnly(True)
        self.logs.setObjectName("logs")
        logs_layout.addLayout(logs_filters)
        logs_layout.addLayout(logs_buttons)
        logs_layout.addWidget(self.logs)

        history_tab = QWidget()
        history_layout = QVBoxLayout(history_tab)
        history_layout.setContentsMargins(0, 8, 0, 0)
        history_layout.setSpacing(10)

        history_filters = QHBoxLayout()
        history_filters.setContentsMargins(0, 0, 0, 0)
        history_filters.setSpacing(10)
        history_buttons = QHBoxLayout()
        history_buttons.setContentsMargins(0, 0, 0, 0)
        history_buttons.setSpacing(10)
        self.history_search_input = QLineEdit()
        self.history_search_input.setObjectName("historySearch")
        self.history_search_input.setPlaceholderText("Buscar no histórico")
        self.history_search_input.setMinimumHeight(36)
        self.history_search_input.textChanged.connect(self.apply_history_filters)
        self.history_period_label = QLabel("Filtro período")
        self.history_period_label.setObjectName("historyFilterLabel")
        self.history_period_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.history_period_input = QComboBox()
        self.history_period_input.setObjectName("historyPeriod")
        self.history_period_input.setMinimumHeight(36)
        self.history_period_input.addItem("Todos", "all")
        self.history_period_input.addItem("Hoje", "today")
        self.history_period_input.addItem("Últimos 7 dias", "7")
        self.history_period_input.addItem("Últimos 30 dias", "30")
        self.history_period_input.currentIndexChanged.connect(self.apply_history_filters)
        self.refresh_history_button = QPushButton("Atualizar histórico")
        self.refresh_history_button.setObjectName("compactToolButton")
        self.refresh_history_button.clicked.connect(self.refresh_history_with_message)
        self.copy_affiliate_link_button = QPushButton("Copiar link afiliado")
        self.copy_affiliate_link_button.setObjectName("compactToolButton")
        self.copy_affiliate_link_button.clicked.connect(self.copy_selected_affiliate_link)
        self.export_history_button = QPushButton("Exportar CSV")
        self.export_history_button.setObjectName("compactToolButton")
        self.export_history_button.clicked.connect(self.export_history_csv)
        self.clear_history_button = QPushButton("Limpar histórico")
        self.clear_history_button.setObjectName("compactDangerButton")
        self.clear_history_button.clicked.connect(self.clear_history)
        history_filters.addWidget(self.history_search_input, 2)
        history_filters.addWidget(self.history_period_label)
        history_filters.addWidget(self.history_period_input)
        self.history_summary_label = QLabel("0 registros")
        self.history_summary_label.setObjectName("historySummary")
        self.history_summary_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        history_filters.addWidget(self.history_summary_label, 1)
        history_buttons.addWidget(self.refresh_history_button)
        history_buttons.addWidget(self.copy_affiliate_link_button)
        history_buttons.addWidget(self.export_history_button)
        history_buttons.addWidget(self.clear_history_button)
        history_buttons.addStretch(1)

        self.history_table = QTableWidget(0, 5)
        self.history_table.setObjectName("historyTable")
        self.history_table.setHorizontalHeaderLabels([
            "Data",
            "Link original",
            "Grupo destino",
            "Link afiliado",
            "Texto",
        ])
        self.history_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.history_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.history_table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.history_table.setItemDelegate(NoFocusTableDelegate(self.history_table))
        self.history_table.verticalHeader().setVisible(False)
        header = self.history_table.horizontalHeader()
        header.setStretchLastSection(True)
        header.resizeSection(0, 145)
        header.resizeSection(1, 240)
        header.resizeSection(2, 160)
        header.resizeSection(3, 240)

        history_layout.addLayout(history_filters)
        history_layout.addLayout(history_buttons)
        history_layout.addWidget(self.history_table)

        indicators_tab = QWidget()
        indicators_layout = QVBoxLayout(indicators_tab)
        indicators_layout.setContentsMargins(0, 8, 0, 0)
        indicators_layout.setSpacing(10)

        indicators_filters = QHBoxLayout()
        indicators_filters.setContentsMargins(0, 0, 0, 0)
        indicators_filters.setSpacing(10)
        self.indicators_period_label = QLabel("Filtro período")
        self.indicators_period_label.setObjectName("historyFilterLabel")
        self.indicators_period_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.indicators_period_input = QComboBox()
        self.indicators_period_input.setObjectName("historyPeriod")
        self.indicators_period_input.setMinimumHeight(36)
        self.indicators_period_input.addItem("Dia anterior", "yesterday")
        self.indicators_period_input.addItem("Últimos 7 dias", "7")
        self.indicators_period_input.addItem("Últimos 30 dias", "30")
        self.indicators_period_input.addItem("Personalizado", "custom")
        self.indicators_period_input.currentIndexChanged.connect(self.update_indicator_period_fields)
        self.indicators_start_date = QDateEdit()
        self.indicators_start_date.setObjectName("historyPeriod")
        self.indicators_start_date.setCalendarPopup(True)
        self.indicators_start_date.setDisplayFormat("dd/MM/yyyy")
        self.indicators_start_date.setMinimumHeight(36)
        self.indicators_end_date = QDateEdit()
        self.indicators_end_date.setObjectName("historyPeriod")
        self.indicators_end_date.setCalendarPopup(True)
        self.indicators_end_date.setDisplayFormat("dd/MM/yyyy")
        self.indicators_end_date.setMinimumHeight(36)
        self.refresh_indicators_button = QPushButton("Atualizar indicadores")
        self.refresh_indicators_button.setObjectName("compactToolButton")
        self.refresh_indicators_button.clicked.connect(self.refresh_indicators)
        self.indicators_summary_label = QLabel("Nenhum dado carregado")
        self.indicators_summary_label.setObjectName("historySummary")
        self.indicators_summary_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        indicators_filters.addWidget(self.indicators_period_label)
        indicators_filters.addWidget(self.indicators_period_input)
        indicators_filters.addWidget(QLabel("De"))
        indicators_filters.addWidget(self.indicators_start_date)
        indicators_filters.addWidget(QLabel("até"))
        indicators_filters.addWidget(self.indicators_end_date)
        indicators_filters.addWidget(self.refresh_indicators_button)
        indicators_filters.addWidget(self.indicators_summary_label, 1)

        indicators_cards = QHBoxLayout()
        indicators_cards.setContentsMargins(0, 0, 0, 0)
        indicators_cards.setSpacing(10)
        self.indicators_sales_label = QLabel("0")
        self.indicators_commission_label = QLabel("R$ 0,00")
        indicators_cards.addWidget(self.metric_card("Vendas", self.indicators_sales_label))
        indicators_cards.addWidget(self.metric_card("Comissão", self.indicators_commission_label))
        indicators_cards.addStretch(1)

        indicators_content = QHBoxLayout()
        indicators_content.setContentsMargins(0, 0, 0, 0)
        indicators_content.setSpacing(10)
        self.indicators_table = QTableWidget(0, 3)
        self.indicators_table.setObjectName("historyTable")
        self.indicators_table.setHorizontalHeaderLabels([
            "Categoria",
            "Vendas",
            "Comissão",
        ])
        self.indicators_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.indicators_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.indicators_table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.indicators_table.setItemDelegate(NoFocusTableDelegate(self.indicators_table))
        self.indicators_table.verticalHeader().setVisible(False)
        indicators_header = self.indicators_table.horizontalHeader()
        indicators_header.setStretchLastSection(True)
        indicators_header.resizeSection(0, 180)
        indicators_header.resizeSection(1, 90)
        self.indicators_chart = QTextEdit()
        self.indicators_chart.setObjectName("logs")
        self.indicators_chart.setReadOnly(True)
        self.indicators_chart.setPlaceholderText("O gráfico por categoria aparecerá aqui.")
        indicators_content.addWidget(self.indicators_table, 1)
        indicators_content.addWidget(self.indicators_chart, 1)

        indicators_layout.addLayout(indicators_filters)
        indicators_layout.addLayout(indicators_cards)
        indicators_layout.addLayout(indicators_content, 1)

        # ===================== Relatório de conversões =====================
        conversoes_tab = QWidget()
        conversoes_layout = QVBoxLayout(conversoes_tab)
        conversoes_layout.setContentsMargins(0, 8, 0, 0)
        conversoes_layout.setSpacing(10)

        conversoes_intro = QLabel(
            "Produtos convertidos no período (dados da API de conversões da Shopee), "
            "com quantidade vendida, preço, comissão e data da última compra. "
            "Clique nos títulos das colunas para reordenar."
        )
        conversoes_intro.setWordWrap(True)
        conversoes_intro.setObjectName("historySummary")
        conversoes_layout.addWidget(conversoes_intro)

        conversoes_filters = QHBoxLayout()
        conversoes_filters.setContentsMargins(0, 0, 0, 0)
        conversoes_filters.setSpacing(10)
        self.conversoes_period_label = QLabel("Filtro período")
        self.conversoes_period_label.setObjectName("historyFilterLabel")
        self.conversoes_period_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.conversoes_period_input = QComboBox()
        self.conversoes_period_input.setObjectName("historyPeriod")
        self.conversoes_period_input.setMinimumHeight(36)
        self.conversoes_period_input.addItem("Dia anterior", "yesterday")
        self.conversoes_period_input.addItem("Últimos 7 dias", "7")
        self.conversoes_period_input.addItem("Últimos 30 dias", "30")
        self.conversoes_period_input.addItem("Personalizado", "custom")
        self.conversoes_period_input.currentIndexChanged.connect(self.update_conversoes_period_fields)
        self.conversoes_start_date = QDateEdit()
        self.conversoes_start_date.setObjectName("historyPeriod")
        self.conversoes_start_date.setCalendarPopup(True)
        self.conversoes_start_date.setDisplayFormat("dd/MM/yyyy")
        self.conversoes_start_date.setMinimumHeight(36)
        self.conversoes_end_date = QDateEdit()
        self.conversoes_end_date.setObjectName("historyPeriod")
        self.conversoes_end_date.setCalendarPopup(True)
        self.conversoes_end_date.setDisplayFormat("dd/MM/yyyy")
        self.conversoes_end_date.setMinimumHeight(36)

        self.conversoes_sort_label = QLabel("Ordenar por")
        self.conversoes_sort_label.setObjectName("historyFilterLabel")
        self.conversoes_sort_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.conversoes_sort_input = QComboBox()
        self.conversoes_sort_input.setObjectName("historyPeriod")
        self.conversoes_sort_input.setMinimumHeight(36)
        self.conversoes_sort_input.addItem("Maior comissão", "comissao")
        self.conversoes_sort_input.addItem("Mais vendidos", "quantidade")
        self.conversoes_sort_input.addItem("Nome (A-Z)", "nome")
        self.conversoes_sort_input.addItem("Data da compra", "data")
        self.conversoes_sort_input.currentIndexChanged.connect(self.render_conversoes_atual)

        self.refresh_conversoes_button = QPushButton("Atualizar conversões")
        self.refresh_conversoes_button.setObjectName("compactToolButton")
        self.refresh_conversoes_button.clicked.connect(self.refresh_conversoes)
        self.exportar_conversoes_button = QPushButton("Exportar Excel")
        self.exportar_conversoes_button.setObjectName("compactToolButton")
        self.exportar_conversoes_button.clicked.connect(self.exportar_conversoes_excel)
        self.conversoes_summary_label = QLabel("Nenhum dado carregado")
        self.conversoes_summary_label.setObjectName("historySummary")
        self.conversoes_summary_label.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

        conversoes_filters.addWidget(self.conversoes_period_label)
        conversoes_filters.addWidget(self.conversoes_period_input)
        conversoes_filters.addWidget(QLabel("De"))
        conversoes_filters.addWidget(self.conversoes_start_date)
        conversoes_filters.addWidget(QLabel("até"))
        conversoes_filters.addWidget(self.conversoes_end_date)
        conversoes_filters.addWidget(self.conversoes_sort_label)
        conversoes_filters.addWidget(self.conversoes_sort_input)
        conversoes_filters.addWidget(self.refresh_conversoes_button)
        conversoes_filters.addWidget(self.exportar_conversoes_button)
        conversoes_filters.addWidget(self.conversoes_summary_label, 1)

        conversoes_cards = QHBoxLayout()
        conversoes_cards.setContentsMargins(0, 0, 0, 0)
        conversoes_cards.setSpacing(10)
        self.conversoes_produtos_label = QLabel("0")
        self.conversoes_qtd_label = QLabel("0")
        self.conversoes_comissao_label = QLabel("R$ 0,00")
        conversoes_cards.addWidget(self.metric_card("Produtos", self.conversoes_produtos_label))
        conversoes_cards.addWidget(self.metric_card("Vendas", self.conversoes_qtd_label))
        conversoes_cards.addWidget(self.metric_card("Comissão", self.conversoes_comissao_label))
        conversoes_cards.addStretch(1)

        self.conversoes_table = QTableWidget(0, 6)
        self.conversoes_table.setObjectName("historyTable")
        self.conversoes_table.setHorizontalHeaderLabels([
            "Produto",
            "Loja",
            "Qtd",
            "Preço",
            "Comissão",
            "Última compra",
        ])
        self.conversoes_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.conversoes_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.conversoes_table.setFocusPolicy(Qt.FocusPolicy.NoFocus)
        self.conversoes_table.setItemDelegate(NoFocusTableDelegate(self.conversoes_table))
        self.conversoes_table.verticalHeader().setVisible(False)
        self.conversoes_table.setSortingEnabled(True)
        conversoes_header = self.conversoes_table.horizontalHeader()
        conversoes_header.setStretchLastSection(True)
        conversoes_header.resizeSection(0, 300)
        conversoes_header.resizeSection(1, 140)
        conversoes_header.resizeSection(2, 60)
        conversoes_header.resizeSection(3, 90)
        conversoes_header.resizeSection(4, 100)

        conversoes_layout.addLayout(conversoes_filters)
        conversoes_layout.addLayout(conversoes_cards)
        conversoes_layout.addWidget(self.conversoes_table, 1)

        self._conversoes_report_atual = None
        self.update_conversoes_period_fields()


        email_report_tab = QWidget()
        email_report_layout = QVBoxLayout(email_report_tab)
        email_report_layout.setContentsMargins(0, 8, 0, 0)
        email_report_layout.setSpacing(12)

        email_report_intro = QLabel(
            "Envia por email, todo dia no horário abaixo, um resumo de vendas e "
            "comissão por categoria de origem (cash-b, redes sociais, siteconversor, "
            "grupoofertas etc; ontem + mês atual até agora). Funciona mesmo "
            "com o bot parado, desde que o app esteja aberto."
        )
        email_report_intro.setWordWrap(True)
        email_report_intro.setObjectName("historySummary")

        self.relatorio_email_ativo_input = QCheckBox("Enviar relatório automaticamente todo dia")

        email_report_form = QFormLayout()
        email_report_form.setContentsMargins(0, 0, 0, 0)
        email_report_form.setHorizontalSpacing(14)
        email_report_form.setVerticalSpacing(10)
        email_report_form.setLabelAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        email_report_form.setFieldGrowthPolicy(QFormLayout.AllNonFixedFieldsGrow)

        self.relatorio_email_hora_input = QLineEdit()
        self.relatorio_email_hora_input.setPlaceholderText("HH:MM (ex: 11:00)")
        self.relatorio_email_hora_input.setMinimumHeight(36)
        email_report_form.addRow(self.label_with_help(
            "Horário de envio",
            "Formato 24h HH:MM. O relatório é conferido a cada minuto enquanto o app estiver aberto.",
        ), self.relatorio_email_hora_input)

        self.relatorio_email_destinatarios_input = QLineEdit()
        self.relatorio_email_destinatarios_input.setPlaceholderText("Destinatários, separados por vírgula")
        self.relatorio_email_destinatarios_input.setMinimumHeight(36)
        email_report_form.addRow(self.label_with_help(
            "Destinatários",
            "Um ou mais emails, separados por vírgula.",
        ), self.relatorio_email_destinatarios_input)

        self.relatorio_email_remetente_input = QLineEdit()
        self.relatorio_email_remetente_input.setPlaceholderText("seuemail@gmail.com")
        self.relatorio_email_remetente_input.setMinimumHeight(36)
        email_report_form.addRow(self.label_with_help(
            "Remetente (Gmail)",
            "O email do Gmail que vai enviar o relatório.",
        ), self.relatorio_email_remetente_input)

        self.relatorio_email_senha_input = QLineEdit()
        self.relatorio_email_senha_input.setPlaceholderText("Senha de app do Gmail (16 caracteres)")
        self.relatorio_email_senha_input.setEchoMode(QLineEdit.Password)
        self.relatorio_email_senha_input.setMinimumHeight(36)
        email_report_form.addRow(self.label_with_help(
            "Senha de app",
            "No Gmail: Conta Google > Segurança > Verificação em duas etapas > Senhas de app. "
            "Não é a senha normal da conta.",
        ), self.relatorio_email_senha_input)

        self.relatorio_email_servidor_input = QLineEdit()
        self.relatorio_email_servidor_input.setPlaceholderText("smtp.gmail.com")
        self.relatorio_email_servidor_input.setMinimumHeight(36)
        email_report_form.addRow(self.label_with_help(
            "Servidor SMTP",
            "Opcional. Padrão smtp.gmail.com para contas Gmail.",
        ), self.relatorio_email_servidor_input)

        self.relatorio_email_porta_input = QLineEdit()
        self.relatorio_email_porta_input.setPlaceholderText("587")
        self.relatorio_email_porta_input.setMinimumHeight(36)
        email_report_form.addRow(self.label_with_help(
            "Porta SMTP",
            "Opcional. Padrão 587.",
        ), self.relatorio_email_porta_input)

        email_report_actions = QHBoxLayout()
        email_report_actions.setContentsMargins(0, 0, 0, 0)
        email_report_actions.setSpacing(10)
        self.relatorio_email_testar_button = QPushButton("Enviar relatório de teste agora")
        self.relatorio_email_testar_button.setObjectName("compactToolButton")
        self.relatorio_email_testar_button.clicked.connect(self.enviar_relatorio_email_teste)
        self.relatorio_email_status_label = QLabel("Nunca enviado")
        self.relatorio_email_status_label.setObjectName("historySummary")
        email_report_actions.addWidget(self.relatorio_email_testar_button)
        email_report_actions.addWidget(self.relatorio_email_status_label, 1)

        email_report_layout.addWidget(email_report_intro)
        email_report_layout.addWidget(self.relatorio_email_ativo_input)
        email_report_layout.addLayout(email_report_form)
        email_report_layout.addLayout(email_report_actions)
        email_report_layout.addStretch(1)

        self.data_tabs.addTab(logs_tab, "Logs")
        self.data_tabs.addTab(history_tab, "Histórico")
        self.data_tabs.addTab(indicators_tab, "Indicadores")
        self.data_tabs.addTab(conversoes_tab, "Relatório de conversões")
        self.data_tabs.addTab(email_report_tab, "Relatório por email")
        self.update_indicator_period_fields()

        right_layout.addWidget(config_title)
        right_layout.addWidget(self.config_tabs)
        right_layout.addWidget(self.data_tabs, 2)

        outer.addWidget(left, 0)
        outer.addWidget(right, 1)
        self.setCentralWidget(root)

    def label(self, text):
        item = QLabel(text)
        item.setObjectName("fieldLabel")
        item.setMinimumHeight(38)
        item.setWordWrap(True)
        item.setAlignment(Qt.AlignVCenter | Qt.AlignLeft)
        return item

    def label_with_help(self, text, help_text):
        container = QFrame()
        container.setObjectName("labelHelpContainer")
        layout = QHBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(6)

        label = self.label(text)
        help_button = QPushButton("?")
        help_button.setObjectName("helpButton")
        help_button.setFixedSize(20, 20)
        help_button.setToolTip(help_text)
        help_button.setCursor(Qt.CursorShape.WhatsThisCursor)

        layout.addWidget(help_button)
        layout.addWidget(label, 1)
        return container

    def card_title(self, text):
        item = QLabel(text)
        item.setObjectName("cardTitle")
        return item

    def config_card(self, title):
        card = QFrame()
        card.setObjectName("configCard")
        card.setMinimumWidth(0)
        card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        return card

    def metric_card(self, label, value_label):
        card = QFrame()
        card.setObjectName("metricCard")
        layout = QVBoxLayout(card)
        layout.setContentsMargins(14, 12, 14, 12)
        value_label.setObjectName("metricValue")
        name = QLabel(label)
        name.setObjectName("metricLabel")
        layout.addWidget(value_label)
        layout.addWidget(name)
        return card

    def load_fields(self):
        s = self.settings
        self.grupo_origem_input.setText(s.grupo_origem)
        self.destinos_input.setText("; ".join(s.grupos_destino))
        self.canais_destino_input.setText("; ".join(s.canais_destino))
        self.app_id_input.setText(s.shopee_api_app_id)
        self.secret_input.setText(s.shopee_api_secret)
        self.sub_ids_input.setText(", ".join(s.shopee_api_sub_ids))
        self.intervalo_input.setValue(5)
        self.timeout_previa_input.setValue(15)
        self.espera_minima_input.setValue(5)
        self.processar_historico_input.setChecked(s.processar_historico_ao_iniciar)
        self.ignorar_links_enviados_input.setChecked(s.ignorar_links_ja_enviados)
        self.aguardar_previa_input.setChecked(s.aguardar_previa_link)
        self.modo_simulacao_input.setChecked(s.modo_simulacao)
        self.headless_input.setChecked(s.headless)
        self.agendamento_ativo_input.setChecked(s.agendamento_ativo)
        self.horario_inicio_input.setText(s.horario_inicio)
        self.horario_fim_input.setText(s.horario_fim)
        self.recuperacao_automatica_input.setChecked(s.recuperacao_automatica)
        self.recuperacao_max_tentativas_input.setText(str(s.recuperacao_max_tentativas))
        self.recuperacao_intervalo_input.setValue(max(5, round(s.recuperacao_intervalo_segundos)))
        self.ia_revisao_input.setChecked(s.ia_revisao_ativa)
        self.ia_pausar_limite_input.setChecked(s.ia_pausar_ao_limite)
        self.ia_retomar_dia_input.setChecked(s.ia_retomar_no_dia_seguinte)
        self.ia_api_key_input.setText(s.ia_gemini_api_key)
        modelo_index = self.ia_modelo_input.findData(s.ia_modelo)
        self.ia_modelo_input.setCurrentIndex(modelo_index if modelo_index >= 0 else 0)
        tipo_revisao_index = self.ia_tipo_revisao_input.findData(s.ia_tipo_revisao)
        self.ia_tipo_revisao_input.setCurrentIndex(tipo_revisao_index if tipo_revisao_index >= 0 else 1)
        self.ia_limite_diario_input.setText(str(s.ia_limite_diario_revisoes))
        self.ia_timeout_input.setValue(max(3, round(s.ia_timeout_segundos)))
        self.ia_status_box.setPlainText(s.ia_status_mensagem or "Revisão por IA desativada.")
        origem_index = self.modo_origem_ofertas_input.findData(s.modo_origem_ofertas)
        self.modo_origem_ofertas_input.setCurrentIndex(origem_index if origem_index >= 0 else 0)
        tipo_ofertas_index = self.shopee_ofertas_tipo_input.findData(s.shopee_ofertas_tipo_busca)
        self.shopee_ofertas_tipo_input.setCurrentIndex(tipo_ofertas_index if tipo_ofertas_index >= 0 else 0)
        self.atualizar_campos_tipo_busca_shopee()
        ordenacao_index = self.shopee_ofertas_ordenacao_input.findData(s.shopee_ofertas_ordenacao)
        self.shopee_ofertas_ordenacao_input.setCurrentIndex(ordenacao_index if ordenacao_index >= 0 else 0)
        self.shopee_ofertas_colecao_input.setText(s.shopee_ofertas_link_colecao)
        self.shopee_ofertas_cupom_input.setText(s.shopee_ofertas_cupom)
        self.shopee_ofertas_palavra_input.setText(s.shopee_ofertas_palavra_chave)
        self.shopee_ofertas_shop_id_input.setText(s.shopee_ofertas_shop_id)
        self.shopee_ofertas_categoria_input.definir_selecionados(s.shopee_ofertas_categorias)
        self.shopee_ofertas_comissao_input.setText(str(s.shopee_ofertas_comissao_minima))
        self.shopee_ofertas_desconto_input.setText(str(s.shopee_ofertas_desconto_minimo))
        self.shopee_ofertas_preco_min_input.setText(s.shopee_ofertas_preco_minimo)
        self.shopee_ofertas_preco_max_input.setText(s.shopee_ofertas_preco_maximo)
        self.shopee_ofertas_vendas_input.setText(str(s.shopee_ofertas_vendas_minimas))
        self.shopee_ofertas_avaliacao_input.setText(s.shopee_ofertas_avaliacao_minima)
        self.shopee_ofertas_intervalo_input.setValue(max(60, round(s.shopee_ofertas_intervalo_envio_segundos)))
        self.shopee_ofertas_nao_repetir_input.setChecked(s.shopee_ofertas_nao_repetir)
        theme_index = self.tema_visual_input.findData(s.tema_visual)
        self.tema_visual_input.setCurrentIndex(theme_index if theme_index >= 0 else 0)
        self.relatorio_email_ativo_input.setChecked(s.relatorio_email_ativo)
        self.relatorio_email_hora_input.setText(s.relatorio_email_hora or "11:00")
        self.relatorio_email_destinatarios_input.setText(s.relatorio_email_destinatarios)
        self.relatorio_email_remetente_input.setText(s.relatorio_email_remetente)
        self.relatorio_email_senha_input.setText(s.relatorio_email_senha_app)
        self.relatorio_email_servidor_input.setText(s.relatorio_email_servidor_smtp or "smtp.gmail.com")
        self.relatorio_email_porta_input.setText(str(s.relatorio_email_porta_smtp or 587))
        if s.relatorio_email_ultimo_envio:
            self.relatorio_email_status_label.setText(f"Último envio: {s.relatorio_email_ultimo_envio}")
        else:
            self.relatorio_email_status_label.setText("Nunca enviado")
        self.update_validation_badges()
        self.update_configuration_status()

    def collect_settings(self):
        destinos_texto = self.destinos_input.text().replace("\n", ";")
        destinos = [
            item.strip()
            for item in destinos_texto.split(";")
            if item.strip()
        ]
        canais_texto = self.canais_destino_input.text().replace("\n", ";")
        canais = [
            item.strip()
            for item in canais_texto.split(";")
            if item.strip()
        ]
        sub_ids = [
            item.strip()
            for item in self.sub_ids_input.text().split(",")
            if item.strip()
        ]

        self.settings.grupo_origem = self.grupo_origem_input.text().strip()
        self.settings.canais_destino = canais
        self.settings.grupos_destino = destinos
        self.settings.shopee_api_app_id = self.app_id_input.text().strip()
        self.settings.shopee_api_secret = self.secret_input.text().strip()
        self.settings.shopee_api_sub_ids = sub_ids
        self.settings.intervalo_ms = 5000
        self.settings.timeout_previa_link_ms = 15000
        self.settings.timeout_imagem_previa_link_ms = 15000
        self.settings.espera_minima_previa_link_ms = 5000
        self.settings.processar_historico_ao_iniciar = self.processar_historico_input.isChecked()
        self.settings.ignorar_links_ja_enviados = self.ignorar_links_enviados_input.isChecked()
        self.settings.aguardar_previa_link = self.aguardar_previa_input.isChecked()
        self.settings.modo_simulacao = self.modo_simulacao_input.isChecked()
        self.settings.headless = self.headless_input.isChecked()
        self.settings.agendamento_ativo = self.agendamento_ativo_input.isChecked()
        self.settings.horario_inicio = self.horario_inicio_input.text().strip() or "07:00"
        self.settings.horario_fim = self.horario_fim_input.text().strip() or "23:00"
        self.settings.recuperacao_automatica = self.recuperacao_automatica_input.isChecked()
        self.settings.recuperacao_max_tentativas = 5
        self.settings.recuperacao_intervalo_segundos = 30
        self.settings.ia_revisao_ativa = self.ia_revisao_input.isChecked()
        self.settings.ia_pausar_ao_limite = self.ia_pausar_limite_input.isChecked()
        self.settings.ia_retomar_no_dia_seguinte = self.ia_retomar_dia_input.isChecked()
        self.settings.ia_gemini_api_key = self.ia_api_key_input.text().strip()
        self.settings.ia_modelo = self.ia_modelo_input.currentData() or "gemini-3.1-flash-lite"
        self.settings.ia_tipo_revisao = self.ia_tipo_revisao_input.currentData() or "contextual"
        try:
            limite_diario = int(self.ia_limite_diario_input.text().strip())
        except ValueError:
            limite_diario = 0
        self.settings.ia_limite_diario_revisoes = max(0, limite_diario)
        self.settings.ia_timeout_segundos = self.ia_timeout_input.value()
        self.settings.ia_status_mensagem = self.ia_status_box.toPlainText().strip()
        self.settings.modo_origem_ofertas = self.modo_origem_ofertas_input.currentData() or "whatsapp"
        self.settings.shopee_ofertas_tipo_busca = self.shopee_ofertas_tipo_input.currentData() or "geral"
        self.settings.shopee_ofertas_link_colecao = self.shopee_ofertas_colecao_input.text().strip()
        self.settings.shopee_ofertas_cupom = self.shopee_ofertas_cupom_input.text().strip()
        self.settings.shopee_ofertas_palavra_chave = self.shopee_ofertas_palavra_input.text().strip()
        self.settings.shopee_ofertas_shop_id = self.shopee_ofertas_shop_id_input.text().strip()
        self.settings.shopee_ofertas_categorias = self.shopee_ofertas_categoria_input.dados_selecionados()
        # Navegador do bot fixado em Chrome (removidas as opções de Edge, perfil
        # pessoal e modo CDP - a coleção agora vem da API, sem depender de scraping).
        self.settings.shopee_navegador_tipo = "chrome"
        self.settings.shopee_usar_perfil_chrome_pessoal = False
        self.settings.shopee_ofertas_ordenacao = self.shopee_ofertas_ordenacao_input.currentData() or "mais_vendidos"
        self.settings.shopee_ofertas_comissao_minima = self.safe_int_from_field(self.shopee_ofertas_comissao_input)
        self.settings.shopee_ofertas_desconto_minimo = self.safe_int_from_field(self.shopee_ofertas_desconto_input)
        self.settings.shopee_ofertas_preco_minimo = self.shopee_ofertas_preco_min_input.text().strip()
        self.settings.shopee_ofertas_preco_maximo = self.shopee_ofertas_preco_max_input.text().strip()
        self.settings.shopee_ofertas_vendas_minimas = self.safe_int_from_field(self.shopee_ofertas_vendas_input)
        self.settings.shopee_ofertas_avaliacao_minima = self.shopee_ofertas_avaliacao_input.text().strip()
        self.settings.shopee_ofertas_quantidade_por_ciclo = 10
        self.settings.shopee_ofertas_intervalo_envio_segundos = self.shopee_ofertas_intervalo_input.value()
        self.settings.shopee_ofertas_nao_repetir = self.shopee_ofertas_nao_repetir_input.isChecked()
        self.settings.tema_visual = self.tema_visual_input.currentData() or "padrao"
        self.settings.relatorio_email_ativo = self.relatorio_email_ativo_input.isChecked()
        self.settings.relatorio_email_hora = self.relatorio_email_hora_input.text().strip() or "11:00"
        self.settings.relatorio_email_destinatarios = self.relatorio_email_destinatarios_input.text().strip()
        self.settings.relatorio_email_remetente = self.relatorio_email_remetente_input.text().strip()
        self.settings.relatorio_email_senha_app = self.relatorio_email_senha_input.text().strip()
        self.settings.relatorio_email_servidor_smtp = self.relatorio_email_servidor_input.text().strip() or "smtp.gmail.com"
        try:
            porta_smtp = int(self.relatorio_email_porta_input.text().strip() or 587)
        except ValueError:
            porta_smtp = 587
        self.settings.relatorio_email_porta_smtp = porta_smtp
        return self.settings

    def safe_int_from_field(self, widget, default=0):
        try:
            return max(0, int(widget.text().strip() or default))
        except ValueError:
            return default

    def is_nonnegative_int_text(self, value):
        return bool(re.fullmatch(r"\d+", value.strip() or "0"))

    def is_valid_hhmm(self, value):
        if not re.fullmatch(r"\d{2}:\d{2}", value or ""):
            return False
        hora, minuto = [int(parte) for parte in value.split(":")]
        return 0 <= hora <= 23 and 0 <= minuto <= 59

    def validate_settings(self):
        self.clear_field_errors()
        s = self.collect_settings()
        errors = []
        invalid_widgets = []

        if s.modo_origem_ofertas == "whatsapp" and not s.grupo_origem:
            errors.append("Informe o grupo de origem.")
            invalid_widgets.append(self.grupo_origem_input)

        if not s.normalized_destinos() and not s.normalized_canais_destino():
            errors.append("Informe pelo menos um grupo ou canal de destino.")
            invalid_widgets.append(self.destinos_input)
            invalid_widgets.append(self.canais_destino_input)

        if not s.shopee_api_app_id:
            errors.append("Informe o AppID da API Shopee.")
            invalid_widgets.append(self.app_id_input)

        if not s.shopee_api_secret:
            errors.append("Informe o Secret da API Shopee.")
            invalid_widgets.append(self.secret_input)

        sub_id_errors = self.validate_sub_ids(s.shopee_api_sub_ids)
        if sub_id_errors:
            errors.extend(sub_id_errors)
            invalid_widgets.append(self.sub_ids_input)

        if s.aguardar_previa_link and s.espera_minima_previa_link_ms > s.timeout_previa_link_ms:
            errors.append("A espera mínima da prévia não pode ser maior que o tempo máximo da prévia.")
            invalid_widgets.extend([self.espera_minima_input, self.timeout_previa_input])

        if s.agendamento_ativo:
            if not self.is_valid_hhmm(s.horario_inicio):
                errors.append("Informe o horário inicial no formato HH:MM.")
                invalid_widgets.append(self.horario_inicio_input)
            if not self.is_valid_hhmm(s.horario_fim):
                errors.append("Informe o horário final no formato HH:MM.")
                invalid_widgets.append(self.horario_fim_input)

        if s.ia_revisao_ativa and not s.ia_gemini_api_key:
            errors.append("Informe a chave Gemini ou desative a revisão por IA.")
            invalid_widgets.append(self.ia_api_key_input)

        if s.ia_revisao_ativa and s.ia_limite_diario_revisoes <= 0:
            errors.append("Informe um limite diário de revisões por IA maior que zero.")
            invalid_widgets.append(self.ia_limite_diario_input)

        integer_fields = [
            (self.shopee_ofertas_comissao_input, "Comissão mínima"),
            (self.shopee_ofertas_desconto_input, "Desconto mínimo"),
            (self.shopee_ofertas_vendas_input, "Vendas mínimas"),
        ]
        for widget, label in integer_fields:
            if widget.text().strip() and not self.is_nonnegative_int_text(widget.text()):
                errors.append(f"{label} deve ser um número inteiro.")
                invalid_widgets.append(widget)

        self.mark_field_errors(invalid_widgets)
        return errors

    def validate_api_settings(self):
        self.clear_field_errors()
        s = self.collect_settings()
        errors = []
        invalid_widgets = []

        if not s.shopee_api_app_id:
            errors.append("Informe o AppID da API Shopee.")
            invalid_widgets.append(self.app_id_input)

        if not s.shopee_api_secret:
            errors.append("Informe o Secret da API Shopee.")
            invalid_widgets.append(self.secret_input)

        sub_id_errors = self.validate_sub_ids(s.shopee_api_sub_ids)
        if sub_id_errors:
            errors.extend(sub_id_errors)
            invalid_widgets.append(self.sub_ids_input)

        self.mark_field_errors(invalid_widgets)
        return errors

    def validate_group_settings(self):
        self.clear_field_errors()
        s = self.collect_settings()
        errors = []
        invalid_widgets = []

        if not s.grupo_origem:
            errors.append("Informe o grupo de origem.")
            invalid_widgets.append(self.grupo_origem_input)

        if not s.normalized_destinos() and not s.normalized_canais_destino():
            errors.append("Informe pelo menos um grupo ou canal de destino.")
            invalid_widgets.append(self.destinos_input)
            invalid_widgets.append(self.canais_destino_input)

        self.mark_field_errors(invalid_widgets)
        return errors

    def validate_sub_ids(self, sub_ids):
        if len(sub_ids) > 5:
            return ["Informe no máximo 5 Sub IDs."]

        invalidos = [
            sub_id
            for sub_id in sub_ids
            if not re.fullmatch(r"[A-Za-z0-9]{1,100}", sub_id)
        ]

        if not invalidos:
            return []

        return [
            "Sub IDs podem usar apenas letras sem acento e números, sem espaços ou símbolos. "
            f"Revise: {', '.join(invalidos)}"
        ]

    def configuration_missing_items(self):
        s = self.collect_settings()
        missing = []
        if s.modo_origem_ofertas == "whatsapp" and not s.grupo_origem:
            missing.append("grupo de origem")
        if not s.normalized_destinos() and not s.normalized_canais_destino():
            missing.append("grupo ou canal de destino")
        if not s.shopee_api_app_id:
            missing.append("AppID Shopee")
        if not s.shopee_api_secret:
            missing.append("Secret Shopee")
        return missing

    def abrir_shopee_para_login(self):
        if self.runner and self.runner.is_running():
            QMessageBox.warning(
                self,
                "Bot em execução",
                "Pare o bot antes de abrir o navegador para login na Shopee "
                "(os dois usam o mesmo perfil de navegador e não podem abrir ao mesmo tempo).",
            )
            return

        self.save_fields(show_message=False)

        self.shopee_login_button.setEnabled(False)
        self.signals.log.emit("Abrindo navegador para login na Shopee...")
        threading.Thread(target=self._run_shopee_login, daemon=True).start()

    def _run_shopee_login(self):
        play = None
        browser = None
        try:
            play = sync_playwright().start()
            # Perfil dedicado do bot para a Shopee (o MESMO que o bot usa ao
            # rodar). Assim o login feito aqui vale para os acessos do bot.
            profile_dir = str(self.resolve_app_path("perfil_shopee_bot"))
            try:
                browser = play.chromium.launch_persistent_context(
                    **montar_kwargs_launch(
                        user_data_dir=profile_dir,
                        channel="chrome",
                        headless_visual=False,
                    )
                )
                aplicar_anti_deteccao_no_contexto(browser)
                self.signals.log.emit("Usando o Google Chrome instalado no PC (perfil dedicado do bot).")
            except Exception as e:
                self.signals.log.emit(f"Google Chrome de verdade falhou ao abrir: {e}")
                browser = play.chromium.launch_persistent_context(
                    **montar_kwargs_launch(
                        user_data_dir=profile_dir,
                        channel=None,
                        headless_visual=False,
                    )
                )
                aplicar_anti_deteccao_no_contexto(browser)
                self.signals.log.emit(
                    "Usando o navegador embutido do Playwright (pode ter mais chance de pedir captcha)."
                )

            pagina = browser.new_page()
            # Começa pela HOME, não pelo link da coleção. A rota de coleção é
            # "sensível" e dispara o antibot da Shopee quando acessada a frio /
            # sem sessão - era um dos motivos de cair no captcha logo de cara.
            # A ideia aqui é você LOGAR primeiro (na home / tela de login) e só
            # depois navegar para produtos/coleções já autenticado.
            pagina.goto("https://shopee.com.br/", wait_until="domcontentloaded", timeout=30000)
            pagina.bring_to_front()

            # O perfil pode abrir uma aba em branco por conta própria ao
            # iniciar; fechamos essas abas extras para não confundir com a
            # aba de verdade que acabamos de abrir.
            for outra_pagina in list(browser.pages):
                if outra_pagina is pagina:
                    continue
                try:
                    if outra_pagina.url in ("about:blank", ""):
                        outra_pagina.close()
                except Exception:
                    pass

            self.signals.log.emit(
                "Faça seu login na Shopee na janela que abriu, com calma. Se aparecer "
                "captcha, resolva ele normalmente (agora ele deve carregar, porque é o "
                "mesmo tipo de navegador que você usa). Navegue até uma página de produto "
                "ou coleção para confirmar que entrou. Quando terminar, FECHE a janela - o "
                "bot guarda esse login no perfil dele e reusa nos próximos acessos à Shopee. "
                "Só inicie o bot depois de fechar a janela."
            )
            pagina.wait_for_event("close", timeout=0)
            self.signals.log.emit("Janela de login fechada. Já pode usar o bot normalmente.")
        except Exception as e:
            self.signals.log.emit(f"Falha ao abrir navegador para login na Shopee: {self.friendly_error(e)}")
        finally:
            try:
                if browser:
                    browser.close()
            except Exception:
                pass
            try:
                if play:
                    play.stop()
            except Exception:
                pass
            self.signals.shopee_login_button_enable.emit()

    def atualizar_campos_tipo_busca_shopee(self, *_):
        """
        Faz o "Tipo de busca" valer de verdade: só o campo correspondente ao
        tipo escolhido fica habilitado. Os outros ficam desabilitados
        visualmente (o texto/seleção continua salvo, mas o backend também
        ignora esses campos quando não correspondem ao tipo escolhido - ver
        montar_parametros_ofertas e oferta_passa_filtros em afiliados.py).
        """

        tipo = self.shopee_ofertas_tipo_input.currentData() or "geral"
        self.shopee_ofertas_palavra_input.setEnabled(tipo == "palavra_chave")
        self.shopee_ofertas_shop_id_input.setEnabled(tipo == "loja")
        self.shopee_ofertas_categoria_input.setEnabled(tipo == "categoria")

    def update_configuration_status(self):
        if self.runner and self.runner.is_running():
            return

        missing = self.configuration_missing_items()
        if missing:
            self.set_status("Configuração incompleta.")
        elif self.status_label.text() in ("Configuração incompleta.", "Pronto para iniciar"):
            self.set_status("Pronto para iniciar")

    def show_first_setup_guide(self):
        self.config_tabs.setCurrentIndex(0)
        QMessageBox.information(
            self,
            "Primeira configuração",
            "Para configurar o Bot.ee pela primeira vez:\n\n"
            "1. Na aba Execução, escolha o modo de operação:\n"
            "   - Grupo do WhatsApp: o bot lê ofertas de um grupo origem.\n"
            "   - Buscar ofertas na Shopee: o bot busca ofertas direto pela API.\n"
            "2. Ainda em Execução, informe um ou mais grupos de destino.\n"
            "3. Na aba API, preencha Shopee AppID e Shopee Secret.\n"
            "4. Use Sub IDs apenas se quiser marcar campanhas nos relatórios.\n"
            "5. Clique em Testar API Shopee.\n"
            "6. Se usar Grupo do WhatsApp, abra a aba Modo Grupo, informe o grupo de origem e clique em Testar grupos.\n"
            "7. Se usar Modo Shopee, abra a aba Modo Shopee e ajuste os filtros desejados.\n"
            "8. Salve as configurações e inicie o bot.",
        )

    def clear_field_errors(self):
        self.mark_field_errors(
            [
                self.grupo_origem_input,
                self.destinos_input,
                self.canais_destino_input,
                self.app_id_input,
                self.secret_input,
                self.sub_ids_input,
                self.intervalo_input,
                self.timeout_previa_input,
                self.espera_minima_input,
                self.horario_inicio_input,
                self.horario_fim_input,
                self.recuperacao_max_tentativas_input,
                self.recuperacao_intervalo_input,
                self.ia_api_key_input,
                self.ia_limite_diario_input,
                self.ia_timeout_input,
                self.modo_origem_ofertas_input,
                self.shopee_ofertas_colecao_input,
                self.shopee_ofertas_cupom_input,
                self.shopee_ofertas_tipo_input,
                self.shopee_ofertas_palavra_input,
                self.shopee_ofertas_shop_id_input,
                self.shopee_ofertas_categoria_input,
                self.shopee_ofertas_ordenacao_input,
                self.shopee_ofertas_comissao_input,
                self.shopee_ofertas_desconto_input,
                self.shopee_ofertas_preco_min_input,
                self.shopee_ofertas_preco_max_input,
                self.shopee_ofertas_vendas_input,
                self.shopee_ofertas_avaliacao_input,
                self.shopee_ofertas_intervalo_input,
            ],
            invalid=False,
        )

    def mark_field_errors(self, widgets, invalid=True):
        for widget in widgets:
            if isinstance(widget, SecondsInput):
                widget.set_invalid(invalid)
                continue

            widget.setProperty("invalid", invalid)
            widget.style().unpolish(widget)
            widget.style().polish(widget)

    def create_settings_backup(self, show_log=False):
        try:
            backup_path = backup_user_config()
            if backup_path and show_log:
                self.add_log(f"Backup das configurações criado: {backup_path}")
            return backup_path
        except Exception as e:
            if show_log:
                self.add_log(f"Não consegui criar backup das configurações: {self.friendly_error(e)}")
            return None

    def save_fields(self, show_message=True):
        self.create_settings_backup(show_log=show_message)
        self.collect_settings().save()
        self.update_configuration_status()
        if show_message:
            self.add_log("Configurações salvas.")

    def start_bot(self):
        errors = self.validate_settings()
        if errors:
            message = "Corrija os campos abaixo antes de iniciar:\n\n"
            message += "\n".join(f"- {error}" for error in errors)
            QMessageBox.warning(self, "Configurações incompletas", message)
            self.set_status("Corrija as configurações para iniciar.")
            return

        self.save_fields(show_message=False)
        self.add_log("Configurações salvas automaticamente antes de iniciar.")
        if self.settings.modo_simulacao:
            self.add_log("Modo simulação ativo: o bot vai converter e formatar, mas não vai enviar ofertas.")
        self.runner = BotRunner(
            self.settings,
            on_log=self.signals.log.emit,
            on_status=self.signals.status.emit,
            on_stats=self.signals.stats.emit,
        )
        self.runner.start()
        self.history_timer.start()
        self.start_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.set_form_enabled(False)

    def test_api_shopee(self):
        errors = self.validate_api_settings()
        api_errors = [
            error
            for error in errors
            if "grupo" not in error.lower()
        ]
        if api_errors:
            message = "Corrija os campos abaixo antes de testar a API:\n\n"
            message += "\n".join(f"- {error}" for error in api_errors)
            QMessageBox.warning(self, "Configurações incompletas", message)
            return

        self.save_fields(show_message=False)
        self.start_test("Testando API Shopee...")
        threading.Thread(target=self._run_api_test, daemon=True).start()

    def _run_api_test(self):
        try:
            self.signals.log.emit("Testando API Shopee com um link de exemplo...")
            conversor = ConversorAfiliados(app_config=self.settings)
            link = conversor.converter_link(TESTE_API_SHOPEE_URL)
            self.settings.api_validada_em = datetime.now().isoformat(timespec="minutes")
            self.settings.save()
            self.signals.log.emit("API Shopee respondeu corretamente.")
            self.signals.log.emit(f"Link de teste gerado: {link}")
            self.signals.status.emit("API Shopee validada.")
        except Exception as e:
            self.settings.api_validada_em = ""
            self.settings.save()
            self.signals.log.emit(f"Falha no teste da API Shopee: {self.friendly_error(e)}")
            self.signals.status.emit("Teste da API Shopee falhou.")
        finally:
            self.signals.test_finished.emit()

    def test_shopee_offers(self):
        errors = self.validate_api_settings()
        if errors:
            message = "Corrija os campos abaixo antes de testar a busca Shopee:\n\n"
            message += "\n".join(f"- {error}" for error in errors)
            QMessageBox.warning(self, "Configurações incompletas", message)
            return

        offer_errors = []
        offer_widgets = []
        integer_fields = [
            (self.shopee_ofertas_comissao_input, "Comissão mínima"),
            (self.shopee_ofertas_desconto_input, "Desconto mínimo"),
            (self.shopee_ofertas_vendas_input, "Vendas mínimas"),
        ]
        for widget, label in integer_fields:
            if widget.text().strip() and not self.is_nonnegative_int_text(widget.text()):
                offer_errors.append(f"{label} deve ser um número inteiro.")
                offer_widgets.append(widget)

        if offer_errors:
            self.mark_field_errors(offer_widgets)
            message = "Corrija os filtros abaixo antes de testar a busca Shopee:\n\n"
            message += "\n".join(f"- {error}" for error in offer_errors)
            QMessageBox.warning(self, "Filtros inválidos", message)
            return

        self.save_fields(show_message=False)
        self.update_shopee_offers_status("Buscando ofertas na Shopee...")
        self.start_test("Testando busca Shopee...")
        threading.Thread(target=self._run_shopee_offers_test, daemon=True).start()

    def _run_shopee_offers_test(self):
        try:
            self.signals.log.emit("Buscando ofertas Shopee pela API...")
            conversor = ConversorAfiliados(app_config=self.settings)
            resultado = conversor.buscar_ofertas_shopee(self.settings)
            ofertas = resultado.get("ofertas") or []
            total_recebido = resultado.get("total_recebido") or 0
            total_filtrado = resultado.get("total_filtrado") or 0
            diagnostico_colecao = self.format_shopee_collection_diagnostic(resultado)

            if not ofertas:
                linhas = [
                    "Busca concluída, mas nenhuma oferta passou pelos filtros atuais.",
                    "",
                ]
                if diagnostico_colecao:
                    linhas.extend([diagnostico_colecao, ""])
                linhas.extend([
                    f"Recebidas pela API: {total_recebido}",
                    f"Após filtros: {total_filtrado}",
                    "",
                    "Tente reduzir comissão mínima, vendas mínimas, avaliação ou filtros de preço.",
                ])
                mensagem = "\n".join(linhas)
            else:
                linhas = [
                    "Busca Shopee concluída com sucesso.",
                    "",
                ]
                if diagnostico_colecao:
                    linhas.extend([diagnostico_colecao, ""])
                linhas.extend([
                    f"Recebidas pela API: {total_recebido}",
                    f"Após filtros: {total_filtrado}",
                    "",
                    "Primeiras ofertas encontradas:",
                ])
                for index, oferta in enumerate(ofertas[:5], start=1):
                    preco = self.format_price_preview(oferta)
                    comissao = oferta.get("comissao") or 0
                    desconto = oferta.get("desconto") or 0
                    linhas.append(
                        f"{index}. {oferta.get('nome') or 'Produto sem nome'}\n"
                        f"   {preco} | Comissão: {comissao:g}% | Desconto: {desconto:g}%\n"
                        f"   Link: {oferta.get('link') or 'não informado'}"
                    )
                mensagem = "\n".join(linhas)

            self.signals.shopee_offers_status.emit(mensagem)
            self.signals.log.emit(f"Busca Shopee concluída: {len(ofertas)} oferta(s) pronta(s) para uso.")
            if diagnostico_colecao:
                self.signals.log.emit(diagnostico_colecao.replace("\n", " | "))
            self.signals.status.emit("Busca Shopee testada.")
        except Exception as e:
            mensagem = (
                "Não consegui buscar ofertas pela API Shopee.\n\n"
                f"Detalhe: {self.friendly_error(e)}"
            )
            self.signals.shopee_offers_status.emit(mensagem)
            self.signals.log.emit(f"Falha na busca Shopee: {self.friendly_error(e)}")
            self.signals.status.emit("Busca Shopee falhou.")
        finally:
            self.signals.test_finished.emit()

    def format_shopee_collection_diagnostic(self, resultado):
        link_colecao = (resultado.get("colecao_link") or "").strip()
        if not link_colecao:
            return ""

        colecao_id = (resultado.get("colecao_id") or "").strip()
        parametro_aplicado = bool(resultado.get("colecao_parametro_aplicado"))
        parametros = (resultado.get("parametros_usados") or "").strip()

        linhas = [
            "Diagnóstico do link de coleção:",
            f"Link informado: {link_colecao}",
            f"ID detectado: {colecao_id or 'não detectado'}",
            f"Parâmetro de coleção aplicado pela API: {'sim' if parametro_aplicado else 'não'}",
        ]
        if parametros:
            linhas.append(f"Parâmetros aceitos na chamada: {parametros}")
        if not parametro_aplicado:
            linhas.append(
                "Aviso: a API não confirmou uso do filtro de coleção; "
                "a busca pode ter retornado ofertas gerais com os demais filtros."
            )
        return "\n".join(linhas)

    def format_price_preview(self, oferta):
        preco_pix = float(oferta.get("preco_pix") or 0)
        preco_parcelado = float(oferta.get("preco_parcelado") or 0)
        preco = float(oferta.get("preco") or 0)
        if preco_pix and preco_parcelado and abs(preco_pix - preco_parcelado) >= 0.01:
            return f"Pix: {self.format_money(preco_pix)} | Parcelado: {self.format_money(preco_parcelado)}"
        if preco_pix:
            return f"Pix: {self.format_money(preco_pix)}"
        return f"Preço: {self.format_money(preco)}"

    def test_whatsapp_groups(self):
        errors = self.validate_group_settings()
        group_errors = [
            error
            for error in errors
            if "grupo" in error.lower()
        ]
        if group_errors:
            message = "Corrija os campos abaixo antes de testar os grupos:\n\n"
            message += "\n".join(f"- {error}" for error in group_errors)
            QMessageBox.warning(self, "Configurações incompletas", message)
            return

        self.save_fields(show_message=False)
        self.start_test("Testando grupos no WhatsApp...")
        threading.Thread(target=self._run_groups_test, daemon=True).start()

    def _run_groups_test(self):
        play = None
        zap = None
        try:
            self.signals.log.emit("Abrindo WhatsApp para testar os grupos...")
            self.settings.profile_dir = str(self.resolve_app_path(self.settings.profile_dir))
            play = sync_playwright().start()
            zap = WhatsApp(play, headless=self.settings.headless, app_config=self.settings)

            self.signals.log.emit(f"Testando grupo de origem: {self.settings.grupo_origem}")
            zap.preparar_monitoramento_origem(self.settings.grupo_origem)
            self.signals.log.emit("Grupo de origem encontrado.")

            for grupo, tipo_destino in self.settings.normalized_destinos_com_tipo():
                rotulo_destino = "canal" if tipo_destino == "canal" else "grupo"
                self.signals.log.emit(f"Testando {rotulo_destino} de destino: {grupo}")
                zap.abrir_destino_para_envio(grupo, tipo=tipo_destino, timeout_ms=20000)
                self.signals.log.emit(f"{rotulo_destino.capitalize()} de destino encontrado: {grupo}")

            self.settings.grupos_validados_em = datetime.now().isoformat(timespec="minutes")
            self.settings.save()
            self.signals.status.emit("Grupos do WhatsApp validados.")
            self.signals.log.emit("Teste de grupos concluído com sucesso.")
        except Exception as e:
            self.settings.grupos_validados_em = ""
            self.settings.save()
            self.signals.log.emit(f"Falha no teste de grupos: {self.friendly_error(e)}")
            self.signals.status.emit("Teste de grupos falhou.")
        finally:
            try:
                if zap:
                    zap.fechar()
            except Exception:
                self.signals.log.emit("Não consegui fechar a janela do WhatsApp automaticamente.")
            try:
                if play:
                    play.stop()
            except Exception:
                pass
            self.signals.test_finished.emit()

    def test_ai_gemini(self):
        self.clear_field_errors()
        self.collect_settings()
        errors = []
        invalid_widgets = []

        if not self.settings.ia_gemini_api_key:
            errors.append("Informe a chave Gemini antes de testar a IA.")
            invalid_widgets.append(self.ia_api_key_input)

        if self.settings.ia_limite_diario_revisoes <= 0:
            errors.append("Informe um limite diário de revisões maior que zero.")
            invalid_widgets.append(self.ia_limite_diario_input)

        if errors:
            self.mark_field_errors(invalid_widgets)
            message = "Corrija os campos abaixo antes de testar a IA:\n\n"
            message += "\n".join(f"- {error}" for error in errors)
            QMessageBox.warning(self, "Configurações incompletas", message)
            return

        self.save_fields(show_message=False)
        self.start_test("Testando IA Gemini...")
        threading.Thread(target=self._run_ai_test, daemon=True).start()

    def _run_ai_test(self):
        texto_teste = (
            "OFERTA TESTE BOT.EE\n\n"
            "Mini processador elétrico 250ml\n"
            "Ideal para decorar seu ambiente com charme e elegância.\n\n"
            "De: R$ 79,90\n"
            "Por: R$ 39,90\n\n"
            "Compre aqui:\n"
            "https://s.shopee.com.br/teste"
        )
        ia_ativa_original = self.settings.ia_revisao_ativa
        revisoes_hoje_original = self.settings.ia_revisoes_hoje
        status_original = self.settings.ia_status_mensagem

        try:
            self.settings.ia_revisao_ativa = True
            revisor = RevisorIA(self.settings, self.signals.log.emit)
            self.signals.log.emit("Testando IA Gemini com uma oferta de exemplo...")
            texto_revisado = revisor.revisar_texto(texto_teste, "https://s.shopee.com.br/teste")

            if texto_revisado == texto_teste:
                mensagem = (
                    "IA Gemini respondeu, mas o texto original foi mantido pelas travas de segurança. "
                    "Confira chave, modelo e modo de revisão."
                )
                self.settings.ia_status_mensagem = mensagem
                self.signals.log.emit(mensagem)
                self.signals.status.emit("Teste da IA concluído com aviso.")
            else:
                mensagem = "IA Gemini validada com sucesso. Exemplo revisado:\n\n" + texto_revisado
                self.settings.ia_status_mensagem = mensagem
                self.signals.log.emit("IA Gemini respondeu corretamente.")
                self.signals.status.emit("IA Gemini validada.")
        except Exception as e:
            mensagem = f"Falha no teste da IA Gemini: {self.friendly_error(e)}"
            self.settings.ia_status_mensagem = mensagem
            self.signals.log.emit(mensagem)
            self.signals.status.emit("Teste da IA Gemini falhou.")
        finally:
            self.settings.ia_revisao_ativa = ia_ativa_original
            self.settings.ia_revisoes_hoje = revisoes_hoje_original
            if not self.settings.ia_status_mensagem:
                self.settings.ia_status_mensagem = status_original
            self.settings.save()
            self.signals.test_finished.emit()

    def start_test(self, status):
        self.set_status(status)
        self.start_button.setEnabled(False)
        self.stop_button.setEnabled(False)
        self.set_form_enabled(False)
        self.test_api_button.setEnabled(False)
        self.test_groups_button.setEnabled(False)
        self.test_ai_button.setEnabled(False)
        self.test_shopee_offers_button.setEnabled(False)
        self.save_button.setEnabled(False)

    def finish_test(self):
        if not (self.runner and self.runner.is_running()):
            self.start_button.setEnabled(True)
            self.stop_button.setEnabled(False)
            self.set_form_enabled(True)

        self.update_validation_badges()
        self.ia_status_box.setPlainText(self.settings.ia_status_mensagem or "")

    def update_shopee_offers_status(self, message):
        if hasattr(self, "shopee_ofertas_status_box"):
            self.shopee_ofertas_status_box.setPlainText(message)

    def change_visual_theme(self, *_):
        if hasattr(self, "tema_visual_input"):
            self.settings.tema_visual = self.tema_visual_input.currentData() or "padrao"
        self.apply_styles()
        self.apply_log_filters()

    def update_validation_badges(self):
        self.api_status_label.setText(self.validation_badge_text("API", self.settings.api_validada_em))
        self.api_status_label.setProperty("valid", bool(self.settings.api_validada_em))
        self.groups_status_label.setText(
            self.validation_badge_text("Grupos", self.settings.grupos_validados_em)
        )
        self.groups_status_label.setProperty("valid", bool(self.settings.grupos_validados_em))

        for label in (self.api_status_label, self.groups_status_label):
            label.style().unpolish(label)
            label.style().polish(label)

    def validation_badge_text(self, name, value):
        if not value:
            return f"{name}: não validado"

        try:
            checked_at = datetime.fromisoformat(value)
            return f"{name}: validado {checked_at.strftime('%d/%m %H:%M')}"
        except ValueError:
            return f"{name}: validado"

    def resolve_app_path(self, value):
        path = Path(value)
        if path.is_absolute():
            return path
        return APP_DIR / path

    def friendly_error(self, error):
        text = str(error)
        if "invalid sub id" in text.lower() or "Sub ID" in text:
            return "Sub ID inválido. Use apenas letras e números, ou deixe o campo vazio."
        if "Credential" in text or "Signature" in text or "unauthorized" in text.lower():
            return "AppID ou Secret da API Shopee parecem incorretos."
        if "Não consegui abrir o grupo" in text or "Não consegui confirmar" in text:
            return text
        if ("Timeout" in text or "timeout" in text) and (
            "whatsapp" in text.lower() or "chat-list" in text.lower() or "grupo" in text.lower()
        ):
            return "Tempo esgotado. Verifique conexão, login no WhatsApp ou nome do grupo."
        return text

    def stop_bot(self):
        if self.runner:
            self.runner.stop()
        self.stop_button.setEnabled(False)
        self.start_button.setEnabled(False)
        self.set_form_enabled(False)

    def closeEvent(self, event):
        if self.runner and self.runner.is_running():
            resposta = QMessageBox.question(
                self,
                "Bot em execução",
                "O bot ainda está rodando. Deseja parar o bot e fechar o aplicativo?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if resposta != QMessageBox.Yes:
                event.ignore()
                return

            self.stop_bot()

        event.accept()

    def set_form_enabled(self, enabled):
        for widget in (
            self.grupo_origem_input,
            self.destinos_input,
            self.canais_destino_input,
            self.app_id_input,
            self.secret_input,
            self.sub_ids_input,
            self.intervalo_input,
            self.timeout_previa_input,
            self.espera_minima_input,
            self.processar_historico_input,
            self.ignorar_links_enviados_input,
            self.aguardar_previa_input,
            self.modo_simulacao_input,
            self.headless_input,
            self.agendamento_ativo_input,
            self.horario_inicio_input,
            self.horario_fim_input,
            self.recuperacao_automatica_input,
            self.recuperacao_max_tentativas_input,
            self.recuperacao_intervalo_input,
            self.ia_revisao_input,
            self.ia_pausar_limite_input,
            self.ia_retomar_dia_input,
            self.ia_api_key_input,
            self.ia_modelo_input,
            self.ia_tipo_revisao_input,
            self.ia_limite_diario_input,
            self.ia_timeout_input,
            self.ia_status_box,
            self.test_ai_button,
            self.modo_origem_ofertas_input,
            self.shopee_ofertas_colecao_input,
            self.shopee_ofertas_cupom_input,
            self.shopee_ofertas_tipo_input,
            self.shopee_ofertas_palavra_input,
            self.shopee_ofertas_shop_id_input,
            self.shopee_ofertas_categoria_input,
            self.shopee_ofertas_ordenacao_input,
            self.shopee_ofertas_comissao_input,
            self.shopee_ofertas_desconto_input,
            self.shopee_ofertas_preco_min_input,
            self.shopee_ofertas_preco_max_input,
            self.shopee_ofertas_vendas_input,
            self.shopee_ofertas_avaliacao_input,
            self.shopee_ofertas_intervalo_input,
            self.shopee_ofertas_nao_repetir_input,
            self.shopee_ofertas_status_box,
            self.test_shopee_offers_button,
            self.tema_visual_input,
            self.test_api_button,
            self.test_groups_button,
            self.first_setup_button,
            self.export_config_button,
            self.import_config_button,
            self.restore_backup_button,
            self.diagnostic_button,
            self.clear_logs_button,
            self.open_logs_button,
            self.copy_logs_button,
            self.export_logs_button,
            self.refresh_indicators_button,
            self.clear_history_button,
            self.save_button,
        ):
            widget.setEnabled(enabled)

        self.refresh_history_button.setEnabled(True)

    def add_log(self, message):
        now = datetime.now().strftime("%H:%M:%S")
        line = f"[{now}] {message}"
        self.log_lines.append(line)
        self.apply_log_filters()
        self.write_log_file(line)

    def classify_log_line(self, line):
        text = line.lower()
        if any(term in text for term in ("erro", "falha", "não consegui", "nao consegui", "traceback", "exception")):
            return "error"
        if any(term in text for term in ("atenção", "atencao", "aviso", "corrija", "captcha", "não validado", "nao validado")):
            return "warning"
        if any(term in text for term in ("ok -", "concluído", "concluido", "validado", "enviado", "copiado", "exportado", "salvo", "pronto")):
            return "success"
        return "info"

    def apply_log_filters(self, *_):
        if not hasattr(self, "logs"):
            return

        search = self.logs_search_input.text().strip().lower()
        log_type = self.logs_type_input.currentData() or "all"
        filtered = []

        for line in self.log_lines:
            category = self.classify_log_line(line)
            if log_type != "all" and category != log_type:
                continue
            if search and search not in line.lower():
                continue
            filtered.append(line)

        self.filtered_log_lines = filtered
        self.render_logs(filtered)

    def render_logs(self, lines):
        theme = getattr(self.settings, "tema_visual", "padrao")
        if theme == "noturno":
            colors = {
                "info": ("transparent", "#ccfbf1"),
                "success": ("#052e2b", "#99f6e4"),
                "warning": ("#422006", "#fde68a"),
                "error": ("#3f1212", "#fecaca"),
            }
        else:
            colors = {
                "info": ("transparent", "#1e293b"),
                "success": ("#ecfdf5", "#065f46"),
                "warning": ("#fffbeb", "#92400e"),
                "error": ("#fff1f2", "#991b1b"),
            }

        parts = []
        for line in lines:
            category = self.classify_log_line(line)
            background, color = colors.get(category, colors["info"])
            parts.append(
                "<div style=\""
                f"background:{background}; color:{color}; "
                "padding:2px 6px; margin:1px 0; white-space:pre-wrap;"
                f"\">{escape(line)}</div>"
            )

        self.logs.setHtml("\n".join(parts))
        self.logs.moveCursor(QTextCursor.MoveOperation.End)

        total = len(self.log_lines)
        shown = len(lines)
        if total == shown:
            self.logs_summary_label.setText(f"{shown} linha(s)")
        else:
            self.logs_summary_label.setText(f"{shown} de {total} linha(s)")

    def cleanup_old_logs(self):
        try:
            self.logs_dir.mkdir(parents=True, exist_ok=True)
            cutoff = datetime.now() - timedelta(days=LOG_RETENTION_DAYS)
            for log_file in self.logs_dir.glob("bot_*.log"):
                try:
                    modified_at = datetime.fromtimestamp(log_file.stat().st_mtime)
                    if modified_at < cutoff:
                        log_file.unlink()
                except OSError:
                    pass
        except OSError:
            pass

    def write_log_file(self, line):
        try:
            self.logs_dir.mkdir(parents=True, exist_ok=True)
            with self.log_file_path.open("a", encoding="utf-8") as file:
                file.write(line + "\n")
        except Exception:
            pass

    def clear_logs(self):
        self.log_lines.clear()
        self.filtered_log_lines.clear()
        self.logs.clear()
        self.logs_summary_label.setText("0 linhas")
        self.add_log("Logs da tela limpos. O arquivo de log foi mantido.")

    def copy_selected_logs(self):
        selected_text = self.logs.textCursor().selectedText().replace("\u2029", "\n").strip()
        if not selected_text:
            QMessageBox.information(self, "Copiar logs", "Nenhum trecho de log foi selecionado.")
            return

        QApplication.clipboard().setText(selected_text)
        QMessageBox.information(self, "Copiar logs", "Logs selecionados copiados.")

    def export_visible_logs(self):
        if not self.filtered_log_lines:
            QMessageBox.information(self, "Exportar logs", "Não há logs visíveis para exportar.")
            return

        default_name = f"logs-bot-ee-{datetime.now().strftime('%Y-%m-%d')}.txt"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar logs",
            str(APP_DIR / default_name),
            "Arquivo de texto (*.txt)",
        )
        if not file_path:
            return

        try:
            Path(file_path).write_text("\n".join(self.filtered_log_lines) + "\n", encoding="utf-8")
            self.add_log(f"Logs exportados para: {file_path}")
            QMessageBox.information(self, "Logs exportados", "Arquivo de logs gerado.")
        except Exception as e:
            QMessageBox.warning(self, "Falha ao exportar", f"Não consegui exportar os logs:\n\n{e}")

    def open_logs_folder(self):
        try:
            self.logs_dir.mkdir(parents=True, exist_ok=True)
            os.startfile(str(self.logs_dir))
        except Exception as e:
            self.add_log(f"Não consegui abrir a pasta de logs: {e}")

    def run_diagnostics(self):
        self.collect_settings()
        self.add_log("=" * 58)
        self.add_log("Diagnóstico do ambiente iniciado.")

        def log_check(ok, success_message, fail_message):
            prefix = "OK" if ok else "ATENÇÃO"
            message = success_message if ok else fail_message
            self.add_log(f"{prefix} - {message}")

        destinos = self.settings.normalized_destinos()
        canais = self.settings.normalized_canais_destino()
        sub_ids = self.settings.normalized_sub_ids()
        sub_id_errors = self.validate_sub_ids(sub_ids)

        log_check(
            bool(self.settings.grupo_origem),
            f"Grupo de origem informado: {self.settings.grupo_origem}",
            "Grupo de origem não informado.",
        )
        log_check(
            bool(destinos or canais),
            f"{len(destinos)} grupo(s) e {len(canais)} canal(is) de destino informado(s).",
            "Nenhum grupo ou canal de destino informado.",
        )
        log_check(
            bool(self.settings.shopee_api_app_id),
            "Shopee AppID informado.",
            "Shopee AppID não informado.",
        )
        log_check(
            bool(self.settings.shopee_api_secret),
            "Shopee Secret informado.",
            "Shopee Secret não informado.",
        )
        log_check(
            not sub_id_errors,
            "Sub IDs válidos.",
            sub_id_errors[0] if sub_id_errors else "Sub IDs inválidos.",
        )

        temp_log_file = self.logs_dir / ".diagnostico.tmp"
        try:
            self.logs_dir.mkdir(parents=True, exist_ok=True)
            temp_log_file.write_text("ok", encoding="utf-8")
            temp_log_file.unlink(missing_ok=True)
            log_check(True, f"Pasta de logs pronta: {self.logs_dir}", "")
            self.add_log(f"Logs antigos: limpeza automática após {LOG_RETENTION_DAYS} dias.")
        except Exception as e:
            log_check(False, "", f"Não consegui gravar na pasta de logs: {self.friendly_error(e)}")

        temp_backup_file = CONFIG_BACKUP_DIR / ".diagnostico.tmp"
        try:
            CONFIG_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
            temp_backup_file.write_text("ok", encoding="utf-8")
            temp_backup_file.unlink(missing_ok=True)
            log_check(True, f"Pasta de backups pronta: {CONFIG_BACKUP_DIR}", "")
        except Exception as e:
            log_check(False, "", f"Não consegui gravar na pasta de backups: {self.friendly_error(e)}")

        db = None
        try:
            db_path = self.resolve_app_path(self.settings.database_path)
            db_path.parent.mkdir(parents=True, exist_ok=True)
            db = Database(str(db_path))
            total_mensagens = db.quantidade()
            log_check(True, f"Banco de dados acessível ({total_mensagens} mensagem(ns) registradas).", "")
        except Exception as e:
            log_check(False, "", f"Banco de dados não pôde ser acessado: {self.friendly_error(e)}")
        finally:
            try:
                if db:
                    db.fechar()
            except Exception:
                pass

        try:
            profile_dir = self.resolve_app_path(self.settings.profile_dir)
            profile_dir.mkdir(parents=True, exist_ok=True)
            log_check(True, f"Pasta do perfil do WhatsApp pronta: {profile_dir}", "")
        except Exception as e:
            log_check(False, "", f"Não consegui preparar a pasta do WhatsApp: {self.friendly_error(e)}")

        log_check(
            bool(self.settings.api_validada_em),
            self.validation_badge_text("API", self.settings.api_validada_em),
            "API ainda não validada pelo botão Testar API Shopee.",
        )
        log_check(
            bool(self.settings.grupos_validados_em),
            self.validation_badge_text("Grupos", self.settings.grupos_validados_em),
            "Grupos ainda não validados pelo botão Testar grupos.",
        )

        self.add_log(f"Modo simulação: {'ativo' if self.settings.modo_simulacao else 'desativado'}.")
        self.add_log(f"Navegador minimizado: {'ativo' if self.settings.headless else 'desativado'}.")
        self.add_log("Diagnóstico do ambiente concluído.")
        self.add_log("=" * 58)
        self.set_status("Diagnóstico concluído. Veja os logs.")

    def export_settings(self):
        self.collect_settings()
        default_name = f"bot-ee-config-{datetime.now().strftime('%Y-%m-%d')}.json"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar configurações",
            str(APP_DIR / default_name),
            "Configurações Bot.ee (*.json)",
        )
        if not file_path:
            return

        try:
            self.settings.save(file_path)
            self.add_log(f"Configurações exportadas para: {file_path}")
            QMessageBox.information(self, "Configurações exportadas", "Arquivo de configurações salvo.")
        except Exception as e:
            QMessageBox.warning(self, "Falha ao exportar", f"Não consegui exportar as configurações:\n\n{e}")

    def import_settings(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Importar configurações",
            str(APP_DIR),
            "Configurações Bot.ee (*.json)",
        )
        if not file_path:
            return

        try:
            backup_path = self.create_settings_backup(show_log=False)
            imported = AppSettings.load(file_path)
            self.settings = imported
            self.load_fields()
            self.settings.save()
            self.apply_styles()
            if backup_path:
                self.add_log(f"Backup anterior preservado em: {backup_path}")
            self.add_log(f"Configurações importadas de: {file_path}")
            QMessageBox.information(self, "Configurações importadas", "As configurações foram carregadas.")
        except Exception as e:
            QMessageBox.warning(self, "Falha ao importar", f"Não consegui importar as configurações:\n\n{e}")

    def restore_settings_backup(self):
        CONFIG_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        backups = sorted(
            CONFIG_BACKUP_DIR.glob("config_usuario_*.json"),
            key=lambda item: item.stat().st_mtime,
            reverse=True,
        )
        if not backups:
            QMessageBox.information(
                self,
                "Nenhum backup encontrado",
                "Ainda não existe backup de configurações para restaurar.",
            )
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Restaurar backup de configurações",
            str(backups[0]),
            "Backups Bot.ee (*.json)",
        )
        if not file_path:
            return

        answer = QMessageBox.question(
            self,
            "Restaurar backup",
            "Restaurar este backup vai substituir as configurações atuais.\n\n"
            "Um backup da configuração atual será criado antes da restauração. Deseja continuar?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if answer != QMessageBox.Yes:
            return

        try:
            current_backup = self.create_settings_backup(show_log=False)
            restored = AppSettings.load(file_path)
            self.settings = restored
            self.load_fields()
            self.settings.save()
            self.apply_styles()
            if current_backup:
                self.add_log(f"Backup da configuração anterior criado em: {current_backup}")
            self.add_log(f"Backup restaurado: {file_path}")
            self.set_status("Backup restaurado.")
            QMessageBox.information(self, "Backup restaurado", "As configurações foram restauradas.")
        except Exception as e:
            QMessageBox.warning(self, "Falha ao restaurar", f"Não consegui restaurar o backup:\n\n{e}")

    def show_about(self):
        message = (
            f"Bot.ee v{APP_VERSION}\n\n"
            "Automatiza ofertas do WhatsApp com links afiliados da Shopee.\n\n"
            "Desenvolvido por Eduardo Carreão Freire"
        )
        QMessageBox.information(self, "Sobre o Bot.ee", message)

    def refresh_history(self):
        db = None
        try:
            db_path = self.resolve_app_path(self.settings.database_path)
            db = Database(str(db_path))
            self.history_rows = db.listar_links_enviados(1000)
            self.apply_history_filters()
        except Exception as e:
            self.add_log(f"Não consegui carregar o histórico: {e}")
        finally:
            if db:
                db.fechar()

    def update_indicator_period_fields(self, *_):
        period = self.indicators_period_input.currentData() or "yesterday"
        today = QDate.currentDate()

        if period == "yesterday":
            start = today.addDays(-1)
            end = today.addDays(-1)
        elif period == "7":
            start = today.addDays(-6)
            end = today
        elif period == "30":
            start = today.addDays(-29)
            end = today
        else:
            if not self.indicators_start_date.date().isValid():
                start = today.addDays(-6)
            else:
                start = self.indicators_start_date.date()
            if not self.indicators_end_date.date().isValid():
                end = today
            else:
                end = self.indicators_end_date.date()

        if period != "custom":
            self.indicators_start_date.setDate(start)
            self.indicators_end_date.setDate(end)

        custom = period == "custom"
        self.indicators_start_date.setEnabled(custom)
        self.indicators_end_date.setEnabled(custom)

    def indicator_period_range(self):
        start_date = self.indicators_start_date.date().toPython()
        end_date = self.indicators_end_date.date().toPython()
        start = datetime.combine(start_date, datetime.min.time())
        end = datetime.combine(end_date, datetime.max.time())
        return start, end

    def refresh_indicators(self):
        errors = self.validate_api_settings()
        api_errors = [
            error
            for error in errors
            if "grupo" not in error.lower()
        ]
        if api_errors:
            message = "Corrija os campos abaixo antes de buscar os indicadores:\n\n"
            message += "\n".join(f"- {error}" for error in api_errors)
            QMessageBox.warning(self, "Configurações incompletas", message)
            return

        start, end = self.indicator_period_range()
        if start > end:
            QMessageBox.warning(
                self,
                "Período inválido",
                "A data inicial não pode ser maior que a data final.",
            )
            return

        self.save_fields(show_message=False)
        self.refresh_indicators_button.setEnabled(False)
        self.set_status("Buscando indicadores na Shopee...")
        self.add_log(
            "Buscando indicadores Shopee de "
            f"{start.strftime('%d/%m/%Y')} até {end.strftime('%d/%m/%Y')}..."
        )
        threading.Thread(
            target=self._run_indicators_refresh,
            args=(start, end),
            daemon=True,
        ).start()

    def _run_indicators_refresh(self, start, end):
        try:
            conversor = ConversorAfiliados(app_config=self.settings)
            report = conversor.obter_indicadores_por_sub_id(start, end)
            report["inicio"] = start
            report["fim"] = end
            self.signals.indicators_loaded.emit(report)
            self.signals.log.emit("Indicadores Shopee atualizados.")
            debug = report.get("debug") or {}
            if debug:
                self.signals.log.emit(
                    "Indicadores Shopee: "
                    f"{debug.get('conversoes', 0)} conversão(ões) retornada(s)."
                )
                campos = debug.get("campos_conversao") or []
                if campos:
                    self.signals.log.emit(
                        "Campos do relatório de conversões: " + ", ".join(campos)
                    )
            resumo_status = report.get("resumo_status_pedidos") or {}
            if resumo_status:
                partes_status = ", ".join(
                    f"{status}: {quantidade}" for status, quantidade in resumo_status.items()
                )
                self.signals.log.emit(f"Pedidos por status: {partes_status}")
            self.signals.status.emit("Indicadores atualizados.")
        except Exception as e:
            self.signals.log.emit(f"Falha ao buscar indicadores Shopee: {self.friendly_error(e)}")
            self.signals.status.emit("Falha ao buscar indicadores.")
            self.signals.indicators_loaded.emit({"erro": str(e), "linhas": [], "totais": {}})

    def render_indicators_report(self, report):
        self.refresh_indicators_button.setEnabled(True)
        if report.get("erro"):
            QMessageBox.warning(
                self,
                "Falha ao buscar indicadores",
                "Não consegui buscar os indicadores pela API da Shopee.\n\n"
                "Veja os logs para conferir o detalhe retornado pela API.",
            )

        rows = report.get("linhas") or []
        totals = report.get("totais") or {}
        self.indicators_rows = rows
        self.indicators_sales_label.setText(str(int(totals.get("vendas") or 0)))
        self.indicators_commission_label.setText(self.format_money(totals.get("comissao") or 0))

        self.indicators_table.setRowCount(len(rows))
        for row_index, row in enumerate(rows):
            values = [
                row.get("sub_id") or "",
                str(int(row.get("vendas") or 0)),
                self.format_money(row.get("comissao") or 0),
            ]
            for col_index, value in enumerate(values):
                item = QTableWidgetItem(value)
                self.indicators_table.setItem(row_index, col_index, item)

        self.indicators_chart.setPlainText(self.build_indicators_chart(rows))
        if rows:
            self.indicators_summary_label.setText(f"{len(rows)} categoria(s)")
        else:
            self.indicators_summary_label.setText("Nenhum dado no período")

    def build_indicators_chart(self, rows):
        if not rows:
            return "Nenhum indicador encontrado para o período selecionado."

        max_sales = max(int(row.get("vendas") or 0) for row in rows) or 1
        lines = ["Vendas por categoria", ""]
        for row in rows[:12]:
            sub_id = str(row.get("sub_id") or "sem_subid")
            vendas = int(row.get("vendas") or 0)
            tamanho = max(1, round((vendas / max_sales) * 28)) if vendas else 0
            barra = "█" * tamanho
            lines.append(f"{sub_id[:18]:18} {barra} {vendas}")

        return "\n".join(lines)

    # ===================== Relatório de conversões =====================
    def update_conversoes_period_fields(self, *_):
        period = self.conversoes_period_input.currentData() or "yesterday"
        today = QDate.currentDate()

        if period == "yesterday":
            start = today.addDays(-1)
            end = today.addDays(-1)
        elif period == "7":
            start = today.addDays(-6)
            end = today
        elif period == "30":
            start = today.addDays(-29)
            end = today
        else:
            start = self.conversoes_start_date.date() if self.conversoes_start_date.date().isValid() else today.addDays(-6)
            end = self.conversoes_end_date.date() if self.conversoes_end_date.date().isValid() else today

        if period != "custom":
            self.conversoes_start_date.setDate(start)
            self.conversoes_end_date.setDate(end)

        custom = period == "custom"
        self.conversoes_start_date.setEnabled(custom)
        self.conversoes_end_date.setEnabled(custom)

    def conversoes_period_range(self):
        start_date = self.conversoes_start_date.date().toPython()
        end_date = self.conversoes_end_date.date().toPython()
        start = datetime.combine(start_date, datetime.min.time())
        end = datetime.combine(end_date, datetime.max.time())
        return start, end

    def refresh_conversoes(self):
        errors = self.validate_api_settings()
        api_errors = [error for error in errors if "grupo" not in error.lower()]
        if api_errors:
            message = "Corrija os campos abaixo antes de buscar as conversões:\n\n"
            message += "\n".join(f"- {error}" for error in api_errors)
            QMessageBox.warning(self, "Configurações incompletas", message)
            return

        start, end = self.conversoes_period_range()
        if start > end:
            QMessageBox.warning(
                self,
                "Período inválido",
                "A data inicial não pode ser maior que a data final.",
            )
            return

        self.save_fields(show_message=False)
        self.refresh_conversoes_button.setEnabled(False)
        self.set_status("Buscando conversões na Shopee...")
        self.add_log(
            "Buscando conversões por produto de "
            f"{start.strftime('%d/%m/%Y')} até {end.strftime('%d/%m/%Y')}..."
        )
        threading.Thread(
            target=self._run_conversoes_refresh,
            args=(start, end),
            daemon=True,
        ).start()

    def _run_conversoes_refresh(self, start, end):
        try:
            conversor = ConversorAfiliados(app_config=self.settings)
            report = conversor.obter_conversoes_por_produto(start, end)
            self.signals.conversoes_loaded.emit(report)
            debug = report.get("debug") or {}
            self.signals.log.emit(
                "Conversões Shopee atualizadas: "
                f"{debug.get('conversoes', 0)} conversão(ões), {debug.get('itens', 0)} item(ns)."
            )
            self.signals.status.emit("Conversões atualizadas.")
        except Exception as e:
            self.signals.log.emit(f"Falha ao buscar conversões Shopee: {self.friendly_error(e)}")
            self.signals.status.emit("Falha ao buscar conversões.")
            self.signals.conversoes_loaded.emit({"erro": str(e), "produtos": [], "totais": {}})
        finally:
            self.signals.conversoes_button_enable.emit()

    def render_conversoes_report(self, report):
        self.refresh_conversoes_button.setEnabled(True)
        if report.get("erro"):
            QMessageBox.warning(
                self,
                "Falha ao buscar conversões",
                "Não consegui buscar as conversões pela API da Shopee.\n\n"
                "Veja os logs para conferir o detalhe retornado pela API.",
            )
        self._conversoes_report_atual = report
        self.render_conversoes_atual()

    def render_conversoes_atual(self, *_):
        report = self._conversoes_report_atual
        if not report:
            return

        produtos = list(report.get("produtos") or [])
        totais = report.get("totais") or {}

        criterio = self.conversoes_sort_input.currentData() or "comissao"
        if criterio == "comissao":
            produtos.sort(key=lambda p: p.get("comissao") or 0, reverse=True)
        elif criterio == "quantidade":
            produtos.sort(key=lambda p: p.get("quantidade") or 0, reverse=True)
        elif criterio == "nome":
            produtos.sort(key=lambda p: (p.get("nome") or "").lower())
        elif criterio == "data":
            produtos.sort(
                key=lambda p: p.get("ultima_compra") or datetime.min,
                reverse=True,
            )

        self.conversoes_produtos_label.setText(str(int(totais.get("produtos") or 0)))
        self.conversoes_qtd_label.setText(str(int(totais.get("quantidade") or 0)))
        self.conversoes_comissao_label.setText(self.format_money(totais.get("comissao") or 0))

        # Desliga a ordenação nativa enquanto preenche, para não embaralhar
        # as linhas enquanto inserimos (e para respeitar a ordem do critério).
        self.conversoes_table.setSortingEnabled(False)
        self.conversoes_table.setRowCount(len(produtos))
        for row_index, p in enumerate(produtos):
            ultima = p.get("ultima_compra")
            ultima_txt = ultima.strftime("%d/%m/%Y") if ultima else "-"
            # valor ordenável para a data (timestamp; "-" vira 0).
            ultima_ord = ultima.timestamp() if ultima else 0

            nome_item = QTableWidgetItem(p.get("nome") or "(sem nome)")
            loja_item = QTableWidgetItem(p.get("loja") or "-")

            # Colunas numéricas: texto formatado, ordenação pelo número real.
            qtd_item = NumericTableItem(p.get("quantidade") or 0, str(int(p.get("quantidade") or 0)))
            preco_item = NumericTableItem(p.get("preco") or 0, self.format_money(p.get("preco") or 0))
            comissao_item = NumericTableItem(p.get("comissao") or 0, self.format_money(p.get("comissao") or 0))
            data_item = NumericTableItem(ultima_ord, ultima_txt)

            for col, cell in enumerate(
                [nome_item, loja_item, qtd_item, preco_item, comissao_item, data_item]
            ):
                self.conversoes_table.setItem(row_index, col, cell)

        self.conversoes_table.setSortingEnabled(True)

        if produtos:
            self.conversoes_summary_label.setText(f"{len(produtos)} produto(s)")
        else:
            self.conversoes_summary_label.setText("Nenhuma conversão no período")

    def _conversoes_produtos_ordenados(self):
        """Retorna os produtos do relatório atual já na ordem do critério
        selecionado em 'Ordenar por' (mesma ordem exibida na tabela)."""
        report = self._conversoes_report_atual or {}
        produtos = list(report.get("produtos") or [])
        criterio = self.conversoes_sort_input.currentData() or "comissao"
        if criterio == "comissao":
            produtos.sort(key=lambda p: p.get("comissao") or 0, reverse=True)
        elif criterio == "quantidade":
            produtos.sort(key=lambda p: p.get("quantidade") or 0, reverse=True)
        elif criterio == "nome":
            produtos.sort(key=lambda p: (p.get("nome") or "").lower())
        elif criterio == "data":
            produtos.sort(key=lambda p: p.get("ultima_compra") or datetime.min, reverse=True)
        return produtos

    def exportar_conversoes_excel(self):
        produtos = self._conversoes_produtos_ordenados()
        if not produtos:
            QMessageBox.information(
                self,
                "Exportar conversões",
                "Não há conversões carregadas para exportar. Clique em "
                "\"Atualizar conversões\" primeiro.",
            )
            return

        report = self._conversoes_report_atual or {}
        inicio = report.get("inicio")
        fim = report.get("fim")
        sufixo = ""
        if inicio and fim:
            sufixo = f"-{inicio.strftime('%Y%m%d')}-a-{fim.strftime('%Y%m%d')}"
        default_name = f"conversoes-shopee{sufixo}.xlsx"

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar conversões",
            str(APP_DIR / default_name),
            "Planilha Excel (*.xlsx);;Arquivo CSV (*.csv)",
        )
        if not file_path:
            return

        cabecalhos = ["Produto", "Loja", "Quantidade", "Preço", "Comissão", "Última compra"]
        linhas = []
        for p in produtos:
            ultima = p.get("ultima_compra")
            linhas.append([
                p.get("nome") or "(sem nome)",
                p.get("loja") or "",
                int(p.get("quantidade") or 0),
                round(float(p.get("preco") or 0), 2),
                round(float(p.get("comissao") or 0), 2),
                ultima.strftime("%d/%m/%Y") if ultima else "",
            ])

        totais = report.get("totais") or {}
        total_qtd = int(totais.get("quantidade") or 0)
        total_comissao = round(float(totais.get("comissao") or 0), 2)

        escolheu_csv = file_path.lower().endswith(".csv")
        try:
            if escolheu_csv:
                self._exportar_conversoes_csv(file_path, cabecalhos, linhas, total_qtd, total_comissao)
            else:
                caminho_final = self._exportar_conversoes_xlsx(
                    file_path, cabecalhos, linhas, total_qtd, total_comissao, inicio, fim
                )
                file_path = caminho_final
            self.add_log(f"Conversões exportadas para: {file_path}")
            QMessageBox.information(
                self, "Conversões exportadas", f"Arquivo gerado:\n\n{file_path}"
            )
        except Exception as e:
            QMessageBox.warning(
                self, "Falha ao exportar", f"Não consegui exportar as conversões:\n\n{e}"
            )

    def _exportar_conversoes_xlsx(self, file_path, cabecalhos, linhas, total_qtd, total_comissao, inicio, fim):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Sem openpyxl instalado: cai para CSV com o mesmo nome (.csv).
            alt = str(Path(file_path).with_suffix(".csv"))
            self._exportar_conversoes_csv(alt, cabecalhos, linhas, total_qtd, total_comissao)
            self.add_log(
                "openpyxl não está instalado; exportei em CSV. Para gerar .xlsx, "
                "instale com: pip install openpyxl"
            )
            return alt

        wb = Workbook()
        ws = wb.active
        ws.title = "Conversões"

        # Título com o período.
        periodo_txt = ""
        if inicio and fim:
            periodo_txt = f"Período: {inicio.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}"
        ws.append([periodo_txt] if periodo_txt else [""])
        ws.append([])

        # Cabeçalho.
        header_row = ws.max_row + 1
        ws.append(cabecalhos)
        fill = PatternFill("solid", fgColor="1F6F5C")
        for col in range(1, len(cabecalhos) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

        # Dados.
        for linha in linhas:
            ws.append(linha)

        # Formatação de moeda nas colunas Preço (4) e Comissão (5).
        primeira_dado = header_row + 1
        ultima_dado = ws.max_row
        for row in range(primeira_dado, ultima_dado + 1):
            ws.cell(row=row, column=4).number_format = 'R$ #,##0.00'
            ws.cell(row=row, column=5).number_format = 'R$ #,##0.00'

        # Linha de totais.
        ws.append([])
        total_row = ws.max_row + 1
        ws.append(["TOTAL", "", total_qtd, "", total_comissao, ""])
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        ws.cell(row=total_row, column=5).font = Font(bold=True)
        ws.cell(row=total_row, column=5).number_format = 'R$ #,##0.00'

        # Largura das colunas.
        larguras = [48, 22, 12, 14, 14, 16]
        for i, largura in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(i)].width = largura

        ws.freeze_panes = f"A{header_row + 1}"
        wb.save(file_path)
        return file_path

    def _exportar_conversoes_csv(self, file_path, cabecalhos, linhas, total_qtd, total_comissao):
        import csv
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(cabecalhos)
            for linha in linhas:
                # No CSV, usa vírgula decimal (padrão brasileiro) para preço/comissão.
                linha_csv = list(linha)
                linha_csv[3] = f"{linha[3]:.2f}".replace(".", ",")
                linha_csv[4] = f"{linha[4]:.2f}".replace(".", ",")
                writer.writerow(linha_csv)
            writer.writerow([])
            writer.writerow(["TOTAL", "", total_qtd, "", f"{total_comissao:.2f}".replace(".", ","), ""])

    def verificar_envio_relatorio_agendado(self):
        if not self.settings.relatorio_email_ativo:
            return

        agora = datetime.now()
        hora_configurada = str(self.settings.relatorio_email_hora or "11:00").strip()
        if agora.strftime("%H:%M") != hora_configurada:
            return

        hoje = agora.strftime("%Y-%m-%d")
        if self.settings.relatorio_email_ultimo_envio == hoje:
            return

        self.add_log(f"Hora do relatório diário por email ({hora_configurada}); gerando e enviando...")
        threading.Thread(
            target=self._gerar_e_enviar_relatorio,
            kwargs={"marcar_enviado": True},
            daemon=True,
        ).start()

    def enviar_relatorio_email_teste(self):
        self.save_fields(show_message=False)
        self.relatorio_email_testar_button.setEnabled(False)
        self.relatorio_email_status_label.setText("Enviando relatório de teste...")
        threading.Thread(
            target=self._gerar_e_enviar_relatorio,
            kwargs={"marcar_enviado": False},
            daemon=True,
        ).start()

    def _gerar_e_enviar_relatorio(self, marcar_enviado=False):
        try:
            agora = datetime.now()
            ontem_inicio = datetime.combine((agora - timedelta(days=1)).date(), datetime.min.time())
            ontem_fim = datetime.combine((agora - timedelta(days=1)).date(), datetime.max.time())
            # O "mês" do relatório é sempre o mês do dia anterior (ontem), não
            # o mês de hoje: no dia 1 de cada mês, "ontem" ainda é o mês
            # passado, então o acumulado mensal deve mostrar esse mês inteiro
            # (já fechado), não o mês novo que praticamente não teve tempo de
            # gerar dado nenhum ainda. O fim do período também é sempre o fim
            # de "ontem" (não "agora"), pra não misturar dois meses diferentes
            # quando o envio acontece logo no início de um mês novo.
            mes_inicio = datetime.combine(ontem_inicio.date().replace(day=1), datetime.min.time())
            mes_fim = ontem_fim

            conversor = ConversorAfiliados(app_config=self.settings)
            resumo_ontem = conversor.obter_indicadores_por_sub_id(ontem_inicio, ontem_fim)
            resumo_mes = conversor.obter_indicadores_por_sub_id(mes_inicio, mes_fim)

            data_referencia = agora.strftime("%d/%m/%Y %H:%M")
            titulo_mes = f"{MESES_PT[mes_inicio.month]}/{mes_inicio.year} (até ontem)"
            corpo_html = montar_corpo_email(resumo_ontem, resumo_mes, data_referencia, titulo_mes)
            assunto = f"Relatório Shopee — {agora.strftime('%d/%m/%Y')}"

            enviar_email(self.settings, assunto, corpo_html)

            if marcar_enviado:
                self.settings.relatorio_email_ultimo_envio = agora.strftime("%Y-%m-%d")
                self.settings.save()

            self.signals.log.emit("Relatório de vendas/comissão enviado por email com sucesso.")
            self.signals.relatorio_email_status.emit(f"Último envio: {agora.strftime('%d/%m/%Y %H:%M')}")
        except Exception as e:
            self.signals.log.emit(f"Falha ao enviar relatório por email: {self.friendly_error(e)}")
            self.signals.relatorio_email_status.emit(f"Falha no último envio: {self.friendly_error(e)}")
        finally:
            self.signals.relatorio_email_button_enable.emit()

    def format_money(self, value):
        try:
            number = float(value or 0)
        except (TypeError, ValueError):
            number = 0.0
        text = f"{number:,.2f}"
        text = text.replace(",", "X").replace(".", ",").replace("X", ".")
        return f"R$ {text}"

    def refresh_history_with_message(self):
        self.refresh_history()
        QMessageBox.information(
            self,
            "Histórico atualizado",
            f"Histórico atualizado. {len(self.filtered_history_rows)} registro(s) exibido(s).",
        )

    def apply_history_filters(self, *_):
        search = self.history_search_input.text().strip().lower()
        period = self.history_period_input.currentData() or "all"
        now = datetime.now()
        filtered = []

        for row in self.history_rows:
            criado_em, link_original, grupo_destino, link_afiliado, texto_final = row
            created_at = self.parse_history_datetime(criado_em)
            if period == "today" and (not created_at or created_at.date() != now.date()):
                continue
            if period in ("7", "30") and (
                not created_at or created_at < now - timedelta(days=int(period))
            ):
                continue

            searchable = " ".join(
                str(value or "")
                for value in (criado_em, link_original, grupo_destino, link_afiliado, texto_final)
            ).lower()
            if search and search not in searchable:
                continue

            filtered.append(row)

        self.filtered_history_rows = filtered
        self.render_history_table(filtered)

    def render_history_table(self, rows):
        self.history_table.setRowCount(len(rows))

        for row_index, row in enumerate(rows):
            criado_em, link_original, grupo_destino, link_afiliado, texto_final = row
            texto_limpo = (texto_final or "").replace("\n", " ")
            values = [
                self.format_history_date(criado_em),
                link_original or "",
                grupo_destino or "",
                link_afiliado or "",
                texto_limpo[:240],
            ]
            tooltips = [
                self.format_history_date(criado_em),
                link_original or "",
                grupo_destino or "",
                link_afiliado or "",
                texto_limpo,
            ]
            for col_index, value in enumerate(values):
                item = QTableWidgetItem(value)
                item.setToolTip(tooltips[col_index])
                self.history_table.setItem(row_index, col_index, item)

        total = len(self.history_rows)
        shown = len(rows)
        if total == shown:
            self.history_summary_label.setText(f"{shown} registro(s)")
        else:
            self.history_summary_label.setText(f"{shown} de {total} registro(s)")

    def copy_selected_affiliate_link(self):
        row_index = self.history_table.currentRow()
        if row_index < 0 or row_index >= len(self.filtered_history_rows):
            QMessageBox.information(self, "Copiar link afiliado", "Nenhuma linha foi selecionada.")
            return

        link_afiliado = self.filtered_history_rows[row_index][3] or ""
        if not link_afiliado:
            QMessageBox.information(self, "Copiar link afiliado", "Esta linha não tem link afiliado para copiar.")
            return

        QApplication.clipboard().setText(link_afiliado)
        self.add_log("Link afiliado copiado para a área de transferência.")
        QMessageBox.information(self, "Copiar link afiliado", "Link afiliado copiado.")

    def export_history_csv(self):
        if not self.filtered_history_rows:
            QMessageBox.information(self, "Exportar histórico", "Não há registros filtrados para exportar.")
            return

        default_name = f"historico-bot-ee-{datetime.now().strftime('%Y-%m-%d')}.csv"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Exportar histórico",
            str(APP_DIR / default_name),
            "Arquivo CSV (*.csv)",
        )
        if not file_path:
            return

        try:
            with open(file_path, "w", newline="", encoding="utf-8-sig") as file:
                writer = csv.writer(file, delimiter=";")
                writer.writerow(["Data", "Link original", "Grupo destino", "Link afiliado", "Texto"])
                for criado_em, link_original, grupo_destino, link_afiliado, texto_final in self.filtered_history_rows:
                    writer.writerow([
                        self.format_history_date(criado_em),
                        link_original or "",
                        grupo_destino or "",
                        link_afiliado or "",
                        texto_final or "",
                    ])
            self.add_log(f"Histórico exportado para: {file_path}")
            QMessageBox.information(self, "Histórico exportado", "Arquivo CSV gerado.")
        except Exception as e:
            QMessageBox.warning(self, "Falha ao exportar", f"Não consegui exportar o histórico:\n\n{e}")

    def clear_history(self):
        resposta = QMessageBox.question(
            self,
            "Limpar histórico",
            "Deseja apagar o histórico de links enviados? Links antigos poderão ser enviados novamente.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if resposta != QMessageBox.Yes:
            return

        db = None
        try:
            db_path = self.resolve_app_path(self.settings.database_path)
            db = Database(str(db_path))
            db.limpar_links_enviados()
            self.refresh_history()
            self.add_log("Histórico de links enviados limpo.")
            QMessageBox.information(self, "Histórico limpo", "Histórico de links enviados limpo.")
        except Exception as e:
            self.add_log(f"Não consegui limpar o histórico: {e}")
        finally:
            if db:
                db.fechar()

    def format_history_date(self, value):
        try:
            parsed = datetime.fromisoformat(value)
            return parsed.strftime("%d/%m/%Y %H:%M")
        except Exception:
            return str(value or "")

    def parse_history_datetime(self, value):
        try:
            return datetime.fromisoformat(value)
        except Exception:
            return None

    def set_status(self, message):
        self.status_label.setText(message)
        if message.startswith("Oferta enviada."):
            self.refresh_history()
        if message in ("Bot parado.", "Bot finalizado."):
            self.history_timer.stop()
            self.stop_button.setEnabled(False)
            self.start_button.setEnabled(True)
            self.set_form_enabled(True)
            self.refresh_history()

    def update_stats(self, stats):
        self.processadas_label.setText(str(stats.get("processadas", 0)))
        self.enviadas_label.setText(str(stats.get("enviadas", 0)))
        self.erros_label.setText(str(stats.get("erros", 0)))

    def theme_overrides(self):
        theme = getattr(self.settings, "tema_visual", "padrao")
        check_icon_path = (RESOURCE_DIR / "assets" / "check_white.svg").as_posix()

        common = """
            QComboBox {
                background: #f8fafc;
                border: 1px solid #cbd5e1;
                border-radius: 10px;
                padding: 9px;
                selection-background-color: #ef4444;
            }
            QComboBox:focus {
                border: 1px solid #ef4444;
                background: #ffffff;
            }
            QComboBox::drop-down {
                border: none;
                width: 28px;
            }
            QComboBox QLineEdit {
                background: transparent;
                border: none;
                padding: 0px;
            }
        """

        if theme == "claro":
            return common + """
                QWidget {
                    color: #172033;
                    background: #f7f9fc;
                }
                #sidebar {
                    background: #ffffff;
                    border: 1px solid #dbe3ee;
                }
                #appTitle {
                    color: #0f172a;
                }
                #appSubtitle {
                    color: #3b5f8f;
                }
                #appVersion {
                    color: #64748b;
                }
                #sidebarFieldLabel {
                    color: #64748b;
                }
                #themeSelect {
                    color: #172033;
                    background: #f8fafc;
                    border-color: #cbd5e1;
                }
                #themeSelect:focus {
                    background: #ffffff;
                    border-color: #2563eb;
                }
                #themeSelect QLineEdit {
                    color: #172033;
                    background: transparent;
                    border: none;
                }
                #statusPill {
                    color: #1d4ed8;
                    background: #eff6ff;
                    border-color: #bfdbfe;
                }
                #metricCard {
                    background: #f8fafc;
                    border-color: #cbd5e1;
                }
                #metricValue {
                    color: #1d4ed8;
                }
                #metricLabel {
                    color: #475569;
                }
                #primaryButton {
                    background: #2563eb;
                }
                #primaryButton:hover {
                    background: #1d4ed8;
                }
                #secondaryButton {
                    background: #475569;
                }
                #secondaryButton:hover {
                    background: #334155;
                }
                #ghostButton {
                    color: #1e293b;
                    background: #ffffff;
                    border-color: #cbd5e1;
                }
                #ghostButton:hover {
                    color: #0f172a;
                    background: #eff6ff;
                    border-color: #2563eb;
                }
                #logs {
                    background: #ffffff;
                    color: #1e293b;
                    border-color: #cbd5e1;
                }
                QLineEdit:focus, QTextEdit:focus, QComboBox:focus {
                    border-color: #2563eb;
                }
            """

        if theme == "noturno":
            noturno_styles = """
                QWidget {
                    color: #e5e7eb;
                    background: #0b1120;
                }
                #content {
                    background: #111827;
                    border-color: #253246;
                }
                #sidebar {
                    background: #020617;
                    border: 1px solid #1f2937;
                }
                #appTitle {
                    color: #f8fafc;
                }
                #appSubtitle {
                    color: #99f6e4;
                }
                #appVersion {
                    color: #94a3b8;
                }
                #sidebarFieldLabel {
                    color: #99f6e4;
                }
                #themeSelect {
                    color: #ccfbf1;
                    background: #0f172a;
                    border-color: #334155;
                }
                #themeSelect:focus {
                    background: #111827;
                    border-color: #14b8a6;
                }
                #themeSelect QLineEdit {
                    color: #ccfbf1;
                    background: transparent;
                    border: none;
                }
                #sectionTitle, #cardTitle {
                    color: #f8fafc;
                }
                #fieldLabel {
                    color: #cbd5e1;
                }
                #configCard {
                    background: #172033;
                    border-color: #334155;
                }
                #settingsPanel {
                    background: transparent;
                }
                #statusPill {
                    color: #ccfbf1;
                    background: #0f2f2b;
                    border-color: #115e59;
                }
                #metricCard {
                    background: #172033;
                    border-color: #334155;
                }
                #metricValue {
                    color: #5eead4;
                }
                #metricLabel {
                    color: #cbd5e1;
                }
                QLineEdit, QTextEdit, QComboBox {
                    color: #f8fafc;
                    background: #0f172a;
                    border-color: #334155;
                }
                QLineEdit:focus, QTextEdit:focus, QComboBox:focus {
                    background: #111827;
                    border-color: #14b8a6;
                }
                QCheckBox {
                    color: #dbeafe;
                }
                QCheckBox::indicator {
                    width: 16px;
                    height: 16px;
                    border: 1px solid #e5e7eb;
                    border-radius: 4px;
                    background: #020617;
                }
                QCheckBox::indicator:checked {
                    background: #14b8a6;
                    border-color: #ffffff;
                    image: url("__CHECK_ICON__");
                }
                #validationBadge {
                    color: #cbd5e1;
                    background: #0f172a;
                    border-color: #334155;
                }
                #validationBadge[valid="true"] {
                    color: #ccfbf1;
                    background: #0f2f2b;
                    border-color: #115e59;
                }
                #stepButton, #toolButton, #compactToolButton, #helpButton {
                    color: #e5e7eb;
                    background: #1f2937;
                    border-color: #334155;
                }
                #stepButton:hover, #toolButton:hover, #compactToolButton:hover, #helpButton:hover {
                    background: #253246;
                    border-color: #14b8a6;
                }
                #primaryButton {
                    color: #042f2e;
                    background: #5eead4;
                }
                #primaryButton:hover {
                    background: #2dd4bf;
                }
                #secondaryButton {
                    background: #334155;
                }
                #secondaryButton:hover {
                    background: #475569;
                }
                #ghostButton {
                    color: #e5e7eb;
                    background: transparent;
                    border-color: #475569;
                }
                #ghostButton:hover {
                    color: #ccfbf1;
                    background: #0f172a;
                    border-color: #14b8a6;
                }
                #logs {
                    background: #020617;
                    color: #ccfbf1;
                    border-color: #334155;
                }
                #dataTabs QTabBar::tab,
                #configTabs QTabBar::tab {
                    color: #cbd5e1;
                    background: #172033;
                    border-color: #334155;
                }
                #dataTabs QTabBar::tab:selected,
                #configTabs QTabBar::tab:selected {
                    color: #f8fafc;
                    background: #111827;
                    border-color: #475569;
                }
                #historyTable {
                    color: #e5e7eb;
                    background: #0f172a;
                    alternate-background-color: #111827;
                    border-color: #334155;
                    gridline-color: #253246;
                    selection-background-color: #1f6f78;
                    selection-color: #f8fafc;
                }
                QTableWidget#historyTable::item:selected,
                QTableWidget#historyTable::item:selected:active,
                QTableWidget#historyTable::item:selected:!active {
                    color: #f8fafc;
                    background: #1f6f78;
                    border: 0px;
                    outline: none;
                }
                QTableWidget#historyTable::item:focus {
                    border: none;
                    outline: none;
                }
                #historyTable QHeaderView::section {
                    color: #e5e7eb;
                    background: #172033;
                    border-right-color: #334155;
                }
            """
            return common + noturno_styles.replace("__CHECK_ICON__", check_icon_path)

        return common

    def apply_styles(self):
        check_icon_path = (RESOURCE_DIR / "assets" / "check_white.svg").as_posix()
        base_styles = """
            QWidget {
                font-family: "Segoe UI";
                font-size: 13px;
                color: #18222f;
                background: #f5f7fb;
            }
            #sidebar {
                background: #101827;
                border-radius: 16px;
                min-width: 270px;
                max-width: 310px;
            }
            #sidebar QLabel {
                background: transparent;
            }
            #content {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 16px;
            }
            #appTitle {
                color: #ffffff;
                font-size: 26px;
                font-weight: 700;
                background: transparent;
            }
            #appSubtitle {
                color: #b8c7dc;
                background: transparent;
            }
            #appVersion {
                color: #94a3b8;
                background: transparent;
                font-size: 12px;
            }
            #sidebarFieldLabel {
                color: #94a3b8;
                background: transparent;
                font-size: 12px;
                font-weight: 700;
                padding: 0px;
            }
            #themeRow {
                background: transparent;
            }
            #themeSelect {
                color: #e2e8f0;
                background: #111c2d;
                border: 1px solid #334155;
                border-radius: 10px;
                padding: 9px 10px;
                min-height: 24px;
                text-align: center;
                selection-background-color: #ef4444;
            }
            #themeSelect:focus {
                border: 1px solid #ef4444;
                background: #132238;
            }
            #themeSelect::drop-down {
                border: none;
                width: 28px;
            }
            #themeSelect QLineEdit {
                color: #e2e8f0;
                background: transparent;
                border: none;
                padding: 0px;
            }
            #statusPill {
                color: #dbeafe;
                background: #1d2a3f;
                border: 1px solid #2b3b55;
                border-radius: 10px;
                padding: 12px;
            }
            #metricCard {
                background: #18263b;
                border: 1px solid #314460;
                border-radius: 12px;
            }
            #metricCard QLabel {
                background: transparent;
            }
            #metricValue {
                color: #ffffff;
                font-size: 24px;
                font-weight: 700;
                background: transparent;
            }
            #metricLabel {
                color: #b8c7dc;
                background: transparent;
            }
            #sectionTitle {
                font-size: 20px;
                font-weight: 700;
                color: #111827;
            }
            #fieldLabel {
                color: #475569;
                font-weight: 600;
                padding: 6px 8px;
                min-height: 20px;
            }
            #labelHelpContainer {
                background: transparent;
            }
            #helpButton {
                color: #64748b;
                background: #eef2f7;
                border: 1px solid #cbd5e1;
                border-radius: 10px;
                padding: 0px;
                font-size: 12px;
                font-weight: 700;
            }
            #helpButton:hover {
                color: #111827;
                background: #e2e8f0;
                border-color: #94a3b8;
            }
            #configCard {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
            }
            #settingsPanel {
                background: transparent;
            }
            #cardTitle {
                color: #111827;
                background: transparent;
                font-size: 15px;
                font-weight: 700;
                padding-bottom: 4px;
            }
            QLineEdit, QTextEdit {
                background: #f8fafc;
                border: 1px solid #cbd5e1;
                border-radius: 10px;
                padding: 6px 12px;
                selection-background-color: #ef4444;
            }
            QLineEdit:focus, QTextEdit:focus {
                border: 1px solid #ef4444;
                background: #ffffff;
            }
            QLineEdit[invalid="true"], QTextEdit[invalid="true"] {
                border: 1px solid #ef4444;
                background: #fff1f2;
            }
            #secondsInput {
                min-height: 28px;
                max-width: 92px;
            }
            #secondsInput[invalid="true"] {
                border: 1px solid #ef4444;
                background: #fff1f2;
            }
            #stepButton {
                color: #18222f;
                background: #eef2f7;
                border: 1px solid #cbd5e1;
                border-radius: 10px;
                padding: 0px;
                font-size: 18px;
                font-weight: 700;
            }
            #stepButton:hover {
                background: #e2e8f0;
                border-color: #94a3b8;
            }
            QCheckBox {
                color: #334155;
                spacing: 8px;
                padding: 7px 8px;
                min-height: 22px;
            }
            QCheckBox::indicator {
                width: 16px;
                height: 16px;
                border: 1px solid #94a3b8;
                border-radius: 4px;
                background: #ffffff;
            }
            QCheckBox::indicator:checked {
                background: #14b8a6;
                border-color: #ffffff;
                image: url("__CHECK_ICON__");
            }
            #validationBadge {
                color: #64748b;
                background: #f8fafc;
                border: 1px solid #e2e8f0;
                border-radius: 8px;
                padding: 7px 10px;
                font-size: 12px;
                font-weight: 700;
            }
            #validationBadge[valid="true"] {
                color: #166534;
                background: #f0fdf4;
                border-color: #bbf7d0;
            }
            QPushButton {
                border: none;
                border-radius: 10px;
                padding: 11px 14px;
                font-weight: 700;
            }
            #primaryButton {
                color: #ffffff;
                background: #ef4444;
            }
            #primaryButton:hover {
                background: #dc2626;
            }
            #secondaryButton {
                color: #ffffff;
                background: #334155;
            }
            #secondaryButton:hover {
                background: #1f2937;
            }
            #toolButton {
                color: #1f2937;
                background: #eef2f7;
                border: 1px solid #cbd5e1;
                padding: 8px 12px;
            }
            #toolButton:hover {
                background: #e2e8f0;
                border-color: #94a3b8;
            }
            #dangerToolButton {
                color: #991b1b;
                background: #fff1f2;
                border: 1px solid #fecdd3;
            }
            #dangerToolButton:hover {
                background: #ffe4e6;
                border-color: #fda4af;
            }
            #compactToolButton {
                color: #1f2937;
                background: #eef2f7;
                border: 1px solid #cbd5e1;
                border-radius: 8px;
                padding: 7px 12px;
                font-size: 12px;
                font-weight: 700;
            }
            #compactToolButton:hover {
                background: #e2e8f0;
                border-color: #94a3b8;
            }
            #compactDangerButton {
                color: #991b1b;
                background: #fff1f2;
                border: 1px solid #fecdd3;
                border-radius: 8px;
                padding: 7px 12px;
                font-size: 12px;
                font-weight: 700;
            }
            #compactDangerButton:hover {
                background: #ffe4e6;
                border-color: #fda4af;
            }
            #ghostButton {
                color: #e2e8f0;
                background: transparent;
                border: 1px solid #334155;
            }
            #ghostButton:hover {
                color: #ffffff;
                background: #111827;
                border-color: #64748b;
            }
            QPushButton:disabled {
                color: #94a3b8;
                background: #e2e8f0;
            }
            #logs {
                background: #0f172a;
                color: #dbeafe;
                border: 1px solid #1e293b;
                border-radius: 12px;
                min-height: 210px;
                font-family: Consolas, "Courier New";
                font-size: 12px;
            }
            #dataTabs::pane,
            #configTabs::pane {
                border: none;
                background: transparent;
                margin-top: 6px;
            }
            #dataTabs QTabBar::tab,
            #configTabs QTabBar::tab {
                color: #475569;
                background: #eef2f7;
                border: 1px solid #dbe3ee;
                border-bottom: none;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                padding: 9px 16px;
                margin-right: 4px;
                font-weight: 700;
            }
            #dataTabs QTabBar::tab:selected,
            #configTabs QTabBar::tab:selected {
                color: #111827;
                background: #ffffff;
                border-color: #cbd5e1;
            }
            #historyTable {
                background: #ffffff;
                alternate-background-color: #f8fafc;
                border: 1px solid #dbe3ee;
                border-radius: 12px;
                gridline-color: #e2e8f0;
                selection-background-color: #bfd7f6;
                selection-color: #111827;
            }
            QTableWidget#historyTable::item:selected,
            QTableWidget#historyTable::item:selected:active,
            QTableWidget#historyTable::item:selected:!active {
                color: #111827;
                background: #bfd7f6;
                border: 0px;
                outline: none;
            }
            QTableWidget#historyTable::item:focus {
                border: none;
                outline: none;
            }
            #historyTable QHeaderView::section {
                color: #334155;
                background: #eef2f7;
                border: none;
                border-right: 1px solid #dbe3ee;
                padding: 8px;
                font-weight: 700;
            }
        """.replace("__CHECK_ICON__", check_icon_path)
        self.setStyleSheet(base_styles + self.theme_overrides())


def _filtrar_avisos_qt_benignos(tipo_msg, contexto, mensagem):
    """
    Alguns avisos que o Qt imprime no terminal sao puramente cosmeticos e nao
    indicam nenhum problema real no app. O mais comum no Windows e:
      "QFont::setPointSize: Point size <= 0 (-1), must be greater than 0"
    Isso acontece porque o Qt consulta a fonte padrao do Windows antes do
    app.setFont(...) ser aplicado, e essa fonte do sistema as vezes vem
    definida so em pixels (sem "point size"). Nao afeta em nada o funcionamento
    do bot. Filtramos apenas essas mensagens conhecidas; qualquer outro aviso
    ou erro do Qt continua aparecendo normalmente no terminal.
    """
    mensagens_ignoradas = (
        "QFont::setPointSize: Point size <= 0",
    )
    if any(trecho in mensagem for trecho in mensagens_ignoradas):
        return
    if tipo_msg == QtMsgType.QtWarningMsg:
        print(f"Aviso do Qt: {mensagem}")
    elif tipo_msg in (QtMsgType.QtCriticalMsg, QtMsgType.QtFatalMsg):
        print(f"Erro do Qt: {mensagem}")


def main():
    qInstallMessageHandler(_filtrar_avisos_qt_benignos)
    app = QApplication(sys.argv)
    app.setFont(QFont("Segoe UI", 10))
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()

