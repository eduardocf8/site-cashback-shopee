from decimal import Decimal

from django.contrib.auth import get_user_model
from django.db.models import Count, Q, QuerySet, Sum
from django.utils import timezone

from accounts.models import Indicacao
from saques.models import Saque

from .models import Pedido

FORMATO_MOEDA = '"R$" #,##0.00'
FORMATO_DATA = "dd/mm/yyyy hh:mm"
# "%" sem aspas faz o Excel multiplicar o valor por 100 (formato nativo de percentual,
# pensado pra frações tipo 0.5). Nosso percentual já vem calculado (ex: 170.6), então
# precisa do símbolo entre aspas - texto literal, sem a multiplicação automática.
FORMATO_PERCENTUAL = '0.0"%"'


def _no_periodo(queryset: QuerySet, campo_data: str, data_inicio, data_fim) -> QuerySet:
    if data_inicio:
        queryset = queryset.filter(**{f"{campo_data}__date__gte": data_inicio})
    if data_fim:
        queryset = queryset.filter(**{f"{campo_data}__date__lte": data_fim})
    return queryset


ORIGEM_SITE = "site"
ORIGEM_FORA = "fora"


def obter_pedidos_filtrados(data_inicio=None, data_fim=None, status=None, origem=None) -> QuerySet:
    """Pedidos filtrados por período (via data_compra), status e origem - base
    compartilhada entre a tela de analytics e as exportações, pra manter tudo batendo.

    "origem" distingue pedidos gerados por aqui (têm um Click vinculado, geram cashback
    de verdade pra um usuário) dos que sobram sem Click - outras campanhas, compras
    pessoais etc. (ver pedidos/admin.py::OrigemFilter) - continuam guardados no banco
    com um valor_cashback calculado, mas não são pagos a ninguém de verdade."""
    pedidos = _no_periodo(Pedido.objects.all(), "data_compra", data_inicio, data_fim)
    if status:
        pedidos = pedidos.filter(status=status)
    if origem == ORIGEM_SITE:
        pedidos = pedidos.filter(click__isnull=False)
    elif origem == ORIGEM_FORA:
        pedidos = pedidos.filter(click__isnull=True)
    return pedidos


def obter_analytics(data_inicio=None, data_fim=None, status=None, origem=None) -> dict:
    """Agrega os números do negócio pro período/status/origem informado - usado pela
    tela /admin/pedidos/pedido/analytics/. Cada bloco filtra pela sua própria data
    "natural" (pedido por data_compra, saque por criado_em, indicação por criado_em,
    usuário por date_joined), porque não faz sentido contar, por exemplo, um saque
    solicitado fora do período só porque o pedido que originou aquele cashback foi
    comprado dentro dele. O filtro de origem só se aplica aos pedidos - saques,
    indicações e novos usuários não têm essa distinção.
    """
    pedidos = obter_pedidos_filtrados(data_inicio, data_fim, status, origem)

    totais_pedidos = pedidos.aggregate(
        total_comissao=Sum("valor_comissao"), total_cashback=Sum("valor_cashback"), total=Count("id")
    )
    total_comissao = totais_pedidos["total_comissao"] or Decimal("0")
    total_cashback = totais_pedidos["total_cashback"] or Decimal("0")
    total_pedidos = totais_pedidos["total"] or 0

    por_status_bruto = {
        linha["status"]: linha
        for linha in pedidos.values("status").annotate(
            total=Count("id"), comissao=Sum("valor_comissao"), cashback=Sum("valor_cashback")
        )
    }
    resumo_status = [
        {
            "status": chave,
            "status_label": label,
            "total": (por_status_bruto.get(chave) or {}).get("total") or 0,
            "comissao": (por_status_bruto.get(chave) or {}).get("comissao") or Decimal("0"),
            "cashback": (por_status_bruto.get(chave) or {}).get("cashback") or Decimal("0"),
        }
        for chave, label in Pedido.STATUS_CHOICES
    ]
    cashback_por_status = {linha["status"]: linha["cashback"] for linha in resumo_status}

    saques = _no_periodo(Saque.objects.all(), "criado_em", data_inicio, data_fim)
    totais_saques = saques.aggregate(total_valor=Sum("valor"), total=Count("id"))
    saques_por_status_bruto = {
        linha["status"]: linha
        for linha in saques.values("status").annotate(total=Count("id"), valor=Sum("valor"))
    }
    saques_por_status = [
        {
            "status": chave,
            "status_label": label,
            "total": (saques_por_status_bruto.get(chave) or {}).get("total") or 0,
            "valor": (saques_por_status_bruto.get(chave) or {}).get("valor") or Decimal("0"),
        }
        for chave, label in Saque.STATUS_CHOICES
    ]

    indicacoes = _no_periodo(Indicacao.objects.all(), "criado_em", data_inicio, data_fim)
    total_indicacoes = indicacoes.count()
    indicacoes_concluidas = indicacoes.filter(pedido_bonus_indicador__isnull=False).count()
    ranking_indicadores = list(
        indicacoes.values("indicador__username")
        .annotate(
            total_indicacoes=Count("id"),
            concluidas=Count("id", filter=Q(pedido_bonus_indicador__isnull=False)),
        )
        .order_by("-total_indicacoes", "-concluidas")[:20]
    )

    novos_usuarios = _no_periodo(get_user_model().objects.all(), "date_joined", data_inicio, data_fim).count()

    return {
        "total_comissao": total_comissao,
        "total_cashback": total_cashback,
        "total_pedidos": total_pedidos,
        "margem_retida": total_comissao - total_cashback,
        "percentual_repassado": (total_cashback / total_comissao * 100) if total_comissao else Decimal("0"),
        "ticket_medio_cashback": (total_cashback / total_pedidos) if total_pedidos else Decimal("0"),
        "resumo_status": resumo_status,
        "saldo_a_liberar": cashback_por_status.get(Pedido.STATUS_PENDENTE, Decimal("0"))
        + cashback_por_status.get(Pedido.STATUS_VALIDADO, Decimal("0")),
        "saldo_liberado": cashback_por_status.get(Pedido.STATUS_LIBERADO, Decimal("0")),
        "total_saques_valor": totais_saques["total_valor"] or Decimal("0"),
        "total_saques": totais_saques["total"] or 0,
        "saques_por_status": saques_por_status,
        "total_indicacoes": total_indicacoes,
        "indicacoes_concluidas": indicacoes_concluidas,
        "ranking_indicadores": ranking_indicadores,
        "novos_usuarios": novos_usuarios,
    }


def gerar_planilha_analytics(data_inicio=None, data_fim=None, status=None, origem=None):
    """Gera a planilha (.xlsx) de analytics já formatada - mesma base de dados de
    obter_analytics()/obter_pedidos_filtrados(), pra bater exatamente com o que a tela
    mostra. Import do openpyxl fica dentro da função de propósito: só é usado aqui, não
    precisa pagar esse custo de import em todo o resto do admin."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    dados = obter_analytics(data_inicio, data_fim, status, origem)
    pedidos = (
        obter_pedidos_filtrados(data_inicio, data_fim, status, origem)
        .select_related("usuario")
        .order_by("-data_compra")
    )

    fonte_titulo = Font(bold=True, size=14)
    fonte_subtitulo = Font(italic=True, color="666666")
    fonte_secao = Font(bold=True, size=12)
    fonte_cabecalho = Font(bold=True, color="FFFFFF")
    preenchimento_cabecalho = PatternFill("solid", fgColor="205067")

    def escrever_cabecalho(planilha, linha, coluna_inicial, titulos):
        for deslocamento, titulo in enumerate(titulos):
            celula = planilha.cell(row=linha, column=coluna_inicial + deslocamento, value=titulo)
            celula.font = fonte_cabecalho
            celula.fill = preenchimento_cabecalho
        return linha + 1

    def definir_larguras(planilha, valores):
        for indice, valor in enumerate(valores, start=1):
            planilha.column_dimensions[get_column_letter(indice)].width = valor

    livro = Workbook()

    resumo = livro.active
    resumo.title = "Resumo"
    definir_larguras(resumo, [32, 16, 16, 16])
    resumo["A1"] = "Analytics — cash-b"
    resumo["A1"].font = fonte_titulo
    periodo = "{} a {}".format(
        data_inicio.strftime("%d/%m/%Y") if data_inicio else "—",
        data_fim.strftime("%d/%m/%Y") if data_fim else "—",
    )
    status_label = dict(Pedido.STATUS_CHOICES).get(status, "todos") if status else "todos"
    origem_label = {ORIGEM_SITE: "gerados no site", ORIGEM_FORA: "fora do site"}.get(origem, "todos")
    resumo["A2"] = f"Período: {periodo}  ·  Status do pedido: {status_label}  ·  Origem: {origem_label}"
    resumo["A2"].font = fonte_subtitulo

    linha = 4
    resumo.cell(row=linha, column=1, value="Indicadores gerais").font = fonte_secao
    linha += 1
    linha = escrever_cabecalho(resumo, linha, 1, ["Indicador", "Valor"])
    kpis = [
        ("Comissão total", dados["total_comissao"], FORMATO_MOEDA),
        ("Cashback repassado", dados["total_cashback"], FORMATO_MOEDA),
        ("Margem retida", dados["margem_retida"], FORMATO_MOEDA),
        ("% da comissão repassado", dados["percentual_repassado"], FORMATO_PERCENTUAL),
        ("Total de pedidos", dados["total_pedidos"], "0"),
        ("Ticket médio de cashback", dados["ticket_medio_cashback"], FORMATO_MOEDA),
        ("Saldo a liberar (pendente + validado)", dados["saldo_a_liberar"], FORMATO_MOEDA),
        ("Saldo liberado", dados["saldo_liberado"], FORMATO_MOEDA),
        ("Total sacado", dados["total_saques_valor"], FORMATO_MOEDA),
        ("Nº de saques no período", dados["total_saques"], "0"),
        ("Novos usuários no período", dados["novos_usuarios"], "0"),
        ("Indicações no período", dados["total_indicacoes"], "0"),
        ("Indicações concluídas (dobro pago)", dados["indicacoes_concluidas"], "0"),
    ]
    for rotulo, valor, formato in kpis:
        resumo.cell(row=linha, column=1, value=rotulo)
        resumo.cell(row=linha, column=2, value=valor).number_format = formato
        linha += 1

    linha += 2
    resumo.cell(row=linha, column=1, value="Pedidos por status").font = fonte_secao
    linha += 1
    linha = escrever_cabecalho(resumo, linha, 1, ["Status", "Pedidos", "Comissão", "Cashback"])
    for item in dados["resumo_status"]:
        resumo.cell(row=linha, column=1, value=item["status_label"])
        resumo.cell(row=linha, column=2, value=item["total"])
        resumo.cell(row=linha, column=3, value=item["comissao"]).number_format = FORMATO_MOEDA
        resumo.cell(row=linha, column=4, value=item["cashback"]).number_format = FORMATO_MOEDA
        linha += 1

    linha += 2
    resumo.cell(row=linha, column=1, value="Saques por status").font = fonte_secao
    linha += 1
    linha = escrever_cabecalho(resumo, linha, 1, ["Status", "Saques", "Valor"])
    for item in dados["saques_por_status"]:
        resumo.cell(row=linha, column=1, value=item["status_label"])
        resumo.cell(row=linha, column=2, value=item["total"])
        resumo.cell(row=linha, column=3, value=item["valor"]).number_format = FORMATO_MOEDA
        linha += 1

    linha += 2
    resumo.cell(row=linha, column=1, value="Ranking de indicadores").font = fonte_secao
    linha += 1
    if dados["ranking_indicadores"]:
        linha = escrever_cabecalho(resumo, linha, 1, ["Usuário", "Indicações", "Concluídas"])
        for item in dados["ranking_indicadores"]:
            resumo.cell(row=linha, column=1, value=item["indicador__username"])
            resumo.cell(row=linha, column=2, value=item["total_indicacoes"])
            resumo.cell(row=linha, column=3, value=item["concluidas"])
            linha += 1
    else:
        resumo.cell(row=linha, column=1, value="Nenhuma indicação no período filtrado.")

    aba_pedidos = livro.create_sheet("Pedidos")
    definir_larguras(aba_pedidos, [22, 20, 14, 14, 14, 14, 20, 20, 20])
    escrever_cabecalho(
        aba_pedidos,
        1,
        1,
        [
            "Order ID",
            "Usuário",
            "Status",
            "Valor do pedido",
            "Comissão",
            "Cashback",
            "Data da compra",
            "Data de validação",
            "Data de liberação",
        ],
    )
    aba_pedidos.freeze_panes = "A2"
    linha = 2
    for pedido in pedidos.iterator():
        aba_pedidos.cell(row=linha, column=1, value=pedido.order_id)
        aba_pedidos.cell(row=linha, column=2, value=pedido.usuario.username if pedido.usuario else "")
        aba_pedidos.cell(row=linha, column=3, value=pedido.get_status_display())
        aba_pedidos.cell(row=linha, column=4, value=pedido.valor_pedido).number_format = FORMATO_MOEDA
        aba_pedidos.cell(row=linha, column=5, value=pedido.valor_comissao).number_format = FORMATO_MOEDA
        aba_pedidos.cell(row=linha, column=6, value=pedido.valor_cashback).number_format = FORMATO_MOEDA
        for coluna, valor_data in ((7, pedido.data_compra), (8, pedido.data_validacao), (9, pedido.data_liberacao)):
            # Excel não aceita datetime com timezone - convertemos pro horário local e
            # removemos o tzinfo antes de escrever na célula.
            if valor_data:
                celula = aba_pedidos.cell(row=linha, column=coluna, value=timezone.localtime(valor_data).replace(tzinfo=None))
                celula.number_format = FORMATO_DATA
        linha += 1

    return livro
