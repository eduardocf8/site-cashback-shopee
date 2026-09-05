from datetime import date, datetime, timedelta, timezone as dt_timezone
from decimal import Decimal
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook

from django.contrib.auth import get_user_model
from django.core import mail
from django.db import connection
from django.test import TestCase, override_settings
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import timezone

from accounts.models import Indicacao
from links.models import Click
from saques.models import Saque

from .analytics import obter_analytics, obter_serie_diaria, origem_detalhada
from .models import CampanhaCashback, Pedido
from .notificacoes import notificar_indicador_bonus_pendente
from .services import calcular_data_prevista_liberacao, liberar_saldo, mapear_status, resolver_click, sincronizar


class MapearStatusTests(TestCase):
    def test_termos_de_cancelamento(self):
        for bruto in ["CANCELLED", "INVALID", "REJECTED", "UNPAID", "FRAUD_ORDER"]:
            self.assertEqual(mapear_status(bruto), Pedido.STATUS_CANCELADO, bruto)

    def test_termos_de_validacao(self):
        for bruto in ["COMPLETED", "CONFIRMED", "PAID", "SUCCESS"]:
            self.assertEqual(mapear_status(bruto), Pedido.STATUS_VALIDADO, bruto)

    def test_desconhecido_cai_para_pendente(self):
        self.assertEqual(mapear_status("PENDING_REVIEW"), Pedido.STATUS_PENDENTE)
        self.assertEqual(mapear_status("ALGO_QUE_NUNCA_VIMOS"), Pedido.STATUS_PENDENTE)
        self.assertEqual(mapear_status(""), Pedido.STATUS_PENDENTE)
        self.assertEqual(mapear_status(None), Pedido.STATUS_PENDENTE)


class CalcularDataPrevistaLiberacaoTests(TestCase):
    def test_sem_data_validacao_retorna_none(self):
        self.assertIsNone(calcular_data_prevista_liberacao(None))

    def test_soma_dois_meses_dentro_do_mesmo_ano(self):
        validacao = datetime(2026, 3, 15, tzinfo=dt_timezone.utc)
        self.assertEqual(calcular_data_prevista_liberacao(validacao), date(2026, 5, 1))

    def test_vira_o_ano_quando_ultrapassa_dezembro(self):
        validacao = datetime(2026, 11, 20, tzinfo=dt_timezone.utc)
        self.assertEqual(calcular_data_prevista_liberacao(validacao), date(2027, 1, 1))

        validacao = datetime(2026, 12, 5, tzinfo=dt_timezone.utc)
        self.assertEqual(calcular_data_prevista_liberacao(validacao), date(2027, 2, 1))

    def test_dia_do_mes_da_validacao_nao_importa(self):
        self.assertEqual(
            calcular_data_prevista_liberacao(datetime(2026, 1, 31, tzinfo=dt_timezone.utc)),
            calcular_data_prevista_liberacao(datetime(2026, 1, 1, tzinfo=dt_timezone.utc)),
        )


class ResolverClickTests(TestCase):
    def setUp(self):
        self.usuario = get_user_model().objects.create_user(
            username="compradora", password="senha123", cpf="39053344705"
        )
        self.click = Click.objects.create(
            usuario=self.usuario,
            tipo=Click.TIPO_HOME,
            url_original="https://shopee.com.br/",
            link_gerado="https://shope.ee/abc",
        )

    def test_identifica_click_com_subids_separados_por_virgula(self):
        utm = f"{self.click.sub_id_usuario()},{self.click.sub_id_click()}"
        encontrado = resolver_click(utm)
        self.assertEqual(encontrado, self.click)

    def test_identifica_click_com_subids_separados_por_pipe(self):
        utm = f"{self.click.sub_id_usuario()}|{self.click.sub_id_click()}"
        encontrado = resolver_click(utm)
        self.assertEqual(encontrado, self.click)

    def test_utm_content_vazio_retorna_none(self):
        self.assertIsNone(resolver_click(""))
        self.assertIsNone(resolver_click(None))

    def test_utm_content_sem_uuid_conhecido_retorna_none(self):
        self.assertIsNone(resolver_click("facebook,instagram"))


@override_settings(
    SHOPEE_AFFILIATE_APP_ID="app123",
    SHOPEE_AFFILIATE_SECRET="segredo123",
    SHOPEE_CASHBACK_PERCENTUAL=100,
)
class SincronizarTests(TestCase):
    def setUp(self):
        self.usuario = get_user_model().objects.create_user(
            username="compradora", password="senha123", cpf="39053344705"
        )
        self.click = Click.objects.create(
            usuario=self.usuario,
            tipo=Click.TIPO_HOME,
            url_original="https://shopee.com.br/",
            link_gerado="https://shope.ee/abc",
        )

    def _pagina(self, orders, has_next_page=False, scroll_id=""):
        return {
            "nodes": [
                {
                    "conversionId": "999",
                    "purchaseTime": 1700000000,
                    "utmContent": f"{self.click.sub_id_usuario()},{self.click.sub_id_click()}",
                    "orders": orders,
                }
            ],
            "pageInfo": {"hasNextPage": has_next_page, "scrollId": scroll_id},
        }

    @patch("pedidos.services.buscar_conversoes")
    def test_cria_pedido_pendente_e_calcula_cashback_com_100_por_cento(self, mock_buscar):
        mock_buscar.return_value = self._pagina(
            [
                {
                    "orderId": "ORD1",
                    "orderStatus": "PENDING",
                    "items": [{"completeTime": None, "itemTotalCommission": "8.00", "actualAmount": "100.00"}],
                }
            ]
        )

        resultado = sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD1")
        self.assertEqual(pedido.status, Pedido.STATUS_PENDENTE)
        self.assertEqual(pedido.usuario, self.usuario)
        self.assertEqual(pedido.click, self.click)
        self.assertEqual(pedido.valor_pedido, Decimal("100.00"))
        self.assertEqual(pedido.valor_comissao, Decimal("8.00"))
        self.assertEqual(pedido.valor_cashback, Decimal("8.00"))
        self.assertEqual(resultado, {"novos": 1, "atualizados": 0, "nao_identificados": 0})

    @patch("pedidos.services.buscar_conversoes")
    def test_valor_pedido_soma_actual_amount_de_todos_os_itens(self, mock_buscar):
        mock_buscar.return_value = self._pagina(
            [
                {
                    "orderId": "ORD-VALOR",
                    "orderStatus": "PENDING",
                    "items": [
                        {"completeTime": None, "itemTotalCommission": "1.00", "actualAmount": "50.00"},
                        {"completeTime": None, "itemTotalCommission": "2.00", "actualAmount": "30.00"},
                    ],
                }
            ]
        )

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-VALOR")
        self.assertEqual(pedido.valor_pedido, Decimal("80.00"))

    @override_settings(SHOPEE_CASHBACK_PERCENTUAL=80)
    @patch("pedidos.services.buscar_conversoes")
    def test_calcula_cashback_com_percentual_configurado(self, mock_buscar):
        mock_buscar.return_value = self._pagina(
            [{"orderId": "ORD2", "orderStatus": "PENDING", "items": [{"completeTime": None, "itemTotalCommission": "10.00"}]}]
        )

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD2")
        self.assertEqual(pedido.valor_cashback, Decimal("8.00"))

    @patch("pedidos.services.buscar_conversoes")
    def test_atualiza_pedido_existente_quando_status_muda(self, mock_buscar):
        mock_buscar.return_value = self._pagina(
            [{"orderId": "ORD3", "orderStatus": "PENDING", "items": [{"completeTime": None, "itemTotalCommission": "5.00"}]}]
        )
        sincronizar(1690000000, 1700000000)

        mock_buscar.return_value = self._pagina(
            [{"orderId": "ORD3", "orderStatus": "COMPLETED", "items": [{"completeTime": 1700000500, "itemTotalCommission": "5.00"}]}]
        )
        resultado = sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD3")
        self.assertEqual(pedido.status, Pedido.STATUS_VALIDADO)
        self.assertIsNotNone(pedido.data_validacao)
        self.assertEqual(resultado, {"novos": 0, "atualizados": 1, "nao_identificados": 0})
        self.assertEqual(Pedido.objects.count(), 1)

    @patch("pedidos.services.buscar_conversoes")
    def test_pedido_sem_click_identificavel_fica_sem_usuario_mas_e_salvo(self, mock_buscar):
        pagina = self._pagina(
            [
                {
                    "orderId": "ORD4",
                    "orderStatus": "PENDING",
                    "items": [{"completeTime": None, "itemTotalCommission": "3.00", "actualAmount": "40.00"}],
                }
            ]
        )
        pagina["nodes"][0]["utmContent"] = "origem-desconhecida"
        mock_buscar.return_value = pagina

        resultado = sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD4")
        self.assertIsNone(pedido.usuario)
        self.assertIsNone(pedido.click)
        self.assertEqual(resultado["nao_identificados"], 1)
        # Sem Click não tem usuário pra receber o cashback - não faz sentido calcular um
        # valor que nunca vai ser pago a ninguém. valor_comissao e valor_pedido (o que a
        # Shopee realmente paga/o que o comprador realmente pagou) continuam sendo
        # somados normalmente, independente da origem do pedido.
        self.assertEqual(pedido.valor_pedido, Decimal("40.00"))
        self.assertEqual(pedido.valor_comissao, Decimal("3.00"))
        self.assertEqual(pedido.valor_cashback, Decimal("0"))

    @patch("pedidos.services.buscar_conversoes")
    def test_segue_paginacao_ate_acabar(self, mock_buscar):
        pagina1 = self._pagina(
            [{"orderId": "ORD5", "orderStatus": "PENDING", "items": [{"completeTime": None, "itemTotalCommission": "1.00"}]}],
            has_next_page=True,
            scroll_id="cursor-1",
        )
        pagina2 = self._pagina(
            [{"orderId": "ORD6", "orderStatus": "PENDING", "items": [{"completeTime": None, "itemTotalCommission": "2.00"}]}],
            has_next_page=False,
        )
        mock_buscar.side_effect = [pagina1, pagina2]

        resultado = sincronizar(1690000000, 1700000000)

        self.assertEqual(resultado["novos"], 2)
        self.assertTrue(Pedido.objects.filter(order_id="ORD5").exists())
        self.assertTrue(Pedido.objects.filter(order_id="ORD6").exists())
        mock_buscar.assert_any_call(1690000000, 1700000000, None)
        mock_buscar.assert_any_call(1690000000, 1700000000, "cursor-1")

    @patch("pedidos.services.buscar_conversoes")
    def test_define_data_prevista_liberacao_ao_validar(self, mock_buscar):
        mock_buscar.return_value = self._pagina(
            [{"orderId": "ORD7", "orderStatus": "COMPLETED", "items": [{"completeTime": 1700000500, "itemTotalCommission": "5.00"}]}]
        )

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD7")
        self.assertEqual(pedido.status, Pedido.STATUS_VALIDADO)
        esperado = calcular_data_prevista_liberacao(pedido.data_validacao)
        self.assertEqual(pedido.data_prevista_liberacao, esperado)

    @patch("pedidos.services.buscar_conversoes")
    def test_pedido_ja_liberado_nao_regride_ao_ressincronizar(self, mock_buscar):
        mock_buscar.return_value = self._pagina(
            [{"orderId": "ORD8", "orderStatus": "COMPLETED", "items": [{"completeTime": 1700000500, "itemTotalCommission": "5.00"}]}]
        )
        sincronizar(1690000000, 1700000000)
        Pedido.objects.filter(order_id="ORD8").update(
            status=Pedido.STATUS_LIBERADO, data_liberacao=timezone.now()
        )

        # A Shopee continua reportando COMPLETED (ela não sabe que já liberamos o saldo).
        resultado = sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD8")
        self.assertEqual(pedido.status, Pedido.STATUS_LIBERADO)
        self.assertIsNotNone(pedido.data_liberacao)
        self.assertEqual(resultado, {"novos": 0, "atualizados": 1, "nao_identificados": 0})

    @patch("pedidos.services.buscar_conversoes")
    def test_guarda_nome_imagem_do_produto_e_motivo_de_cancelamento(self, mock_buscar):
        mock_buscar.return_value = self._pagina(
            [
                {
                    "orderId": "ORD9",
                    "orderStatus": "CANCELLED",
                    "items": [
                        {
                            "completeTime": None,
                            "itemTotalCommission": "0",
                            "itemName": "Fone de ouvido Bluetooth",
                            "imageUrl": "https://cf.shopee.com.br/file/foto.jpg",
                            "fraudReason": "Pedido cancelado pelo comprador",
                        }
                    ],
                }
            ]
        )

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD9")
        self.assertEqual(pedido.status, Pedido.STATUS_CANCELADO)
        self.assertEqual(pedido.produto_nome, "Fone de ouvido Bluetooth")
        self.assertEqual(pedido.produto_imagem_url, "https://cf.shopee.com.br/file/foto.jpg")
        self.assertEqual(pedido.motivo_cancelamento, "Pedido cancelado pelo comprador")

    @patch("pedidos.services.buscar_conversoes")
    def test_pedido_valido_nao_tem_motivo_de_cancelamento(self, mock_buscar):
        mock_buscar.return_value = self._pagina(
            [{"orderId": "ORD10", "orderStatus": "PENDING", "items": [{"completeTime": None, "itemTotalCommission": "1.00"}]}]
        )

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD10")
        self.assertEqual(pedido.motivo_cancelamento, "")

    @patch("pedidos.services.buscar_conversoes")
    def test_sincroniza_muitos_pedidos_com_poucas_consultas_ao_banco(self, mock_buscar):
        # Uma conta com uso real pode ter milhares de pedidos - se processarmos um por
        # um, cada request estoura o tempo limite do servidor em produção (foi o que
        # aconteceu). Esse teste garante que o número de consultas não cresce junto
        # com a quantidade de pedidos.
        quantidade = 300
        orders = [
            {
                "orderId": f"ORD-LOTE-{i}",
                "orderStatus": "PENDING",
                "items": [{"completeTime": None, "itemTotalCommission": "1.00"}],
            }
            for i in range(quantidade)
        ]
        mock_buscar.return_value = self._pagina(orders)

        with CaptureQueriesContext(connection) as contexto:
            resultado = sincronizar(1690000000, 1700000000)

        self.assertEqual(resultado["novos"], quantidade)
        self.assertEqual(Pedido.objects.count(), quantidade)
        self.assertLess(len(contexto), 20)


class CampanhaCashbackModelTests(TestCase):
    """CampanhaCashback.multiplicador_em substitui o antigo CASHBACK_MULTIPLICADOR_CAMPANHA
    fixo no .env - o multiplicador de um pedido passa a depender da data_compra real,
    não do momento em que a sincronização roda (ver ROADMAP.md, Fase 44)."""

    def test_sem_nenhuma_campanha_cadastrada_multiplicador_e_1(self):
        momento = datetime(2024, 6, 1, tzinfo=dt_timezone.utc)
        self.assertEqual(CampanhaCashback.multiplicador_em(momento), Decimal("1"))

    def test_momento_none_multiplicador_e_1(self):
        self.assertEqual(CampanhaCashback.multiplicador_em(None), Decimal("1"))

    def test_momento_dentro_da_janela_usa_o_multiplicador_da_campanha(self):
        CampanhaCashback.objects.create(
            multiplicador=Decimal("2"),
            inicio=datetime(2024, 6, 1, tzinfo=dt_timezone.utc),
            fim=datetime(2024, 6, 30, tzinfo=dt_timezone.utc),
        )
        momento = datetime(2024, 6, 15, tzinfo=dt_timezone.utc)
        self.assertEqual(CampanhaCashback.multiplicador_em(momento), Decimal("2"))

    def test_momento_antes_do_inicio_nao_conta(self):
        CampanhaCashback.objects.create(
            multiplicador=Decimal("2"),
            inicio=datetime(2024, 6, 1, tzinfo=dt_timezone.utc),
            fim=datetime(2024, 6, 30, tzinfo=dt_timezone.utc),
        )
        momento = datetime(2024, 5, 31, 23, 59, tzinfo=dt_timezone.utc)
        self.assertEqual(CampanhaCashback.multiplicador_em(momento), Decimal("1"))

    def test_momento_depois_do_fim_nao_conta(self):
        CampanhaCashback.objects.create(
            multiplicador=Decimal("2"),
            inicio=datetime(2024, 6, 1, tzinfo=dt_timezone.utc),
            fim=datetime(2024, 6, 30, tzinfo=dt_timezone.utc),
        )
        momento = datetime(2024, 7, 1, tzinfo=dt_timezone.utc)
        self.assertEqual(CampanhaCashback.multiplicador_em(momento), Decimal("1"))

    def test_sem_data_de_fim_continua_valendo_indefinidamente(self):
        CampanhaCashback.objects.create(
            multiplicador=Decimal("2"), inicio=datetime(2024, 6, 1, tzinfo=dt_timezone.utc), fim=None,
        )
        momento = datetime(2030, 1, 1, tzinfo=dt_timezone.utc)
        self.assertEqual(CampanhaCashback.multiplicador_em(momento), Decimal("2"))

    def test_multiplicador_atual_usa_agora(self):
        CampanhaCashback.objects.create(
            multiplicador=Decimal("3"),
            inicio=timezone.now() - timedelta(days=1),
            fim=timezone.now() + timedelta(days=1),
        )
        self.assertEqual(CampanhaCashback.multiplicador_atual(), Decimal("3"))


@override_settings(
    SHOPEE_AFFILIATE_APP_ID="app123", SHOPEE_AFFILIATE_SECRET="segredo123",
    SHOPEE_CASHBACK_PERCENTUAL=100,
    CASHBACK_MINIMO_VENDA_DIRETA=1.6, CASHBACK_MINIMO_VENDA_INDIRETA=1,
)
class CashbackMinimoGarantidoTests(TestCase):
    """Quando a comissão real da Shopee resultaria em menos que o piso mínimo
    garantido, vale o piso - a única parte do cálculo em que a cash-b pode pagar mais
    do que recebeu de comissão (o resto é sempre uma fração da comissão real, nunca um
    prejuízo). Venda direta (link/vitrine) só tem o piso maior quando o item comprado
    é comprovadamente o mesmo do link/card clicado (Click.item_id_alvo) - ver
    ROADMAP.md, Fase 41."""

    def setUp(self):
        self.usuario = get_user_model().objects.create_user(
            username="compradora", password="senha123", cpf="39053344705"
        )

    def _pagina(self, click, order_id, comissao, valor_item, item_id=1):
        return {
            "nodes": [
                {
                    "conversionId": "999",
                    "purchaseTime": 1700000000,
                    "utmContent": f"{click.sub_id_usuario()},{click.sub_id_click()}",
                    "orders": [
                        {
                            "orderId": order_id,
                            "orderStatus": "PENDING",
                            "items": [
                                {
                                    "itemId": item_id,
                                    "completeTime": None,
                                    "itemTotalCommission": comissao,
                                    "actualAmount": valor_item,
                                }
                            ],
                        }
                    ],
                }
            ],
            "pageInfo": {"hasNextPage": False, "scrollId": ""},
        }

    @patch("pedidos.services.buscar_conversoes")
    def test_venda_direta_com_comissao_baixa_usa_o_piso_de_1_6_por_cento(self, mock_buscar):
        click = Click.objects.create(
            usuario=self.usuario, tipo=Click.TIPO_PRODUTO, item_id_alvo=1,
            url_original="https://shopee.com.br/produto-i.1.1", link_gerado="https://shope.ee/abc",
        )
        # Comissão real (0,50) seria só 0,5% dos 100 - abaixo do piso de 1,6% (1,60).
        # item_id do pedido (1) bate com o item_id_alvo do click - venda direta de verdade.
        mock_buscar.return_value = self._pagina(click, "ORD-DIRETA-BAIXA", "0.50", "100.00", item_id=1)

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-DIRETA-BAIXA")
        self.assertEqual(pedido.valor_comissao, Decimal("0.50"))
        self.assertEqual(pedido.valor_cashback, Decimal("1.60"))

    @patch("pedidos.services.buscar_conversoes")
    def test_vitrine_tambem_conta_como_venda_direta_pro_piso_quando_item_bate(self, mock_buscar):
        click = Click.objects.create(
            usuario=self.usuario, tipo=Click.TIPO_VITRINE, item_id_alvo=1,
            url_original="https://shopee.com.br/produto-i.1.1", link_gerado="https://shope.ee/abc",
        )
        mock_buscar.return_value = self._pagina(click, "ORD-VITRINE-BAIXA", "0.50", "100.00", item_id=1)

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-VITRINE-BAIXA")
        self.assertEqual(pedido.valor_cashback, Decimal("1.60"))

    @patch("pedidos.services.buscar_conversoes")
    def test_venda_indireta_com_comissao_baixa_usa_o_piso_de_1_por_cento(self, mock_buscar):
        click = Click.objects.create(
            usuario=self.usuario, tipo=Click.TIPO_HOME,
            url_original="https://shopee.com.br/", link_gerado="https://shope.ee/abc",
        )
        mock_buscar.return_value = self._pagina(click, "ORD-INDIRETA-BAIXA", "0.50", "100.00", item_id=42)

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-INDIRETA-BAIXA")
        self.assertEqual(pedido.valor_cashback, Decimal("1.00"))

    @patch("pedidos.services.buscar_conversoes")
    def test_comissao_acima_do_piso_nao_e_afetada(self, mock_buscar):
        click = Click.objects.create(
            usuario=self.usuario, tipo=Click.TIPO_PRODUTO, item_id_alvo=1,
            url_original="https://shopee.com.br/produto-i.1.1", link_gerado="https://shope.ee/abc",
        )
        # Comissão real de 5,00 (5% dos 100) já é bem maior que o piso de 1,6%.
        mock_buscar.return_value = self._pagina(click, "ORD-ACIMA-DO-PISO", "5.00", "100.00", item_id=1)

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-ACIMA-DO-PISO")
        self.assertEqual(pedido.valor_cashback, Decimal("5.00"))

    @patch("pedidos.services.buscar_conversoes")
    def test_multiplicador_de_campanha_tambem_dobra_o_piso(self, mock_buscar):
        # purchaseTime da _pagina é sempre 1700000000 - campanha cobrindo esse instante.
        CampanhaCashback.objects.create(
            multiplicador=Decimal("2"),
            inicio=datetime(2023, 1, 1, tzinfo=dt_timezone.utc),
            fim=datetime(2023, 12, 31, tzinfo=dt_timezone.utc),
        )
        click = Click.objects.create(
            usuario=self.usuario, tipo=Click.TIPO_PRODUTO, item_id_alvo=1,
            url_original="https://shopee.com.br/produto-i.1.1", link_gerado="https://shope.ee/abc",
        )
        mock_buscar.return_value = self._pagina(click, "ORD-PISO-CAMPANHA", "0.50", "100.00", item_id=1)

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-PISO-CAMPANHA")
        # Piso (1,60) x multiplicador de campanha (2) = 3,20.
        self.assertEqual(pedido.valor_cashback, Decimal("3.20"))

    @patch("pedidos.services.buscar_conversoes")
    def test_pedido_sem_click_nao_aplica_piso_nenhum(self, mock_buscar):
        click = Click.objects.create(
            usuario=self.usuario, tipo=Click.TIPO_PRODUTO, item_id_alvo=1,
            url_original="https://shopee.com.br/produto-i.1.1", link_gerado="https://shope.ee/abc",
        )
        pagina = self._pagina(click, "ORD-SEM-CLICK", "0.50", "100.00", item_id=1)
        pagina["nodes"][0]["utmContent"] = "origem-desconhecida"
        mock_buscar.return_value = pagina

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-SEM-CLICK")
        self.assertIsNone(pedido.usuario)
        self.assertEqual(pedido.valor_cashback, Decimal("0"))

    @patch("pedidos.services.buscar_conversoes")
    def test_comprar_produto_diferente_do_link_so_ganha_o_piso_indireto(self, mock_buscar):
        """O golpe que a Fase 41 fecha: converter o link de QUALQUER produto e comprar
        outro completamente diferente não deveria destravar o piso maior de venda
        direta."""
        click = Click.objects.create(
            usuario=self.usuario, tipo=Click.TIPO_PRODUTO, item_id_alvo=1,
            url_original="https://shopee.com.br/produto-i.1.1", link_gerado="https://shope.ee/abc",
        )
        # Comprou o item 999, não o item 1 que estava no link.
        mock_buscar.return_value = self._pagina(click, "ORD-ITEM-DIFERENTE", "0.50", "100.00", item_id=999)

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-ITEM-DIFERENTE")
        self.assertEqual(pedido.valor_cashback, Decimal("1.00"))

    @patch("pedidos.services.buscar_conversoes")
    def test_click_sem_item_id_alvo_identificado_so_ganha_o_piso_indireto(self, mock_buscar):
        """Link curto que não deu pra identificar na hora do clique (ver Fase 35) não
        consegue provar o vínculo com o item comprado - conta como não bate, por
        segurança (mais seguro errar pro lado de baixo)."""
        click = Click.objects.create(
            usuario=self.usuario, tipo=Click.TIPO_PRODUTO, item_id_alvo=None,
            url_original="https://s.shopee.com.br/abc123", link_gerado="https://shope.ee/abc",
        )
        mock_buscar.return_value = self._pagina(click, "ORD-SEM-ITEM-ALVO", "0.50", "100.00", item_id=1)

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-SEM-ITEM-ALVO")
        self.assertEqual(pedido.valor_cashback, Decimal("1.00"))

    @patch("pedidos.services.buscar_conversoes")
    def test_pedido_com_varios_itens_so_o_que_bate_ganha_o_piso_maior(self, mock_buscar):
        click = Click.objects.create(
            usuario=self.usuario, tipo=Click.TIPO_PRODUTO, item_id_alvo=1,
            url_original="https://shopee.com.br/produto-i.1.1", link_gerado="https://shope.ee/abc",
        )
        pagina = {
            "nodes": [
                {
                    "conversionId": "999",
                    "purchaseTime": 1700000000,
                    "utmContent": f"{click.sub_id_usuario()},{click.sub_id_click()}",
                    "orders": [
                        {
                            "orderId": "ORD-MULTI-ITEM",
                            "orderStatus": "PENDING",
                            "items": [
                                {
                                    "itemId": 1, "completeTime": None,
                                    "itemTotalCommission": "0.50", "actualAmount": "100.00",
                                },
                                {
                                    "itemId": 2, "completeTime": None,
                                    "itemTotalCommission": "0.30", "actualAmount": "100.00",
                                },
                            ],
                        }
                    ],
                }
            ],
            "pageInfo": {"hasNextPage": False, "scrollId": ""},
        }
        mock_buscar.return_value = pagina

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-MULTI-ITEM")
        # Item 1 (bate com o link): piso de 1,6% de 100 = 1,60. Item 2 (não bate):
        # piso de 1% de 100 = 1,00. Total: 2,60.
        self.assertEqual(pedido.valor_cashback, Decimal("2.60"))


@override_settings(SHOPEE_CASHBACK_PERCENTUAL=20)
class MultiplicadorCampanhaTests(TestCase):
    """O multiplicador de campanha fica congelado no pedido quando ele é registrado.

    Sem isso, como a Shopee reenvia o mesmo pedido em toda sincronização seguinte e o
    cashback é recalculado do zero a cada vez, uma campanha de "cashback em dobro"
    seria desfeita sozinha assim que acabasse.
    """

    def setUp(self):
        self.usuario = get_user_model().objects.create_user(
            username="compradora", password="senha123", cpf="39053344705"
        )
        self.click = Click.objects.create(
            usuario=self.usuario,
            tipo=Click.TIPO_HOME,
            url_original="https://shopee.com.br/",
            link_gerado="https://shope.ee/abc",
        )

    def _pagina(self, order_id, status="PENDING"):
        return {
            "nodes": [
                {
                    "conversionId": "999",
                    "purchaseTime": 1700000000,
                    "utmContent": f"{self.click.sub_id_usuario()},{self.click.sub_id_click()}",
                    "orders": [
                        {
                            "orderId": order_id,
                            "orderStatus": status,
                            "items": [{"completeTime": None, "itemTotalCommission": "20.00"}],
                        }
                    ],
                }
            ],
            "pageInfo": {"hasNextPage": False, "scrollId": ""},
        }

    def _campanha(self, inicio=None, fim=None, multiplicador="2"):
        return CampanhaCashback.objects.create(
            multiplicador=Decimal(multiplicador),
            inicio=inicio or datetime(2023, 1, 1, tzinfo=dt_timezone.utc),
            fim=fim,
        )

    @patch("pedidos.services.buscar_conversoes")
    def test_pedido_comprado_durante_a_campanha_recebe_o_dobro(self, mock_buscar):
        self._campanha()  # cobre o purchaseTime fixo da _pagina (1700000000)
        mock_buscar.return_value = self._pagina("ORD-CAMP-1")

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-CAMP-1")
        # 20% de R$20 = R$4, dobrado pela campanha = R$8
        self.assertEqual(pedido.valor_cashback, Decimal("8.00"))
        self.assertEqual(pedido.multiplicador_campanha, Decimal("2"))

    @patch("pedidos.services.buscar_conversoes")
    def test_campanha_encerrada_nao_reduz_cashback_de_pedido_ja_registrado(self, mock_buscar):
        campanha = self._campanha()
        mock_buscar.return_value = self._pagina("ORD-CAMP-2")
        sincronizar(1690000000, 1700000000)

        # Campanha acabou e a Shopee reenvia o mesmo pedido, agora validado - o
        # multiplicador gravado na primeira vez continua valendo, não importa se a
        # campanha ainda existe ou não nesse momento.
        campanha.delete()
        mock_buscar.return_value = self._pagina("ORD-CAMP-2", status="COMPLETED")
        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-CAMP-2")
        self.assertEqual(pedido.status, Pedido.STATUS_VALIDADO)
        self.assertEqual(pedido.valor_cashback, Decimal("8.00"))
        self.assertEqual(pedido.multiplicador_campanha, Decimal("2"))

    @patch("pedidos.services.buscar_conversoes")
    def test_campanha_nova_nao_dobra_retroativamente_pedido_antigo(self, mock_buscar):
        mock_buscar.return_value = self._pagina("ORD-CAMP-3")
        sincronizar(1690000000, 1700000000)

        # Campanha começa depois, mas cobre a mesma data_compra do pedido antigo (que
        # ainda está na janela de sincronização) - mesmo assim não dobra
        # retroativamente, porque o multiplicador já ficou gravado no pedido.
        self._campanha()
        mock_buscar.return_value = self._pagina("ORD-CAMP-3", status="COMPLETED")
        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-CAMP-3")
        self.assertEqual(pedido.valor_cashback, Decimal("4.00"))
        self.assertEqual(pedido.multiplicador_campanha, Decimal("1"))

    @patch("pedidos.services.buscar_conversoes")
    def test_pedido_registrado_fora_da_campanha_guarda_multiplicador_1(self, mock_buscar):
        mock_buscar.return_value = self._pagina("ORD-CAMP-4")

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-CAMP-4")
        self.assertEqual(pedido.valor_cashback, Decimal("4.00"))
        self.assertEqual(pedido.multiplicador_campanha, Decimal("1"))


class SincronizarBonusIndicacaoTests(TestCase):
    def setUp(self):
        self.indicador = get_user_model().objects.create_user(
            username="indicador", password="senha123", cpf="39053344705"
        )
        self.indicado = get_user_model().objects.create_user(
            username="indicado", password="senha123", cpf="14783246947"
        )
        self.indicacao = Indicacao.objects.create(indicador=self.indicador, indicado=self.indicado)

        self.click_indicador = Click.objects.create(
            usuario=self.indicador, tipo=Click.TIPO_HOME,
            url_original="https://shopee.com.br/", link_gerado="https://shope.ee/indicador",
        )
        self.click_indicado = Click.objects.create(
            usuario=self.indicado, tipo=Click.TIPO_HOME,
            url_original="https://shopee.com.br/", link_gerado="https://shope.ee/indicado",
        )

    def _no(self, click, order_id, comissao, order_status="COMPLETED", purchase_time=1700000000, complete_time=1700000500):
        return {
            "conversionId": order_id,
            "purchaseTime": purchase_time,
            "utmContent": f"{click.sub_id_usuario()},{click.sub_id_click()}",
            "orders": [
                {"orderId": order_id, "orderStatus": order_status, "items": [{"completeTime": complete_time, "itemTotalCommission": comissao}]}
            ],
        }

    def _pagina(self, nodes):
        return {"nodes": nodes, "pageInfo": {"hasNextPage": False, "scrollId": ""}}

    @patch("pedidos.services.buscar_conversoes")
    def test_pedido_em_campanha_nao_consome_o_bonus_de_indicacao(self, mock_buscar):
        CampanhaCashback.objects.create(
            multiplicador=Decimal("2"), inicio=datetime(2023, 1, 1, tzinfo=dt_timezone.utc),
        )
        # O pedido leva só o extra da campanha (5.00 x 2 = 10.00), e a indicação
        # continua pendente esperando a campanha acabar.
        mock_buscar.return_value = self._pagina([self._no(self.click_indicado, "ORD-IND-CAMP", "5.00")])

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-IND-CAMP")
        self.assertEqual(pedido.multiplicador_campanha, Decimal("2"))
        self.assertEqual(pedido.valor_cashback, Decimal("10.00"))
        self.indicacao.refresh_from_db()
        self.assertIsNone(self.indicacao.pedido_bonus_indicado)

    @patch("pedidos.services.buscar_conversoes")
    def test_bonus_na_fila_entra_no_proximo_pedido_depois_da_campanha(self, mock_buscar):
        # Campanha cobre só até o purchaseTime do 1º pedido (1700000000) - o 2º pedido
        # (purchase_time=1700000600) já compra depois dela ter acabado.
        CampanhaCashback.objects.create(
            multiplicador=Decimal("2"),
            inicio=datetime(2023, 1, 1, tzinfo=dt_timezone.utc),
            fim=datetime.fromtimestamp(1700000000, tz=dt_timezone.utc),
        )
        mock_buscar.return_value = self._pagina([self._no(self.click_indicado, "ORD-FILA-1", "5.00")])
        sincronizar(1690000000, 1700000000)

        # O pedido seguinte do indicado, já fora da campanha, pega o bônus que ficou na fila.
        mock_buscar.return_value = self._pagina(
            [
                self._no(self.click_indicado, "ORD-FILA-1", "5.00"),
                self._no(self.click_indicado, "ORD-FILA-2", "5.00", purchase_time=1700000600),
            ]
        )
        sincronizar(1690000000, 1700000000)

        # O pedido da campanha continua com o valor da campanha, sem ganhar o bônus depois.
        pedido_campanha = Pedido.objects.get(order_id="ORD-FILA-1")
        self.assertEqual(pedido_campanha.valor_cashback, Decimal("10.00"))

        pedido_pos_campanha = Pedido.objects.get(order_id="ORD-FILA-2")
        self.assertEqual(pedido_pos_campanha.valor_cashback, Decimal("10.00"))  # 5.00 x 2 de indicação
        self.indicacao.refresh_from_db()
        self.assertEqual(self.indicacao.pedido_bonus_indicado, pedido_pos_campanha)

    @patch("pedidos.services.buscar_conversoes")
    def test_primeira_compra_validada_do_indicado_dobra_cashback(self, mock_buscar):
        mock_buscar.return_value = self._pagina([self._no(self.click_indicado, "ORD-IND-1", "10.00")])

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-IND-1")
        self.assertEqual(pedido.valor_cashback, Decimal("20.00"))
        self.indicacao.refresh_from_db()
        self.assertEqual(self.indicacao.pedido_bonus_indicado, pedido)
        self.assertIsNone(self.indicacao.pedido_bonus_indicador)

    @patch("pedidos.services.buscar_conversoes")
    def test_primeira_compra_validada_do_indicado_notifica_o_indicador_por_email(self, mock_buscar):
        self.indicador.email = "indicador@example.com"
        self.indicador.save()
        mock_buscar.return_value = self._pagina([self._no(self.click_indicado, "ORD-IND-EMAIL", "10.00")])

        sincronizar(1690000000, 1700000000)

        self.assertEqual(len(mail.outbox), 1)
        email = mail.outbox[0]
        self.assertEqual(email.to, ["indicador@example.com"])
        self.assertIn(self.indicado.username, email.body)
        self.assertIn("dobro", email.subject.lower())

    @patch("pedidos.services.buscar_conversoes")
    def test_ressincronizar_o_mesmo_pedido_nao_notifica_de_novo(self, mock_buscar):
        self.indicador.email = "indicador@example.com"
        self.indicador.save()
        mock_buscar.return_value = self._pagina([self._no(self.click_indicado, "ORD-IND-EMAIL2", "10.00")])

        sincronizar(1690000000, 1700000000)
        sincronizar(1690000000, 1700000000)

        self.assertEqual(len(mail.outbox), 1)

    @patch("pedidos.services.buscar_conversoes")
    def test_indicador_ganha_dobro_na_proxima_compra_apos_indicado_validar(self, mock_buscar):
        mock_buscar.return_value = self._pagina([self._no(self.click_indicado, "ORD-IND-1", "10.00")])
        sincronizar(1690000000, 1700000000)

        mock_buscar.return_value = self._pagina([self._no(self.click_indicador, "ORD-REF-1", "5.00")])
        sincronizar(1690000000, 1700000000)

        pedido_indicador = Pedido.objects.get(order_id="ORD-REF-1")
        self.assertEqual(pedido_indicador.valor_cashback, Decimal("10.00"))
        self.indicacao.refresh_from_db()
        self.assertEqual(self.indicacao.pedido_bonus_indicador, pedido_indicador)

    @patch("pedidos.services.buscar_conversoes")
    def test_segunda_compra_do_indicador_nao_recebe_bonus_de_novo(self, mock_buscar):
        mock_buscar.return_value = self._pagina([self._no(self.click_indicado, "ORD-IND-1", "10.00")])
        sincronizar(1690000000, 1700000000)
        mock_buscar.return_value = self._pagina([self._no(self.click_indicador, "ORD-REF-1", "5.00")])
        sincronizar(1690000000, 1700000000)

        mock_buscar.return_value = self._pagina([self._no(self.click_indicador, "ORD-REF-2", "7.00")])
        sincronizar(1690000000, 1700000000)

        pedido_seguinte = Pedido.objects.get(order_id="ORD-REF-2")
        self.assertEqual(pedido_seguinte.valor_cashback, Decimal("7.00"))

    @patch("pedidos.services.buscar_conversoes")
    def test_ressincronizar_o_mesmo_pedido_nao_dobra_de_novo(self, mock_buscar):
        mock_buscar.return_value = self._pagina([self._no(self.click_indicado, "ORD-IND-1", "10.00")])
        sincronizar(1690000000, 1700000000)
        # A Shopee reenvia o mesmo pedido validado em toda sincronização seguinte.
        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-IND-1")
        self.assertEqual(pedido.valor_cashback, Decimal("20.00"))

    @patch("pedidos.services.buscar_conversoes")
    def test_pedido_ainda_pendente_do_indicado_nao_recebe_bonus(self, mock_buscar):
        mock_buscar.return_value = self._pagina(
            [self._no(self.click_indicado, "ORD-IND-1", "10.00", order_status="PENDING", complete_time=None)]
        )

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-IND-1")
        self.assertEqual(pedido.valor_cashback, Decimal("10.00"))
        self.indicacao.refresh_from_db()
        self.assertIsNone(self.indicacao.pedido_bonus_indicado)

    @patch("pedidos.services.buscar_conversoes")
    def test_fila_fifo_quando_indicador_tem_duas_indicacoes_pendentes(self, mock_buscar):
        indicado2 = get_user_model().objects.create_user(
            username="indicado2", password="senha123", cpf="52914637837"
        )
        indicacao2 = Indicacao.objects.create(indicador=self.indicador, indicado=indicado2)
        click_indicado2 = Click.objects.create(
            usuario=indicado2, tipo=Click.TIPO_HOME,
            url_original="https://shopee.com.br/", link_gerado="https://shope.ee/indicado2",
        )

        mock_buscar.return_value = self._pagina(
            [
                self._no(self.click_indicado, "ORD-IND-1", "10.00"),
                self._no(click_indicado2, "ORD-IND-2", "10.00"),
            ]
        )
        sincronizar(1690000000, 1700000000)

        # O indicador faz duas compras que validam na mesma sincronização - a mais
        # antiga (por purchaseTime) deve atender a indicação mais antiga primeiro.
        mock_buscar.return_value = self._pagina(
            [
                self._no(self.click_indicador, "ORD-REF-1", "5.00", purchase_time=1700000200),
                self._no(self.click_indicador, "ORD-REF-2", "5.00", purchase_time=1700000100),
            ]
        )
        sincronizar(1690000000, 1700000000)

        self.indicacao.refresh_from_db()
        indicacao2.refresh_from_db()
        self.assertEqual(self.indicacao.pedido_bonus_indicador.order_id, "ORD-REF-2")
        self.assertEqual(indicacao2.pedido_bonus_indicador.order_id, "ORD-REF-1")

    @patch("pedidos.services.buscar_conversoes")
    def test_usuario_sem_indicacao_nao_e_afetado(self, mock_buscar):
        avulso = get_user_model().objects.create_user(username="avulso", password="senha123", cpf="94834869092")
        click_avulso = Click.objects.create(
            usuario=avulso, tipo=Click.TIPO_HOME,
            url_original="https://shopee.com.br/", link_gerado="https://shope.ee/avulso",
        )
        mock_buscar.return_value = self._pagina([self._no(click_avulso, "ORD-AVULSO-1", "8.00")])

        sincronizar(1690000000, 1700000000)

        pedido = Pedido.objects.get(order_id="ORD-AVULSO-1")
        self.assertEqual(pedido.valor_cashback, Decimal("8.00"))


class NotificarIndicadorBonusPendenteTests(TestCase):
    def setUp(self):
        self.indicador = get_user_model().objects.create_user(
            username="indicador", password="senha123", cpf="39053344705", email="indicador@example.com"
        )
        self.indicado = get_user_model().objects.create_user(
            username="indicado", password="senha123", cpf="14783246947"
        )
        self.indicacao = Indicacao.objects.create(indicador=self.indicador, indicado=self.indicado)

    def test_envia_email_pro_indicador_com_o_nome_do_indicado(self):
        notificar_indicador_bonus_pendente(self.indicacao)

        self.assertEqual(len(mail.outbox), 1)
        email = mail.outbox[0]
        self.assertEqual(email.to, ["indicador@example.com"])
        self.assertIn(self.indicado.username, email.body)
        self.assertIn("dobro", email.subject.lower())

    def test_indicador_sem_email_nao_envia_nada(self):
        self.indicador.email = ""
        self.indicador.save()

        notificar_indicador_bonus_pendente(self.indicacao)

        self.assertEqual(len(mail.outbox), 0)


class LiberarSaldoTests(TestCase):
    def setUp(self):
        self.usuario = get_user_model().objects.create_user(
            username="compradora", password="senha123", cpf="39053344705"
        )

    def _criar_pedido(self, order_id, status, data_prevista_liberacao):
        return Pedido.objects.create(
            order_id=order_id,
            conversion_id="1",
            usuario=self.usuario,
            status=status,
            status_shopee_bruto="COMPLETED",
            valor_comissao=Decimal("10.00"),
            valor_cashback=Decimal("10.00"),
            data_prevista_liberacao=data_prevista_liberacao,
        )

    def test_libera_pedido_validado_com_data_ja_vencida(self):
        ontem = timezone.localdate() - timedelta(days=1)
        pedido = self._criar_pedido("ORD-VENCIDO", Pedido.STATUS_VALIDADO, ontem)

        total = liberar_saldo()

        pedido.refresh_from_db()
        self.assertEqual(total, 1)
        self.assertEqual(pedido.status, Pedido.STATUS_LIBERADO)
        self.assertIsNotNone(pedido.data_liberacao)

    def test_nao_libera_pedido_com_data_futura(self):
        amanha = timezone.localdate() + timedelta(days=1)
        pedido = self._criar_pedido("ORD-FUTURO", Pedido.STATUS_VALIDADO, amanha)

        total = liberar_saldo()

        pedido.refresh_from_db()
        self.assertEqual(total, 0)
        self.assertEqual(pedido.status, Pedido.STATUS_VALIDADO)
        self.assertIsNone(pedido.data_liberacao)

    def test_nao_mexe_em_pedido_pendente_ou_cancelado_mesmo_com_data_vencida(self):
        ontem = timezone.localdate() - timedelta(days=1)
        pendente = self._criar_pedido("ORD-PENDENTE", Pedido.STATUS_PENDENTE, ontem)
        cancelado = self._criar_pedido("ORD-CANCELADO", Pedido.STATUS_CANCELADO, ontem)

        total = liberar_saldo()

        pendente.refresh_from_db()
        cancelado.refresh_from_db()
        self.assertEqual(total, 0)
        self.assertEqual(pendente.status, Pedido.STATUS_PENDENTE)
        self.assertEqual(cancelado.status, Pedido.STATUS_CANCELADO)


class ObterAnalyticsTests(TestCase):
    def setUp(self):
        self.usuario = get_user_model().objects.create_user(
            username="compradora", password="senha123", cpf="39053344705"
        )

    def _criar_pedido(self, order_id, status, comissao, cashback, data_compra, usuario=None, click=None):
        return Pedido.objects.create(
            order_id=order_id,
            conversion_id="1",
            usuario=usuario if usuario is not None else self.usuario,
            click=click,
            status=status,
            status_shopee_bruto="COMPLETED",
            valor_comissao=Decimal(comissao),
            valor_cashback=Decimal(cashback),
            data_compra=data_compra,
        )

    def test_filtro_origem_site_traz_so_pedidos_com_click(self):
        click = Click.objects.create(
            usuario=self.usuario, tipo=Click.TIPO_HOME,
            url_original="https://shopee.com.br/", link_gerado="https://shope.ee/abc",
        )
        self._criar_pedido("COM-CLICK", Pedido.STATUS_VALIDADO, "10.00", "5.00", timezone.now(), click=click)
        self._criar_pedido("SEM-CLICK", Pedido.STATUS_VALIDADO, "10.00", "5.00", timezone.now())

        dados = obter_analytics(origem="site")

        self.assertEqual(dados["total_pedidos"], 1)
        self.assertEqual(dados["total_comissao"], Decimal("10.00"))

    def test_filtro_origem_fora_traz_so_pedidos_sem_click(self):
        click = Click.objects.create(
            usuario=self.usuario, tipo=Click.TIPO_HOME,
            url_original="https://shopee.com.br/", link_gerado="https://shope.ee/abc",
        )
        self._criar_pedido("COM-CLICK", Pedido.STATUS_VALIDADO, "10.00", "5.00", timezone.now(), click=click)
        self._criar_pedido("SEM-CLICK", Pedido.STATUS_VALIDADO, "20.00", "8.00", timezone.now())

        dados = obter_analytics(origem="fora")

        self.assertEqual(dados["total_pedidos"], 1)
        self.assertEqual(dados["total_comissao"], Decimal("20.00"))

    def test_sem_filtro_origem_traz_tudo(self):
        click = Click.objects.create(
            usuario=self.usuario, tipo=Click.TIPO_HOME,
            url_original="https://shopee.com.br/", link_gerado="https://shope.ee/abc",
        )
        self._criar_pedido("COM-CLICK", Pedido.STATUS_VALIDADO, "10.00", "5.00", timezone.now(), click=click)
        self._criar_pedido("SEM-CLICK", Pedido.STATUS_VALIDADO, "20.00", "8.00", timezone.now())

        dados = obter_analytics()

        self.assertEqual(dados["total_pedidos"], 2)

    def test_soma_comissao_e_cashback_de_todos_os_pedidos(self):
        self._criar_pedido("A1", Pedido.STATUS_VALIDADO, "20.00", "10.00", timezone.now())
        self._criar_pedido("A2", Pedido.STATUS_PENDENTE, "8.00", "4.00", timezone.now())

        dados = obter_analytics()

        self.assertEqual(dados["total_comissao"], Decimal("28.00"))
        self.assertEqual(dados["total_cashback"], Decimal("14.00"))
        self.assertEqual(dados["total_pedidos"], 2)
        self.assertEqual(dados["margem_retida"], Decimal("14.00"))

    def test_filtro_por_periodo_exclui_pedido_fora_do_intervalo(self):
        dentro = timezone.make_aware(datetime(2026, 3, 15))
        fora = timezone.make_aware(datetime(2026, 5, 1))
        self._criar_pedido("DENTRO", Pedido.STATUS_VALIDADO, "10.00", "5.00", dentro)
        self._criar_pedido("FORA", Pedido.STATUS_VALIDADO, "10.00", "5.00", fora)

        dados = obter_analytics(data_inicio=date(2026, 3, 1), data_fim=date(2026, 3, 31))

        self.assertEqual(dados["total_pedidos"], 1)
        self.assertEqual(dados["total_comissao"], Decimal("10.00"))

    def test_filtro_por_status(self):
        self._criar_pedido("VALIDADO", Pedido.STATUS_VALIDADO, "10.00", "5.00", timezone.now())
        self._criar_pedido("CANCELADO", Pedido.STATUS_CANCELADO, "10.00", "5.00", timezone.now())

        dados = obter_analytics(status=Pedido.STATUS_VALIDADO)

        self.assertEqual(dados["total_pedidos"], 1)

    def test_saldo_a_liberar_soma_pendente_e_validado_mas_nao_liberado(self):
        self._criar_pedido("PEND", Pedido.STATUS_PENDENTE, "10.00", "5.00", timezone.now())
        self._criar_pedido("VALID", Pedido.STATUS_VALIDADO, "10.00", "3.00", timezone.now())
        self._criar_pedido("LIB", Pedido.STATUS_LIBERADO, "10.00", "7.00", timezone.now())

        dados = obter_analytics()

        self.assertEqual(dados["saldo_a_liberar"], Decimal("8.00"))
        self.assertEqual(dados["saldo_liberado"], Decimal("7.00"))

    def test_resumo_status_inclui_status_sem_nenhum_pedido_zerado(self):
        self._criar_pedido("VALID", Pedido.STATUS_VALIDADO, "10.00", "5.00", timezone.now())

        dados = obter_analytics()
        por_chave = {linha["status"]: linha for linha in dados["resumo_status"]}

        self.assertEqual(por_chave[Pedido.STATUS_CANCELADO]["total"], 0)
        self.assertEqual(por_chave[Pedido.STATUS_CANCELADO]["cashback"], Decimal("0"))

    def test_ranking_indicadores_conta_indicacoes_e_concluidas(self):
        indicador = get_user_model().objects.create_user(username="indicador", password="senha123", cpf="14783246947")
        indicado1 = get_user_model().objects.create_user(username="ind1", password="senha123", cpf="52914637837")
        indicado2 = get_user_model().objects.create_user(username="ind2", password="senha123", cpf="91234567873")
        pedido_bonus = self._criar_pedido("BONUS", Pedido.STATUS_VALIDADO, "10.00", "20.00", timezone.now(), usuario=indicador)
        Indicacao.objects.create(indicador=indicador, indicado=indicado1, pedido_bonus_indicado=pedido_bonus, pedido_bonus_indicador=pedido_bonus)
        Indicacao.objects.create(indicador=indicador, indicado=indicado2)

        dados = obter_analytics()

        self.assertEqual(dados["total_indicacoes"], 2)
        self.assertEqual(dados["indicacoes_concluidas"], 1)
        self.assertEqual(dados["ranking_indicadores"][0]["indicador__username"], "indicador")
        self.assertEqual(dados["ranking_indicadores"][0]["total_indicacoes"], 2)
        self.assertEqual(dados["ranking_indicadores"][0]["concluidas"], 1)

    def test_total_saques_por_status(self):
        Saque.objects.create(
            usuario=self.usuario, valor=Decimal("50.00"), chave_pix="a@a.com",
            tipo_chave_pix="EMAIL", status=Saque.STATUS_PAGO,
        )
        Saque.objects.create(
            usuario=self.usuario, valor=Decimal("30.00"), chave_pix="a@a.com",
            tipo_chave_pix="EMAIL", status=Saque.STATUS_SOLICITADO,
        )

        dados = obter_analytics()

        self.assertEqual(dados["total_saques"], 2)
        self.assertEqual(dados["total_saques_valor"], Decimal("80.00"))
        por_status = {linha["status"]: linha for linha in dados["saques_por_status"]}
        self.assertEqual(por_status[Saque.STATUS_PAGO]["valor"], Decimal("50.00"))
        self.assertEqual(por_status[Saque.STATUS_SOLICITADO]["valor"], Decimal("30.00"))


class ObterSerieDiariaTests(TestCase):
    """Série dia a dia usada no gráfico de linha da tela de analytics."""

    def setUp(self):
        self.usuario = get_user_model().objects.create_user(
            username="compradora", password="senha123", cpf="39053344705"
        )

    def _criar_pedido(self, order_id, comissao, cashback, data_compra):
        return Pedido.objects.create(
            order_id=order_id, conversion_id="1", usuario=self.usuario,
            status=Pedido.STATUS_VALIDADO, status_shopee_bruto="COMPLETED",
            valor_comissao=Decimal(comissao), valor_cashback=Decimal(cashback), data_compra=data_compra,
        )

    def test_um_ponto_por_dia_no_periodo_mesmo_sem_pedido(self):
        serie = obter_serie_diaria(data_inicio=date(2026, 3, 1), data_fim=date(2026, 3, 3))

        self.assertEqual(serie["rotulos"], ["01/03", "02/03", "03/03"])
        self.assertEqual(serie["series"]["pedidos"], [0, 0, 0])

    def test_agrupa_pedidos_por_dia_da_compra(self):
        self._criar_pedido("A1", "10.00", "5.00", timezone.make_aware(datetime(2026, 3, 1, 8)))
        self._criar_pedido("A2", "20.00", "8.00", timezone.make_aware(datetime(2026, 3, 1, 20)))
        self._criar_pedido("B1", "6.00", "3.00", timezone.make_aware(datetime(2026, 3, 2, 12)))

        serie = obter_serie_diaria(data_inicio=date(2026, 3, 1), data_fim=date(2026, 3, 3))

        self.assertEqual(serie["series"]["pedidos"], [2, 1, 0])
        self.assertEqual(serie["series"]["comissao"], [30.0, 6.0, 0])
        self.assertEqual(serie["series"]["cashback"], [13.0, 3.0, 0])

    def test_agrupa_saques_indicacoes_e_novos_usuarios_por_dia(self):
        # criado_em é auto_now_add em Saque/Indicacao - passar no create() é ignorado,
        # então cria e ajusta a data depois com update() (que não passa pelo auto_now_add).
        saque = Saque.objects.create(
            usuario=self.usuario, valor=Decimal("50.00"), chave_pix="a@a.com", tipo_chave_pix="EMAIL",
            status=Saque.STATUS_PAGO,
        )
        Saque.objects.filter(pk=saque.pk).update(criado_em=timezone.make_aware(datetime(2026, 3, 1, 10)))
        indicado = get_user_model().objects.create_user(username="indicado", password="s", cpf="14783246947")
        indicacao = Indicacao.objects.create(indicador=self.usuario, indicado=indicado)
        Indicacao.objects.filter(pk=indicacao.pk).update(criado_em=timezone.make_aware(datetime(2026, 3, 2, 9)))

        serie = obter_serie_diaria(data_inicio=date(2026, 3, 1), data_fim=date(2026, 3, 2))

        self.assertEqual(serie["series"]["saques_quantidade"], [1, 0])
        self.assertEqual(serie["series"]["saques_valor"], [50.0, 0])
        self.assertEqual(serie["series"]["indicacoes"], [0, 1])
        # self.usuario (setUp) e indicado foram criados "agora", fora do período de março -
        # não devem contar em nenhum dos dois dias filtrados.
        self.assertEqual(serie["series"]["novos_usuarios"], [0, 0])

    def test_sem_periodo_usa_os_ultimos_30_dias_terminando_hoje(self):
        serie = obter_serie_diaria()

        hoje = timezone.localdate()
        self.assertEqual(len(serie["rotulos"]), 30)
        self.assertEqual(serie["rotulos"][-1], hoje.strftime("%d/%m"))

    def test_periodo_maior_que_180_dias_e_recortado_pro_final_dele(self):
        serie = obter_serie_diaria(data_inicio=date(2026, 1, 1), data_fim=date(2026, 12, 31))

        self.assertEqual(len(serie["rotulos"]), 180)
        self.assertEqual(serie["rotulos"][-1], "31/12")

    def test_respeita_filtro_de_status_igual_ao_resto_da_tela(self):
        self._criar_pedido("VALID", "10.00", "5.00", timezone.make_aware(datetime(2026, 3, 1, 8)))
        Pedido.objects.create(
            order_id="CANC", conversion_id="2", usuario=self.usuario,
            status=Pedido.STATUS_CANCELADO, status_shopee_bruto="CANCELLED",
            valor_comissao=Decimal("10.00"), valor_cashback=Decimal("0"),
            data_compra=timezone.make_aware(datetime(2026, 3, 1, 9)),
        )

        serie = obter_serie_diaria(data_inicio=date(2026, 3, 1), data_fim=date(2026, 3, 1), status=Pedido.STATUS_VALIDADO)

        self.assertEqual(serie["series"]["pedidos"], [1])


class OrigemDetalhadaTests(TestCase):
    """Diferencia a origem de cada pedido: conversão de link direto, clique num card da
    vitrine de ofertas ou venda indireta (botão "Ir pra Shopee") - além do caso sem
    Click (pedido não gerado por aqui, ver OrigemFilter em pedidos/admin.py)."""

    def setUp(self):
        self.usuario = get_user_model().objects.create_user(
            username="compradora", password="senha123", cpf="39053344705"
        )

    def _pedido(self, order_id, click=None):
        return Pedido.objects.create(
            order_id=order_id, conversion_id="1", usuario=self.usuario, click=click,
            status=Pedido.STATUS_VALIDADO, status_shopee_bruto="COMPLETED",
        )

    def test_sem_click_e_fora_do_site(self):
        pedido = self._pedido("ORD-SEM-CLICK")
        self.assertEqual(origem_detalhada(pedido), "Fora do site")

    def test_click_tipo_produto_e_link_direto(self):
        click = Click.objects.create(
            usuario=self.usuario, tipo=Click.TIPO_PRODUTO,
            url_original="https://shopee.com.br/produto-i.1.1", link_gerado="https://shope.ee/a",
        )
        pedido = self._pedido("ORD-LINK-DIRETO", click=click)
        self.assertEqual(origem_detalhada(pedido), "Link direto")

    def test_click_tipo_vitrine_e_vitrine_de_ofertas(self):
        click = Click.objects.create(
            usuario=self.usuario, tipo=Click.TIPO_VITRINE,
            url_original="https://shopee.com.br/produto-i.2.2", link_gerado="https://shope.ee/b",
        )
        pedido = self._pedido("ORD-VITRINE", click=click)
        self.assertEqual(origem_detalhada(pedido), "Vitrine de ofertas")

    def test_click_tipo_home_e_venda_indireta(self):
        click = Click.objects.create(
            usuario=self.usuario, tipo=Click.TIPO_HOME,
            url_original="https://shopee.com.br/", link_gerado="https://shope.ee/c",
        )
        pedido = self._pedido("ORD-INDIRETA", click=click)
        self.assertEqual(origem_detalhada(pedido), "Venda indireta (Ir pra Shopee)")


class AnalyticsAdminViewTests(TestCase):
    def setUp(self):
        self.staff = get_user_model().objects.create_user(
            username="equipe", password="senha123", cpf="39053344705", is_staff=True
        )
        self.usuario_comum = get_user_model().objects.create_user(
            username="comum", password="senha123", cpf="14783246947"
        )
        Pedido.objects.create(
            order_id="ORD-1", conversion_id="1", usuario=self.usuario_comum,
            status=Pedido.STATUS_VALIDADO, status_shopee_bruto="COMPLETED",
            valor_comissao=Decimal("10.00"), valor_cashback=Decimal("5.00"), data_compra=timezone.now(),
        )

    def test_staff_acessa_a_tela_de_analytics(self):
        self.client.force_login(self.staff)
        resposta = self.client.get(reverse("admin:pedidos_analytics"))
        self.assertEqual(resposta.status_code, 200)
        self.assertContains(resposta, "Comissão total")
        self.assertContains(resposta, "Quantidade de pedidos")

    def test_usuario_comum_nao_acessa_analytics(self):
        self.client.force_login(self.usuario_comum)
        resposta = self.client.get(reverse("admin:pedidos_analytics"))
        self.assertEqual(resposta.status_code, 302)

    def test_anonimo_e_redirecionado_para_login(self):
        resposta = self.client.get(reverse("admin:pedidos_analytics"))
        self.assertEqual(resposta.status_code, 302)

    def test_exportar_csv_traz_o_pedido_filtrado(self):
        self.client.force_login(self.staff)
        resposta = self.client.get(reverse("admin:pedidos_analytics_exportar"))
        self.assertEqual(resposta.status_code, 200)
        self.assertEqual(resposta["Content-Type"], "text/csv")
        conteudo = resposta.content.decode()
        self.assertIn("ORD-1", conteudo)
        self.assertIn("comum", conteudo)

    def test_filtro_origem_site_exclui_pedido_sem_click(self):
        # ORD-1 (setUp) não tem Click vinculado - é "fora do site".
        self.client.force_login(self.staff)
        resposta = self.client.get(reverse("admin:pedidos_analytics"), {"origem": "site"})
        self.assertEqual(resposta.context["dados"]["total_pedidos"], 0)
        self.assertEqual(resposta.context["filtro_origem"], "site")

    def test_filtro_origem_fora_inclui_pedido_sem_click(self):
        self.client.force_login(self.staff)
        resposta = self.client.get(reverse("admin:pedidos_analytics"), {"origem": "fora"})
        self.assertEqual(resposta.context["dados"]["total_pedidos"], 1)

    def test_filtro_origem_invalido_e_ignorado(self):
        self.client.force_login(self.staff)
        resposta = self.client.get(reverse("admin:pedidos_analytics"), {"origem": "lixo"})
        self.assertEqual(resposta.context["dados"]["total_pedidos"], 1)
        self.assertEqual(resposta.context["filtro_origem"], "")

    def test_exportar_csv_respeita_filtro_de_status(self):
        Pedido.objects.create(
            order_id="ORD-CANCELADO", conversion_id="2", usuario=self.usuario_comum,
            status=Pedido.STATUS_CANCELADO, status_shopee_bruto="CANCELLED",
            valor_comissao=Decimal("0"), valor_cashback=Decimal("0"), data_compra=timezone.now(),
        )
        self.client.force_login(self.staff)
        resposta = self.client.get(reverse("admin:pedidos_analytics_exportar"), {"status": Pedido.STATUS_VALIDADO})
        conteudo = resposta.content.decode()
        self.assertIn("ORD-1", conteudo)
        self.assertNotIn("ORD-CANCELADO", conteudo)

    def test_exportar_excel_traz_resumo_e_pedidos_formatados(self):
        self.client.force_login(self.staff)
        resposta = self.client.get(reverse("admin:pedidos_analytics_exportar_excel"))

        self.assertEqual(resposta.status_code, 200)
        self.assertEqual(
            resposta["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        livro = load_workbook(BytesIO(resposta.content))
        self.assertEqual(livro.sheetnames, ["Resumo", "Pedidos"])

        resumo = livro["Resumo"]
        self.assertEqual(resumo["A1"].value, "Analytics — cash-b")

        aba_pedidos = livro["Pedidos"]
        cabecalho = [celula.value for celula in aba_pedidos[1]]
        self.assertEqual(cabecalho[0], "Order ID")
        linhas = list(aba_pedidos.iter_rows(min_row=2, values_only=True))
        self.assertEqual(len(linhas), 1)
        self.assertEqual(linhas[0][0], "ORD-1")
        self.assertEqual(linhas[0][2], "Fora do site")
        self.assertEqual(linhas[0][5], Decimal("10.00"))
        celula_comissao = aba_pedidos.cell(row=2, column=6)
        self.assertEqual(celula_comissao.number_format, '"R$" #,##0.00')

    def test_exportar_excel_usuario_comum_e_redirecionado(self):
        self.client.force_login(self.usuario_comum)
        resposta = self.client.get(reverse("admin:pedidos_analytics_exportar_excel"))
        self.assertEqual(resposta.status_code, 302)
